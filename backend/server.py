"""Grupo RYSA ERP - API principal (FastAPI + PostgreSQL)."""
import os
import json
from dotenv import load_dotenv
from pathlib import Path

_ENV_BASE = Path(__file__).parent
# ENVIRONMENT definido por el entorno del sistema (p. ej. systemd en el VPS).
_env_active_os = os.environ.get("ENVIRONMENT", "").lower()
# Cargar siempre `.env` (base/local). Las variables del SO tienen prioridad.
load_dotenv(_ENV_BASE / '.env', override=False)
# Si el proceso indica el entorno explícitamente, cargar también `.env.<entorno>`.
if _env_active_os in ("development", "production"):
    _env_file = _ENV_BASE / f".env.{_env_active_os}"
    if _env_file.exists():
        load_dotenv(_env_file, override=True)

import io
import uuid
import re
import time
import logging
import mimetypes
import jwt
from typing import List, Optional
from fastapi import FastAPI, APIRouter, HTTPException, Depends, Response, UploadFile, File, Form, Request
from fastapi.responses import StreamingResponse, JSONResponse
from starlette.middleware.cors import CORSMiddleware
from starlette.concurrency import run_in_threadpool
from pydantic import BaseModel, Field, EmailStr, ConfigDict
import pandas as pd
from datetime import datetime, date, timedelta
import httpx
import platform

from deps import (
    db, client, now_utc, iso_now, hash_password, verify_password,
    create_access_token, create_refresh_token, decode_token, hash_token,
    get_current_user, require_permission, has_permission, effective_permissions,
    user_has_permission, next_counter, log_audit, revoke_user_sessions, MODULES,
    ROLE_PERMISSIONS, ACCESS_TOKEN_TTL_SECONDS, REFRESH_TOKEN_TTL_SECONDS,
    ver_todas_ventas, DEFAULT_MODULES_NON_PRIVILEGED,
    es_rol_privilegiado,
)
import pgstore.pos as _pgpos
import pgstore.cxc as _pgcxc
import pgstore.compras as _pgcompras
import storage
import exports
import ocr_invoice as _ocr_invoice
import pac_provider
import moneycalc
import developer as _devmod

_APP_ENV = os.environ.get("ENVIRONMENT", "development").lower()
logging.basicConfig(level=logging.DEBUG if _APP_ENV == "development" else logging.INFO)
logger = logging.getLogger("rysa")

app = FastAPI(title="Grupo RYSA ERP")
api = APIRouter(prefix="/api")

# Roles que se consideran administración del sistema (admin, propietario, desarrollador)
ADMIN_SYSTEM_ROLES = {"admin", "admin_propietario", "admin_desarrollador"}

# Unidades de medida predeterminadas del sistema. Configurables en
# Configuración → Precios y unidades; se ofrecen al crear/editar productos.
UNIDADES_DEFAULT = [
    "PZA", "CAJA", "PAQUETE", "BOLSA", "SIX", "CUBETA", "PAR", "JUEGO",
    "KG", "GR", "LT", "ML", "MT", "ROL", "SERVICIO",
]

# --- Bitácora en memoria de errores no controlados (para developer admin) ---
DEV_ERRORS = []

@app.exception_handler(Exception)
async def unhandled_error_handler(request: Request, exc: Exception):
    import traceback
    # El detalle completo solo se registra internamente; el cliente recibe un
    # mensaje genérico. En producción no se retiene la traza en memoria.
    if _APP_ENV != "production":
        tb = "".join(traceback.format_exception(type(exc), exc, exc.__traceback__))
        tipo = type(exc).__name__
        record = {
            "id": uid(), "fecha": iso_now(),
            "ruta": f"{request.method} {request.url.path}",
            "tipo": tipo, "mensaje": str(exc)[:800], "detalle": tb[-2000:],
            "estado": 500,
            "categoria": ("postgresql" if ("asyncpg" in tb or "DBAPI" in tipo
                                           or "asyncpg" in tipo) else "app"),
            "usuario": _usuario_de_request(request),
        }
        DEV_ERRORS.append(record)
        DEV_ERRORS[:] = DEV_ERRORS[-200:]
    logger.exception("Error no controlado: %s", exc)
    return JSONResponse(status_code=500, content={"detail": "Error interno del servidor"})


def _usuario_de_request(request: Request) -> str:
    """Extrae el email del usuario del token (best-effort, solo para bitácora dev)."""
    try:
        token = request.cookies.get("access_token")
        if not token:
            auth = request.headers.get("Authorization", "")
            token = auth[7:] if auth.startswith("Bearer ") else None
        if not token:
            return ""
        payload = decode_token(token)
        return payload.get("email") or payload.get("sub") or ""
    except Exception:
        return ""


# --- Bitácora de requests fallidos (>=400) para el módulo DESARROLLADOR ------
@app.middleware("http")
async def _dev_failed_request_logger(request: Request, call_next):
    t0 = time.perf_counter()
    response = await call_next(request)
    try:
        if response.status_code >= 400 and request.url.path.startswith("/api"):
            _devmod.record_request_error({
                "id": uid(),
                "fecha": iso_now(),
                "metodo": request.method,
                "ruta": request.url.path,
                "estado": response.status_code,
                "duracion_ms": round((time.perf_counter() - t0) * 1000, 1),
                "usuario": _usuario_de_request(request),
                "categoria": "http",
            })
    except Exception:
        pass
    return response

def es_admin_sistema(user: dict) -> bool:
    """Solo admin/propietario/desarrollador pueden administrar usuarios y módulos."""
    return user_has_permission(user, "usuarios.admin")

def uid() -> str:
    return uuid.uuid4().hex

SEARCH_MAX_LENGTH = 100

def sanitize_search_term(q) -> str:
    """Normaliza y escapa un término de búsqueda para $regex.
    Escapa metacaracteres (búsqueda literal) y limita la longitud para evitar
    patrones arbitrariamente costosos en la base de datos."""
    if not q:
        return ""
    s = str(q).strip()
    if len(s) > SEARCH_MAX_LENGTH:
        s = s[:SEARCH_MAX_LENGTH]
    return re.escape(s)

# ==========================================================================
# ESTRUCTURA COMPLETA DE 85 COLUMNAS (nomenclatura DBF/XBase)
# ==========================================================================
COLS_85 = [
    ("POSICION", "C"), ("CODIGO", "C"), ("DESCRIP", "C"), ("DESCRIPLRG", "M"), ("CLASIFICA", "C"),
    ("CATEGORIA", "C"), ("CATEGOCVE", "C"), ("DEPTOCVE", "C"), ("LINEA", "C"), ("UNIMEDIDA", "C"),
    ("UNIMEDCVE", "C"), ("CVEPROSER", "C"), ("SATOBJIMP", "C"), ("UBICACION", "C"), ("EMPAQUE", "N"),
    ("UNIMEDEMPQ", "C"), ("EXISTENCIA", "N"), ("INSUMO", "L"), ("PROVEEDOR", "C"), ("FECHAALTA", "D"),
    ("ULTFCOSTO", "D"), ("ULTCOSTO", "N"), ("COSTO", "N"), ("COSTODLLS", "N"), ("UTILMINIMO", "N"),
    ("UTILPRECI1", "N"), ("UTILPRECI2", "N"), ("UTILPRECI3", "N"), ("UTILPRECI4", "N"), ("UTILPRECI5", "N"),
    ("EXENTO", "L"), ("IMPUESTO", "N"), ("T_IEPS", "N"), ("IEPS", "N"), ("ISH", "N"),
    ("RET_ISR", "N"), ("RET_IVA", "N"), ("PRECIOVTA", "N"), ("PRECVTACTR", "N"), ("PRECVTAUSO", "N"),
    ("PRECIO1", "N"), ("PRECIO2", "N"), ("PRECIO3", "N"), ("PRECIO4", "N"), ("PRECIO5", "N"),
    ("PRECIOMIN", "N"), ("ULTFDEVCOM", "D"), ("ULTCDEVCOM", "N"), ("ULTFCOMPRA", "D"), ("ULTCCOMPRA", "N"),
    ("ULTFDEVVEN", "D"), ("ULTCDEVVEN", "N"), ("ULTFVENTA", "D"), ("ULTCVENTA", "N"), ("VTA_MES", "N"),
    ("VTA_ANUAL", "N"), ("XENTREGAR", "N"), ("XRECIBIR", "N"), ("STOCKMIN", "N"), ("STOCKMAX", "N"),
    ("PORPEDIR", "L"), ("IMAGEN", "M"), ("FOTO", "M"), ("FICHATEC", "M"), ("NUMSERIES", "L"),
    ("FACTCOMENT", "L"), ("INTEGRADO", "L"), ("VALEXIST", "L"), ("MODIPRECIO", "L"), ("APLIDESCTO", "L"),
    ("TOPECOSTO", "L"), ("INVENTARIO", "L"), ("MOVKARDEX", "L"), ("VENTAWEB", "L"), ("LOTES", "L"),
    ("CONTROLADO", "L"), ("BASCULA", "L"), ("ASOCIADO", "L"), ("FLETE", "L"), ("COMENTARIO", "M"),
    ("ROTACION", "C"), ("ULTPRECIO", "D"), ("COMISION", "N"), ("COMITIPO", "C"), ("STATUS", "C"),
]
COL_ORDER = [n for n, _ in COLS_85]
IMPORT_ALIASES = {"DESCRIPCION": "DESCRIP", "UNIDAD_MEDIDA": "UNIMEDIDA", "STOCK_MINIMO": "STOCKMIN",
                  "ESTADO": "STATUS", "CLASIFICACION": "CLASIFICA"}
STATUS_TO_ESTADO = {"A": "activo", "1": "activo", "ACTIVO": "activo", "B": "baja", "BAJA": "baja",
                    "S": "suspendido", "SUSPENDIDO": "suspendido"}

def _p_num(v):
    if v is None:
        return None
    s = str(v).strip().replace(",", "")
    if s == "" or s.lower() in ("nan", "none"):
        return None
    try:
        return float(s)
    except Exception:
        return "__ERR__"

def _p_bool(v):
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ("true", "1", "1.0", "si", "sí", "s", "yes", "y", "verdadero", ".t.", "t"):
        return True
    if s in ("false", "0", "0.0", "no", "n", "", ".f.", "f", "nan"):
        return False
    return "__ERR__"

def _p_date(v):
    if v is None:
        return None
    s = str(v).strip()
    if s == "" or s.lower() in ("nat", "nan", "none"):
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except Exception:
            pass
    try:
        return pd.to_datetime(s).date().isoformat()
    except Exception:
        return "__ERR__"

def asegurar_codigo_como_barras(doc: dict) -> dict:
    """Añade el código del producto como código de barras si aún no está registrado."""
    codigo = str(doc.get("codigo") or "").strip()
    barras = [str(x).strip() for x in doc.get("codigos_barras") or [] if str(x).strip()]
    if codigo and codigo not in barras:
        barras.insert(0, codigo)
    doc["codigos_barras"] = barras
    return doc

def parse_row(canon: dict):
    """canon: {COLNAME_UPPER: raw}. Devuelve (data lowercase, errores)."""
    data, errores = {}, []
    for name, t in COLS_85:
        raw = canon.get(name, "")
        key = name.lower()
        if t in ("C", "M"):
            data[key] = "" if raw is None else str(raw).strip()
        elif t == "N":
            val = _p_num(raw)
            if val == "__ERR__":
                errores.append({"campo": name, "valor": str(raw), "motivo": "Número inválido"}); data[key] = None
            else:
                data[key] = val
        elif t == "D":
            val = _p_date(raw)
            if val == "__ERR__":
                errores.append({"campo": name, "valor": str(raw), "motivo": "Fecha inválida"}); data[key] = None
            else:
                data[key] = val
        elif t == "L":
            val = _p_bool(raw)
            if val == "__ERR__":
                errores.append({"campo": name, "valor": str(raw), "motivo": "Booleano inválido"}); data[key] = False
            else:
                data[key] = val
    if not data.get("codigo"):
        errores.append({"campo": "CODIGO", "valor": "", "motivo": "Código obligatorio"})
    if not data.get("descrip"):
        errores.append({"campo": "DESCRIP", "valor": "", "motivo": "Descripción obligatoria"})
    return data, errores

def build_product_doc(d: dict) -> dict:
    """Documento de producto: conserva los 85 campos y sincroniza los campos usados por POS/Inventario."""
    iva = d.get("impuesto")
    iva = float(iva) if iva not in (None, 0, "") else 8.0
    costo = float(d.get("costo") or 0)
    precios = []
    for i in range(1, 6):
        con = d.get(f"precio{i}"); util = d.get(f"utilpreci{i}")
        if con:
            con = float(con); sin = round(con / (1 + iva / 100), 2)
            u = round((sin / costo - 1) * 100, 2) if costo else float(util or 0)
        else:
            u = float(util or 0); sin = round(costo * (1 + u / 100), 2); con = round(sin * (1 + iva / 100), 2)
        precios.append({"nombre": f"Precio {i}", "utilidad_pct": u, "precio_sin_iva": sin, "precio_con_iva": round(con, 2)})
    status = str(d.get("status", "")).upper()
    doc = dict(d)  # conserva los 85 campos tal cual
    doc.update({
        "descripcion": d.get("descrip", ""),
        "descripcion_larga": d.get("descriplrg", ""),
        "linea": d.get("linea", ""),
        "clasificacion": d.get("clasifica", ""),
        "unidad_medida": d.get("unimedida") or "PZA",
        "empaque": d.get("empaque") or "",
        "ubicacion": d.get("ubicacion", ""),
        "costo": costo,
        "stock_minimo": float(d.get("stockmin") or 0),
        "iva_tasa": iva,
        "estado": STATUS_TO_ESTADO.get(status, "activo"),
        "precios": precios,
        "precio_minimo": float(d.get("preciomin") or 0),
        "imagen_url": d.get("imagen") or d.get("foto") or "",
        "sku": d.get("codigo", ""),
        "sinonimos": [],
        "sat": {"clave_sat": d.get("cveproser", ""), "unidad_sat": d.get("unimedcve", ""),
                "impuestos": "Exento" if d.get("exento") else "IVA"},
        "controles": {"permitir_venta": True, "controlar_inventario": bool(d.get("inventario", True)),
                      "permitir_inventario_negativo": False,
                      "mostrar_pos": not bool(d.get("insumo", False)),
                      "mostrar_catalogo": bool(d.get("ventaweb", False))},
        "ficha_tecnica": {},
        "proveedores": [d.get("proveedor")] if d.get("proveedor") else [],
        "precio_incluye_iva": True,
    })
    _enriquecer_precios(doc)
    return doc

# =========================================================================
# MODELOS
# =========================================================================
class LoginInput(BaseModel):
    email: str = Field(max_length=320)
    password: str = Field(max_length=128)  # límite anti-DoS contra bcrypt

class UserCreate(BaseModel):
    email: EmailStr
    name: str
    password: str
    role: str = "vendedor"
    modulos: List[str] = Field(default_factory=list)
    sucursal_id: Optional[str] = None

class UserUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    active: Optional[bool] = None
    password: Optional[str] = None
    modulos: Optional[List[str]] = None
    sucursal_id: Optional[str] = None

class PrecioItem(BaseModel):
    nombre: str = "Precio 1"
    utilidad_pct: float = 0.0
    precio_sin_iva: float = 0.0
    precio_con_iva: float = 0.0

class ProductInput(BaseModel):
    # Mass assignment seguro: SOLO se aceptan los campos declarados.
    # Se declaran los 85 campos legacy DBF (COLS_85) + campos ERP modernos.
    model_config = ConfigDict(extra="forbid")

    # --- Campos ERP modernos ---
    codigo: Optional[str] = None
    sku: Optional[str] = ""
    descripcion: str
    descripcion_larga: Optional[str] = ""
    estado: str = "activo"  # activo | baja | suspendido
    linea: Optional[str] = ""
    clasificacion: Optional[str] = ""
    unidad_medida: str = "PZA"
    empaque: Optional[str] = ""
    costo: float = 0.0
    existencia: float = 0.0
    ubicacion: Optional[str] = ""
    stock_minimo: float = 0.0
    iva_tasa: float = 8.0
    # ¿El precio capturado en `precios` incluye IVA? True = los PRECIOn son
    # brutos (con IVA); False = son netos (sin IVA). Se deriva neto/bruto.
    precio_incluye_iva: bool = True
    precios: List[PrecioItem] = Field(default_factory=list)
    precio_minimo: float = 0.0
    sat: dict = Field(default_factory=dict)
    controles: dict = Field(default_factory=dict)
    ficha_tecnica: dict = Field(default_factory=dict)
    proveedores: List[str] = Field(default_factory=list)
    sinonimos: List[str] = Field(default_factory=list)
    imagen_url: Optional[str] = ""
    codigos_barras: List[str] = Field(default_factory=list)

    # --- Campos legacy DBF (nomenclatura XBase) ---
    # Los numéricos se aceptan como texto porque el formulario envía "" en
    # blanco y así se evita romper la validación manteniendo la persistencia.
    posicion: Optional[str] = ""
    descrip: Optional[str] = ""
    descriplrg: Optional[str] = ""
    clasifica: Optional[str] = ""
    categoria: Optional[str] = ""
    categocve: Optional[str] = ""
    deptocve: Optional[str] = ""
    unimedida: Optional[str] = ""
    unimedcve: Optional[str] = ""
    cveproser: Optional[str] = ""
    satobjimp: Optional[str] = ""
    unimedempq: Optional[str] = ""
    insumo: Optional[bool] = False
    proveedor: Optional[str] = ""
    fechaalta: Optional[str] = ""
    ultfcosto: Optional[str] = ""
    ultcosto: Optional[str] = ""
    costodlls: Optional[str] = ""
    utilminimo: Optional[str] = ""
    utilpreci1: Optional[str] = ""
    utilpreci2: Optional[str] = ""
    utilpreci3: Optional[str] = ""
    utilpreci4: Optional[str] = ""
    utilpreci5: Optional[str] = ""
    exento: Optional[bool] = False
    impuesto: Optional[str] = ""
    t_ieps: Optional[str] = ""
    ieps: Optional[str] = ""
    ish: Optional[str] = ""
    ret_isr: Optional[str] = ""
    ret_iva: Optional[str] = ""
    preciovta: Optional[str] = ""
    precvtactr: Optional[str] = ""
    precvtauso: Optional[str] = ""
    precio1: Optional[str] = ""
    precio2: Optional[str] = ""
    precio3: Optional[str] = ""
    precio4: Optional[str] = ""
    precio5: Optional[str] = ""
    preciomin: Optional[str] = ""
    ultfdevcom: Optional[str] = ""
    ultcdevcom: Optional[str] = ""
    ultfcompra: Optional[str] = ""
    ultccompra: Optional[str] = ""
    ultfdevven: Optional[str] = ""
    ultcdevven: Optional[str] = ""
    ultfventa: Optional[str] = ""
    ultcventa: Optional[str] = ""
    vta_mes: Optional[str] = ""
    vta_anual: Optional[str] = ""
    xentregar: Optional[str] = ""
    xrecibir: Optional[str] = ""
    stockmin: Optional[str] = ""
    stockmax: Optional[str] = ""
    porpedir: Optional[bool] = False
    imagen: Optional[str] = ""
    foto: Optional[str] = ""
    fichatec: Optional[str] = ""
    numseries: Optional[bool] = False
    factcoment: Optional[bool] = False
    integrado: Optional[bool] = False
    valexist: Optional[bool] = False
    modiprecio: Optional[bool] = False
    aplidescto: Optional[bool] = False
    topecosto: Optional[bool] = False
    inventario: Optional[bool] = False
    movkardex: Optional[bool] = False
    ventaweb: Optional[bool] = False
    lotes: Optional[bool] = False
    controlado: Optional[bool] = False
    bascula: Optional[bool] = False
    asociado: Optional[bool] = False
    flete: Optional[bool] = False
    comentario: Optional[str] = ""
    rotacion: Optional[str] = ""
    ultprecio: Optional[str] = ""
    comision: Optional[str] = ""
    comitipo: Optional[str] = ""
    status: Optional[str] = ""

class InventoryAdjust(BaseModel):
    tipo: str  # entrada | salida | ajuste | merma | devolucion | correccion
    cantidad: float
    concepto: Optional[str] = ""
    documento: Optional[str] = ""
    costo: Optional[float] = 0.0
    motivo: Optional[str] = ""
    observaciones: Optional[str] = ""

class ClientInput(BaseModel):
    # --- General ---
    codigo: Optional[str] = None            # CLAVE (identificador único)
    nombre: str                             # NOMBRE
    razon_social: Optional[str] = ""
    status: Optional[str] = ""              # STATUS legacy (1 char)
    estado: str = "activo"                  # activo | suspendido | inactivo
    tipo: str = "publico"                   # clasificación moderna
    tipo_clave: Optional[str] = ""          # TIPO legacy (1 char)
    fecha_alta: Optional[str] = ""          # FECHAALTA
    contrasena: Optional[str] = ""          # CONTRASENA
    # --- Contacto ---
    representa: Optional[str] = ""          # REPRESENTA
    tel_oficina: Optional[str] = ""         # TELOFICINA
    tel_residencia: Optional[str] = ""      # TELRESIDEN
    tel_fax: Optional[str] = ""             # TEL_FAX
    telefono: Optional[str] = ""            # compat
    celular: Optional[str] = ""             # CELULAR
    whatsapp: Optional[str] = ""            # compat
    correo: Optional[str] = ""              # compat (1 correo)
    correos: Optional[str] = ""             # CORREOS (varios)
    # --- Dirección ---
    direccion: Optional[str] = ""           # DIRECCION
    calle: Optional[str] = ""               # compat
    numero_exterior: Optional[str] = ""     # NOEXTERIOR
    numero_interior: Optional[str] = ""     # NOINTERIOR
    colonia: Optional[str] = ""             # COLONIA
    ciudad_edo: Optional[str] = ""          # CIUDADEDO
    localidad: Optional[str] = ""           # LOCALIDAD
    municipio: Optional[str] = ""           # compat
    ciudad: Optional[str] = ""              # CIUDAD
    estado_geo: Optional[str] = ""          # ESTADO (geográfico)
    pais: Optional[str] = "México"          # PAIS
    cp: Optional[str] = ""                  # CODPOSTAL
    referencias: Optional[str] = ""         # REFERENCIA
    id_localidad: Optional[str] = ""
    id_colonia: Optional[str] = ""
    id_ciudad: Optional[str] = ""
    id_estado: Optional[str] = ""
    id_pais: Optional[str] = ""
    # --- Fiscales ---
    rfc: Optional[str] = ""                 # RFC
    reg_fiscal: Optional[str] = ""          # REGFISCAL
    uso_cfdi: Optional[str] = ""            # USOCFDI
    resfiscal: Optional[str] = ""           # RESFISCAL
    nregidtrib: Optional[str] = ""          # NREGIDTRIB
    # --- Comercial ---
    vendedor: Optional[str] = ""            # VENDEDOR (legacy, texto)
    vendedor_id: Optional[str] = ""         # id del usuario vendedor asignado
    almacen: Optional[str] = ""             # ALMACEN
    precio_venta: Optional[int] = 1         # PRECIOVTA (lista de precios)
    lista_precios: int = 1                  # compat POS
    condicion_pago: str = "contado"
    # --- Geolocalización (campo / mapa) ---
    latitud: Optional[float] = None
    longitud: Optional[float] = None
    proxima_visita: Optional[str] = ""      # fecha ISO próxima visita programada
    # --- Crédito ---
    credito_autorizado: bool = False        # CREDITO
    limite_credito: float = 0.0             # LIMCREDITO
    lim_descuento: Optional[float] = 0.0    # LIMDESCTO
    descuento_permanente: Optional[float] = 0.0  # % descuento fijo del cliente
    dias_credito: Optional[int] = 0         # DIASCREDIT
    saldo: Optional[float] = 0.0            # SALDO
    venta_credito: Optional[float] = 0.0    # VTACREDITO
    # --- Retenciones ---
    ret_isr: bool = False                   # RET_ISR
    ret_iva: bool = False                   # RET_IVA
    ret_isr_tasa: Optional[float] = 0.0     # RET_ISRTAS
    ret_iva_tasa: Optional[float] = 0.0     # RET_IVATAS
    # --- Estadísticas ---
    mensual: Optional[float] = 0.0          # MENSUAL
    anual: Optional[float] = 0.0            # ANUAL
    ult_fecha_compra: Optional[str] = ""    # ULTFCOMPRA
    ult_monto_compra: Optional[float] = 0.0 # ULTCCOMPRA
    # --- Otros ---
    comentario: Optional[str] = ""          # COMENTARIO
    ofertas: bool = False                   # OFERTAS

class CreditInput(BaseModel):
    credito_autorizado: bool
    limite_credito: float = 0.0

class QuickToggle(BaseModel):
    valor: bool

class AbonoInput(BaseModel):
    monto: float = Field(ge=0)
    metodo: str = "efectivo"  # efectivo | tarjeta | transferencia | deposito | otros
    referencia: Optional[str] = Field(default="", max_length=200)
    nota: Optional[str] = Field(default="", max_length=500)

# =========================================================================
# PROVEEDORES
# =========================================================================
class ProveedorInput(BaseModel):
    nombre: str  # nombre comercial
    razon_social: Optional[str] = ""
    rfc: Optional[str] = ""
    telefono: Optional[str] = ""
    email: Optional[str] = ""
    direccion: Optional[str] = ""
    cp: Optional[str] = ""
    ciudad: Optional[str] = ""
    estado: Optional[str] = ""
    contacto: Optional[str] = ""
    telefono_contacto: Optional[str] = ""
    email_contacto: Optional[str] = ""
    condiciones_pago: Optional[str] = ""
    dias_credito: int = 0
    limite_credito: float = 0.0
    banco: Optional[str] = ""
    cuenta: Optional[str] = ""
    clabe: Optional[str] = ""
    observaciones: Optional[str] = ""
    activo: bool = True
    categoria: Optional[str] = ""

# =========================================================================
# CUENTAS BANCARIAS
# =========================================================================
class CuentaBancariaInput(BaseModel):
    banco: str
    nombre: Optional[str] = ""
    numero_cuenta: str
    clabe: Optional[str] = ""
    titular: Optional[str] = ""
    moneda: str = "MXN"
    tipo_cuenta: Optional[str] = "debito"  # debito | credito | nomina | otros
    activa: bool = True
    alias: Optional[str] = ""
    predeterminada: bool = False

# =========================================================================
# COMPRAS / GASTOS
# =========================================================================
class CompraItemInput(BaseModel):
    product_id: Optional[str] = None
    codigo: str = ""
    descripcion: str = ""
    unidad: str = "PZA"
    cantidad: float = 1.0
    costo: float = 0.0  # costo unitario
    iva_tasa: float = 8.0
    descuento: float = 0.0
    afecta_inventario: bool = True  # ✓ Afecta inventario (compra) vs gasto
    importe: Optional[float] = None

class CompraInput(BaseModel):
    tipo: str = "compra"  # compra | gasto | mixto
    proveedor_id: Optional[str] = None
    proveedor_nombre: Optional[str] = ""
    factura_numero: Optional[str] = ""
    fecha_factura: Optional[str] = ""
    fecha_recepcion: Optional[str] = ""
    fecha_vencimiento: Optional[str] = ""
    concepto: Optional[str] = ""
    categoria: Optional[str] = ""
    subtotal: float = 0.0
    descuento: float = 0.0
    iva: float = 0.0
    otros_impuestos: float = 0.0
    total: float = 0.0
    metodo_pago: str = "efectivo"  # efectivo|transferencia|tarjeta|deposito|credito|otros
    forma_pago: Optional[str] = ""  # contado | credito
    cuenta_bancaria_id: Optional[str] = None
    observaciones: Optional[str] = ""
    items: List[CompraItemInput] = Field(default_factory=list)
    documentos: List[dict] = Field(default_factory=list)  # evidencia/factura
    # ---- Evolución Compras y Gastos ----
    orden_id: Optional[str] = None          # orden de compra relacionada
    recepcion_id: Optional[str] = None      # recepción de mercancía relacionada
    centro_costo_id: Optional[str] = None   # centro de costo del gasto
    centro_costo_nombre: Optional[str] = ""
    sucursal_id: Optional[str] = None       # sucursal destino
    # Costos adicionales de compra (flete, seguro, maniobras, transporte, otros)
    flete: float = 0.0
    seguro: float = 0.0
    maniobras: float = 0.0
    transporte: float = 0.0
    otros_costos: float = 0.0

# ---- ÓRDENES DE COMPRA ----
class OrdenItemInput(BaseModel):
    product_id: Optional[str] = None
    codigo: str = ""
    descripcion: str = ""
    unidad: str = "PZA"
    solicitado: float = 1.0
    costo: float = 0.0
    iva_tasa: float = 8.0

class OrdenCompraInput(BaseModel):
    proveedor_id: Optional[str] = None
    proveedor_nombre: Optional[str] = ""
    fecha_orden: Optional[str] = ""
    fecha_estimada: Optional[str] = ""
    estado: str = "borrador"   # borrador | enviada | cancelada
    notas: Optional[str] = ""
    items: List[OrdenItemInput] = Field(default_factory=list)
    sucursal_id: Optional[str] = None
    cuenta_bancaria_id: Optional[str] = None

class OrdenEstadoInput(BaseModel):
    estado: str  # borrador | enviada | cancelada

# ---- RECEPCIONES DE MERCANCÍA ----
class RecepcionItemInput(BaseModel):
    product_id: Optional[str] = None
    codigo: str = ""
    descripcion: str = ""
    unidad: str = "PZA"
    cantidad: float = 0.0      # cantidad realmente recibida
    costo: float = 0.0
    iva_tasa: float = 8.0

class RecepcionInput(BaseModel):
    orden_id: str
    fecha: Optional[str] = ""
    factura_numero: Optional[str] = ""
    fecha_factura: Optional[str] = ""
    metodo_pago: str = "efectivo"
    forma_pago: Optional[str] = ""   # contado | credito
    cuenta_bancaria_id: Optional[str] = None
    fecha_vencimiento: Optional[str] = ""
    observaciones: Optional[str] = ""
    items: List[RecepcionItemInput] = Field(default_factory=list)
    documentos: List[dict] = Field(default_factory=list)

# ---- PRESUPUESTOS ----
class PresupuestoInput(BaseModel):
    categoria: Optional[str] = ""
    sucursal_id: Optional[str] = None
    centro_costo_id: Optional[str] = None
    centro_costo_nombre: Optional[str] = ""
    periodo: str = ""          # YYYY-MM
    monto: float = 0.0
    notas: Optional[str] = ""

# ---- CENTROS DE COSTO ----
class CentroCostoInput(BaseModel):
    nombre: str
    codigo: Optional[str] = ""
    descripcion: Optional[str] = ""
    activo: bool = True

# ---- GASTOS / COMPRAS RECURRENTES ----
class RecurrenteInput(BaseModel):
    tipo: str = "gasto"        # gasto | compra
    proveedor_id: Optional[str] = None
    proveedor_nombre: Optional[str] = ""
    concepto: Optional[str] = ""
    categoria: Optional[str] = ""
    importe: float = 0.0
    frecuencia: str = "mensual"  # semanal | quincenal | mensual | bimestral | trimestral | anual
    dia: int = 1                # día del periodo en que se genera
    cuenta_bancaria_id: Optional[str] = None
    recordatorio: bool = True
    sucursal_id: Optional[str] = None
    centro_costo_id: Optional[str] = None
    centro_costo_nombre: Optional[str] = ""
    activo: bool = True
    notas: Optional[str] = ""

# ---- PAGO DE CUENTA POR PAGAR ----
class CompraPagoInput(BaseModel):
    monto: float = 0.0
    metodo_pago: str = "efectivo"
    cuenta_bancaria_id: Optional[str] = None
    referencia: Optional[str] = ""
    fecha: Optional[str] = ""
    notas: Optional[str] = ""


class CajaOpen(BaseModel):
    fondo_inicial: float = 0.0
    caja_nombre: Optional[str] = ""
    denominaciones: Optional[dict] = None   # {1000: 2, 500: 1, ...} (billetes y monedas)
    metodo: Optional[str] = "denominaciones"  # denominaciones | solo_monto

class CajaOpenPorUsuario(BaseModel):
    usuario_id: str
    fondo_inicial: float = 0.0
    caja_nombre: Optional[str] = ""
    denominaciones: Optional[dict] = None
    metodo: Optional[str] = "denominaciones"

class CajaMovimiento(BaseModel):
    tipo: str  # entrada | retiro | gasto | ajuste
    concepto: str
    monto: float
    referencia: Optional[str] = ""
    forzar: bool = False              # excedente autorizado (encargado/admin)
    evidencia_url: Optional[str] = "" # foto del comprobante (gasto/retiro)

class CajaClose(BaseModel):
    efectivo_contado: float
    caja_id: Optional[str] = None

class SaleItem(BaseModel):
    product_id: Optional[str] = None  # None para líneas sin inventario (p. ej. recargas)
    codigo: str = Field(max_length=100)
    descripcion: str = Field(max_length=500)
    cantidad: float = Field(ge=0)      # >= 0: 0 se ignora aguas abajo; nunca negativo
    unidad: str = "PZA"
    precio: float = Field(ge=0)
    iva_tasa: float = Field(default=8.0, ge=0, le=100)
    descuento: float = Field(default=0.0, ge=0)  # monto de descuento por linea
    comentario: str = Field(default="", max_length=300)
    costo: Optional[float] = None  # costo unitario tomado del producto (snapshot)

class Pago(BaseModel):
    metodo: str  # efectivo | tarjeta | transferencia | deposito | otros
    monto: float = Field(ge=0)
    card_type: Optional[str] = None  # debito | credito (solo cuando metodo == "tarjeta")

class SaleInput(BaseModel):
    cliente_id: Optional[str] = None
    items: List[SaleItem]
    descuento_global: float = 0.0
    condicion: str = "contado"  # contado | credito
    pagos: List[Pago] = Field(default_factory=list)
    lista_precios: int = 1
    tipo_venta: str = "directa"  # directa | cotizacion
    vendedor_id: Optional[str] = None
    # Convención de precios: True = los precios enviados ya incluyen IVA (brutos,
    # se extrae el neto); False = los precios son netos (se suma el IVA). El total
    # siempre es el bruto. Default True mantiene compatibilidad con clientes antiguos.
    precios_incluyen_iva: bool = True
    idempotency_key: Optional[str] = None  # evita ventas duplicadas en POS
    # Override de inventario negativo (solo roles autorizados, con auditoría).
    allow_negative_inventory: bool = False
    override_reason: Optional[str] = None

class CancelInput(BaseModel):
    motivo: str

# ---- COTIZACIONES (conversión a venta) ----
class CotizacionConvertInput(BaseModel):
    condicion: str = "contado"  # contado | credito
    pagos: List[Pago] = Field(default_factory=list)
    vendedor_id: Optional[str] = None

# ---- PEDIDOS ----
class PedidoItemInput(BaseModel):
    product_id: Optional[str] = None
    codigo: str = ""
    descripcion: str = ""
    unidad: str = "PZA"
    solicitado: float = 1.0
    precio: float = 0.0
    iva_tasa: float = 8.0

class PedidoInput(BaseModel):
    cliente_id: Optional[str] = None
    vendedor_id: Optional[str] = None
    fecha_pedido: Optional[str] = ""
    fecha_entrega: Optional[str] = ""
    notas: Optional[str] = ""
    items: List[PedidoItemInput] = Field(default_factory=list)
    estado: str = "borrador"  # borrador | confirmado | surtido | convertido | cancelado
    sucursal_id: Optional[str] = None

class PedidoEstadoInput(BaseModel):
    estado: str  # borrador | confirmado | surtido | cancelado

class PedidoConvertInput(BaseModel):
    condicion: str = "contado"
    pagos: List[Pago] = Field(default_factory=list)
    vendedor_id: Optional[str] = None

class RecargaInput(BaseModel):
    compania: str
    telefono: str
    monto: float
    metodo: str = "efectivo"
    referencia_tae: Optional[str] = ""
    comision: Optional[float] = 0.0

class SucursalItem(BaseModel):
    nombre: str = ""
    codigo: Optional[str] = ""
    direccion: Optional[str] = ""
    ciudad: Optional[str] = ""
    estado: Optional[str] = ""
    cp: Optional[str] = ""
    telefono: Optional[str] = ""
    activa: bool = True

class SettingsInput(BaseModel):
    empresa_nombre: str = "Grupo RYSA"
    razon_social: Optional[str] = ""
    rfc: Optional[str] = ""
    telefono: Optional[str] = ""
    correo: Optional[str] = ""
    direccion: Optional[str] = ""
    colonia: Optional[str] = ""
    ciudad: Optional[str] = ""
    estado: Optional[str] = ""
    cp: Optional[str] = ""
    pais: Optional[str] = "México"
    iva_tasa: float = 8.0
    moneda: str = "MXN"
    precios_incluyen_iva: bool = True
    listas_precios_nombres: List[str] = Field(default_factory=lambda: ["Precio 1", "Precio 2", "Precio 3", "Precio 4", "Precio 5"])
    listas_precios_pct: List[float] = Field(default_factory=lambda: [40, 30, 20, 15, 10])
    unidades_medida: List[str] = Field(default_factory=lambda: list(UNIDADES_DEFAULT))
    logo_url: Optional[str] = ""
    ticket_config: dict = Field(default_factory=dict)
    sucursales: List[SucursalItem] = Field(default_factory=list)
    storage: dict = Field(default_factory=dict)
    # Impresoras configuradas (lista) + predeterminadas por tipo de documento.
    printers: dict = Field(default_factory=dict)
    # Número de WhatsApp de la EMPRESA para recibir comprobantes (wa.me).
    whatsapp_empresa: Optional[str] = ""
    # QR de comprobante de pago en cotizaciones (§20): configurable, con defaults.
    qr_comprobante: dict = Field(default_factory=lambda: {
        "activo": True, "vigencia_dias": 30, "max_mb": 10, "max_archivos": 6})

# =========================================================================
# INVENTARIO (KARDEX) - helper
# =========================================================================
async def registrar_movimiento(product: dict, tipo: str, entrada: float, salida: float,
                                usuario: dict, documento: str = "", referencia: str = "",
                                costo: float = 0.0, motivo: str = "", observaciones: str = "",
                                venta_id: str = "", caja_id: str = ""):
    anterior = round(float(product.get("existencia", 0)), 3)
    nueva_existencia = round(anterior + entrada - salida, 3)
    await db.products.update_one({"id": product["id"]}, {"$set": {"existencia": nueva_existencia, "updated_at": iso_now()}})
    # Contador de unidades vendidas (para catálogo "más vendidos" en POS).
    # Una devolución/cancelación reduce el contador: las unidades devueltas no
    # deben seguir contando como vendidas.
    if tipo == "venta" and salida > 0:
        await db.products.update_one({"id": product["id"]}, {"$inc": {"vendidas": float(salida)}})
    elif tipo in ("devolucion", "cancelacion") and entrada > 0:
        await db.products.update_one({"id": product["id"]}, {"$inc": {"vendidas": -float(entrada)}})
    now = now_utc()
    await db.inventory_movements.insert_one({
        "id": uid(), "product_id": product["id"], "codigo": product.get("codigo"),
        "descripcion": product.get("descripcion"), "tipo": tipo,
        "documento": documento, "entrada": entrada, "salida": salida,
        "existencia_anterior": anterior, "existencia_resultante": nueva_existencia,
        "costo": round(float(costo or 0), 4), "motivo": motivo, "observaciones": observaciones,
        "usuario_id": usuario.get("id"), "usuario_nombre": usuario.get("name"),
        "referencia": referencia, "venta_id": venta_id or "", "caja_id": caja_id or "",
        "fecha": iso_now(),
        "hora": now.strftime("%H:%M:%S"),
    })
    return nueva_existencia

# =========================================================================
# AUTH
# =========================================================================
def public_user(u: dict) -> dict:
    u = dict(u)
    u.pop("password_hash", None)
    u.pop("_id", None)
    return u

ACCESS_COOKIE = "access_token"
REFRESH_COOKIE = "refresh_token"

# -------------------------------------------------------------------------
# Rate limiting de login persistente en PostgreSQL (compartido entre workers).
# Tabla dedicada `login_attempts`, no un diccionario en memoria.
# -------------------------------------------------------------------------
LOGIN_IP_WINDOW_SECONDS = 60            # ventana de conteo por IP
LOGIN_IP_MAX_FAILURES = 8               # máximo de fallos por minuto y por IP
LOGIN_USER_MAX_FAILURES = 8             # fallos acumulados para bloquear el usuario
LOGIN_USER_LOCK_MINUTES = 15            # duración del bloqueo temporal
_CLEANUP_LAST_RUN = [0.0]


def _iso_to_dt(value) -> Optional[datetime]:
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00"))
    except Exception:
        return None


def _cookie_secure() -> bool:
    return os.environ.get("ENVIRONMENT", "development").lower() == "production"


def set_auth_cookies(response: Response, access_token: str, refresh_token: str):
    secure = _cookie_secure()
    response.set_cookie(ACCESS_COOKIE, access_token, httponly=True, secure=secure,
                        samesite="lax", max_age=ACCESS_TOKEN_TTL_SECONDS, path="/")
    response.set_cookie(REFRESH_COOKIE, refresh_token, httponly=True, secure=secure,
                        samesite="lax", max_age=REFRESH_TOKEN_TTL_SECONDS, path="/")


def clear_auth_cookies(response: Response):
    response.delete_cookie(ACCESS_COOKIE, path="/")
    response.delete_cookie(REFRESH_COOKIE, path="/")


async def store_refresh_token(token: str, user_id: str, token_version: int):
    payload = decode_token(token)
    now = now_utc()
    await db.refresh_tokens.insert_one({
        "_id": payload["jti"],
        "user_id": user_id,
        "token_hash": hash_token(token),
        "token_version": int(token_version or 0),
        "created_at": iso_now(),
        "expires_at": now + timedelta(seconds=REFRESH_TOKEN_TTL_SECONDS),
        "active": True,
    })


async def _cleanup_login_attempts():
    now = time.time()
    if now - _CLEANUP_LAST_RUN[0] < 300:
        return
    _CLEANUP_LAST_RUN[0] = now
    try:
        cutoff_ip = (now_utc() - timedelta(minutes=30)).isoformat()
        cutoff_user = (now_utc() - timedelta(hours=24)).isoformat()
        await db.login_attempts.delete_many({"kind": "ip", "last_at": {"$lt": cutoff_ip}})
        await db.login_attempts.delete_many({"kind": "user", "last_at": {"$lt": cutoff_user}})
    except Exception:
        pass


async def check_login_rate_limit(ip: str, email: str):
    if os.environ.get("LOGIN_RATE_LIMIT", "on").lower() == "off":
        return
    now = now_utc()
    doc = await db.login_attempts.find_one({"_id": "ip:" + ip})
    if doc:
        ws = _iso_to_dt(doc.get("window_start"))
        if (ws and (now - ws) < timedelta(seconds=LOGIN_IP_WINDOW_SECONDS)
                and int(doc.get("count", 0)) >= LOGIN_IP_MAX_FAILURES):
            raise HTTPException(status_code=429,
                                detail="Demasiados intentos. Intenta de nuevo en un minuto.")
    if email:
        u = await db.login_attempts.find_one({"_id": "user:" + email.lower()})
        if u:
            lu = _iso_to_dt(u.get("locked_until"))
            if lu and lu > now:
                raise HTTPException(status_code=429,
                                    detail="Cuenta temporalmente bloqueada por demasiados intentos. Intenta en unos minutos.")


async def record_failed_login(ip: str, email: str):
    now = now_utc()
    ip_key = "ip:" + ip
    doc = await db.login_attempts.find_one({"_id": ip_key})
    if doc:
        ws = _iso_to_dt(doc.get("window_start"))
        if ws and (now - ws) >= timedelta(seconds=LOGIN_IP_WINDOW_SECONDS):
            await db.login_attempts.update_one({"_id": ip_key},
                                               {"$set": {"count": 1, "window_start": iso_now(), "last_at": iso_now()}})
        else:
            await db.login_attempts.update_one({"_id": ip_key},
                                               {"$inc": {"count": 1}, "$set": {"last_at": iso_now()}})
    else:
        await db.login_attempts.insert_one({"_id": ip_key, "kind": "ip", "count": 1,
                                            "window_start": iso_now(), "last_at": iso_now()})
    if email:
        ukey = "user:" + email.lower()
        u = await db.login_attempts.find_one({"_id": ukey})
        fails = int((u or {}).get("fails", 0)) + 1
        locked_until = None
        if fails >= LOGIN_USER_MAX_FAILURES:
            locked_until = (now + timedelta(minutes=LOGIN_USER_LOCK_MINUTES)).isoformat()
        await db.login_attempts.update_one({"_id": ukey},
                                           {"$set": {"kind": "user", "fails": fails,
                                                     "locked_until": locked_until, "last_at": iso_now()}},
                                           upsert=True)


async def reset_login_failures(email: str):
    if email:
        await db.login_attempts.delete_one({"_id": "user:" + email.lower()})


def _password_ok(password: str) -> bool:
    return bool(password) and len(password) >= 12


@api.post("/auth/login")
async def login(data: LoginInput, response: Response, request: Request):
    ip = request.client.host if request.client else "unknown"
    ua = request.headers.get("user-agent", "")
    email = data.email.strip().lower()
    await _cleanup_login_attempts()
    await check_login_rate_limit(ip, email)

    user = await db.users.find_one({"email": email})
    if not user or not verify_password(data.password, user["password_hash"]):
        await record_failed_login(ip, email)
        await log_audit({"id": None, "name": email or "?"}, "LOGIN_FAILURE", "auth",
                        detalle=(email or "sin correo"), ip=ip, user_agent=ua)
        raise HTTPException(status_code=401, detail="Credenciales incorrectas")
    if not user.get("active", True):
        await log_audit({"id": user.get("id"), "name": user.get("name")}, "LOGIN_FAILURE", "auth",
                        registro_id=user.get("id"), detalle=f"Usuario desactivado", ip=ip, user_agent=ua)
        raise HTTPException(status_code=403, detail="Usuario desactivado")

    await reset_login_failures(email)
    token_version = int(user.get("token_version", 0))
    access = create_access_token(user["id"], user["email"], token_version)
    refresh = create_refresh_token(user["id"], token_version)
    await store_refresh_token(refresh, user["id"], token_version)
    set_auth_cookies(response, access, refresh)
    await log_audit(user, "LOGIN_SUCCESS", "auth", registro_id=user["id"],
                    detalle=f"IP {ip}", ip=ip, user_agent=ua)
    return {"token": access, "user": public_user(user)}


@api.post("/auth/refresh")
async def refresh_session(response: Response, request: Request):
    rt = request.cookies.get(REFRESH_COOKIE)
    if not rt:
        clear_auth_cookies(response)
        raise HTTPException(status_code=401, detail="Sesión expirada")
    try:
        payload = decode_token(rt)
    except jwt.ExpiredSignatureError:
        clear_auth_cookies(response)
        raise HTTPException(status_code=401, detail="Sesión expirada")
    except Exception:
        clear_auth_cookies(response)
        raise HTTPException(status_code=401, detail="Token inválido")
    if payload.get("type") != "refresh":
        clear_auth_cookies(response)
        raise HTTPException(status_code=401, detail="Token inválido")

    jti = payload.get("jti")
    record = await db.refresh_tokens.find_one({"_id": jti, "active": True})
    if not record:
        clear_auth_cookies(response)
        raise HTTPException(status_code=401, detail="Sesión expirada")

    user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
    if not user or not user.get("active", True):
        clear_auth_cookies(response)
        raise HTTPException(status_code=401, detail="Sesión expirada")
    token_version = int(user.get("token_version", 0))
    if token_version != int(payload.get("token_version", 0)):
        await db.refresh_tokens.update_one({"_id": jti}, {"$set": {"active": False, "revoked_at": iso_now()}})
        clear_auth_cookies(response)
        raise HTTPException(status_code=401, detail="Sesión invalidada")

    # Rotación: se revoca el refresh actual y se emite uno nuevo (nunca en localStorage).
    await db.refresh_tokens.update_one({"_id": jti}, {"$set": {"active": False, "rotated_at": iso_now()}})
    access = create_access_token(user["id"], user["email"], token_version)
    new_refresh = create_refresh_token(user["id"], token_version)
    await store_refresh_token(new_refresh, user["id"], token_version)
    set_auth_cookies(response, access, new_refresh)
    return {"token": access}


@api.post("/auth/logout")
async def logout(response: Response, request: Request):
    ip = request.client.host if request.client else "unknown"
    ua = request.headers.get("user-agent", "")
    usuario = None
    rt = request.cookies.get(REFRESH_COOKIE)
    if rt:
        try:
            payload = decode_token(rt)
            if payload.get("type") == "refresh":
                await db.refresh_tokens.update_one(
                    {"_id": payload.get("jti"), "active": True},
                    {"$set": {"active": False, "revoked_at": iso_now()}})
                user_doc = await db.users.find_one({"id": payload.get("sub")})
                if user_doc:
                    usuario = user_doc
        except Exception:
            pass
    clear_auth_cookies(response)
    if usuario:
        await log_audit(usuario, "LOGOUT", "auth", registro_id=usuario["id"], ip=ip, user_agent=ua)
    return {"ok": True}

@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return {"user": user, "permissions": sorted(effective_permissions(user))}

# =========================================================================
# USUARIOS
# =========================================================================
def _validar_modulos(modulos: List[str]):
    desconocidos = [m for m in modulos if m not in MODULES]
    if desconocidos:
        raise HTTPException(400, f"Módulos desconocidos: {', '.join(desconocidos)}")

@api.get("/users")
async def list_users(user: dict = Depends(require_permission("usuarios.ver"))):
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(500)
    return users

@api.post("/users")
async def create_user(data: UserCreate, user: dict = Depends(require_permission("usuarios.crear"))):
    _validar_modulos(data.modulos)
    if not _password_ok(data.password):
        raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 12 caracteres")
    if data.role in ADMIN_SYSTEM_ROLES or data.modulos:
        if not es_admin_sistema(user):
            raise HTTPException(403, "Solo administradores/propietario pueden asignar roles privilegiados o módulos")
    email = data.email.strip().lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="El correo ya está registrado")
    # Módulos: si el rol no es privilegiado y no se especifican, se otorgan los
    # módulos por defecto (Productos, Inventario, Clientes, Recargas, Ventas,
    # Caja, Reportes). Solo un admin/propietario puede especificarlos a mano.
    modulos = data.modulos
    if not es_rol_privilegiado(data.role) and not modulos:
        modulos = DEFAULT_MODULES_NON_PRIVILEGED
    sucursal_id = data.sucursal_id or await _default_sucursal_id()
    doc = {"id": uid(), "email": email, "name": data.name, "role": data.role,
           "modulos": modulos, "password_hash": hash_password(data.password),
           "sucursal_id": sucursal_id, "active": True, "token_version": 0, "created_at": iso_now()}
    await db.users.insert_one(doc)
    await log_audit(user, "crear", "usuario", doc["id"],
                    f"Usuario {email} · rol {data.role} · {'+'.join(modulos) or 'sin módulos'}")
    return public_user(doc)

@api.put("/users/{user_id}")
async def update_user(user_id: str, data: UserUpdate, user: dict = Depends(require_permission("usuarios.editar"))):
    target = await db.users.find_one({"id": user_id})
    if not target:
        raise HTTPException(404, "Usuario no encontrado")
    rol_privilegiado = target.get("role") in ADMIN_SYSTEM_ROLES
    cambia_a_admin = data.role in ADMIN_SYSTEM_ROLES if data.role else False
    toca_modulos = data.modulos is not None
    if rol_privilegiado or cambia_a_admin or toca_modulos:
        if not es_admin_sistema(user):
            raise HTTPException(403, "Solo administradores/propietario pueden editar administradores o asignar módulos")
    upd = {k: v for k, v in data.model_dump().items() if v is not None}
    if "modulos" in upd:
        _validar_modulos(upd["modulos"])
    cambia_password = "password" in upd
    if cambia_password:
        if not _password_ok(upd["password"]):
            raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 12 caracteres")
        upd["password_hash"] = hash_password(upd.pop("password"))
    cambia_rol = "role" in upd and upd["role"] != target.get("role")
    desactiva = "active" in upd and upd["active"] is False and target.get("active", True)
    await db.users.update_one({"id": user_id}, {"$set": upd})
    if cambia_password or cambia_rol or desactiva:
        # Invalidar todas las sesiones anteriores del usuario objetivo.
        await revoke_user_sessions(user_id)
        if cambia_password:
            await log_audit(user, "PASSWORD_CHANGE", "usuario", user_id)
        if desactiva:
            await log_audit(user, "ACCOUNT_DISABLED", "usuario", user_id)
        await log_audit(user, "TOKEN_REVOKED", "usuario", user_id,
                        "Sesiones invalidadas por cambio de password/rol o desactivación")
    await log_audit(user, "editar", "usuario", user_id, f"campos: {', '.join(sorted(upd))}")
    return await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})

@api.get("/roles")
async def list_roles(user: dict = Depends(get_current_user)):
    return {
        "roles": {r: sorted(p) for r, p in ROLE_PERMISSIONS.items()},
        "modulos": {k: v["label"] for k, v in MODULES.items()},
    }

# =========================================================================
# PRODUCTOS
# =========================================================================
def calc_precios(costo: float, precios: List[dict], iva_tasa: float) -> List[dict]:
    out = []
    for p in precios:
        util = float(p.get("utilidad_pct", 0) or 0)
        if p.get("precio_sin_iva"):
            sin_iva = float(p["precio_sin_iva"])
            con_iva = round(sin_iva * (1 + iva_tasa / 100), 2)
        elif p.get("precio_con_iva"):
            con_iva = float(p["precio_con_iva"])
            sin_iva = round(con_iva / (1 + iva_tasa / 100), 2)
            util = round((sin_iva / costo - 1) * 100, 2) if costo else util
        else:
            sin_iva = round(costo * (1 + util / 100), 2)
            con_iva = round(sin_iva * (1 + iva_tasa / 100), 2)
        out.append({"nombre": p.get("nombre", "Precio"), "utilidad_pct": util,
                    "precio_sin_iva": sin_iva, "precio_con_iva": round(con_iva, 2)})
    return out

@api.get("/products")
async def list_products(response: Response, estado: Optional[str] = None, q: Optional[str] = None,
                        filtro: Optional[str] = None, categoria: Optional[str] = None,
                        sku: Optional[str] = None, linea: Optional[str] = None,
                        unidad_medida: Optional[str] = None, proveedor: Optional[str] = None,
                        min_costo: Optional[float] = None, max_costo: Optional[float] = None,
                        min_precio: Optional[float] = None, max_precio: Optional[float] = None,
                        skip: int = 0, limit: int = 100,
                        user: dict = Depends(get_current_user)):
    limit = max(1, min(int(limit), 500))
    skip = max(0, int(skip))
    query = _product_export_query(estado=estado, q=q, categoria=categoria, sku=sku,
                                  linea=linea, unidad_medida=unidad_medida, proveedor=proveedor,
                                  min_costo=min_costo, max_costo=max_costo,
                                  min_precio=min_precio, max_precio=max_precio)
    if filtro in ("bajo_stock", "sin_existencia"):
        docs = await db.products.find(query, {"_id": 0}).sort("descripcion", 1).to_list(20000)
        if filtro == "bajo_stock":
            docs = [p for p in docs if 0 < float(p.get("existencia", 0)) <= float(p.get("stock_minimo", 0))]
        else:
            docs = [p for p in docs if float(p.get("existencia", 0)) <= 0]
        total = len(docs)
        docs = docs[skip:skip + limit]
    else:
        total = await db.products.count_documents(query)
        docs = await db.products.find(query, {"_id": 0}).sort("descripcion", 1).skip(skip).limit(limit).to_list(limit)
    response.headers["X-Total-Count"] = str(total)
    return docs

# --- Catálogo POS: más vendidos y favoritos por usuario ---
@api.get("/products/bestsellers")
async def list_bestsellers(estado: Optional[str] = None, limit: int = 24,
                           user: dict = Depends(get_current_user)):
    limit = max(1, min(int(limit), 100))
    query = {"vendidas": {"$gt": 0}}
    if estado:
        query["estado"] = estado
    docs = await db.products.find(query, {"_id": 0}).sort("vendidas", -1).limit(limit).to_list(limit)
    return docs

@api.get("/favorites")
async def list_favorites(user: dict = Depends(get_current_user)):
    favs = await db.favorites.find({"user_id": user["id"]}, {"_id": 0, "product_id": 1}).to_list(5000)
    ids = [f["product_id"] for f in favs]
    if not ids:
        return []
    products = await db.products.find({"id": {"$in": ids}, "estado": "activo"}, {"_id": 0}).to_list(5000)
    by_id = {p["id"]: p for p in products}
    return [by_id[i] for i in ids if i in by_id]

@api.post("/favorites/{product_id}")
async def add_favorite(product_id: str, user: dict = Depends(get_current_user)):
    p = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not p:
        raise HTTPException(404, "Producto no encontrado")
    exists = await db.favorites.find_one({"user_id": user["id"], "product_id": product_id})
    if not exists:
        await db.favorites.insert_one({"id": uid(), "user_id": user["id"], "product_id": product_id,
                                       "fecha": iso_now()})
    return {"ok": True}

@api.delete("/favorites/{product_id}")
async def remove_favorite(product_id: str, user: dict = Depends(get_current_user)):
    await db.favorites.delete_one({"user_id": user["id"], "product_id": product_id})
    return {"ok": True}

@api.get("/products/{product_id}")
async def get_product(product_id: str, user: dict = Depends(get_current_user)):
    p = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not p:
        raise HTTPException(404, "Producto no encontrado")
    return p

@api.post("/products")
async def create_product(data: ProductInput, user: dict = Depends(require_permission("producto.crear"))):
    codigo = (data.codigo or "").strip() or await next_counter("producto", "P", 5)
    if await db.products.find_one({"codigo": codigo}):
        raise HTTPException(400, "El código ya existe")
    doc = data.model_dump()
    doc["codigo"] = codigo
    doc["id"] = uid()
    doc = asegurar_codigo_como_barras(doc)
    _enriquecer_precios(doc)
    doc["created_at"] = iso_now()
    doc["updated_at"] = iso_now()
    existencia_inicial = float(doc.get("existencia", 0))
    doc["existencia"] = 0
    await db.products.insert_one(doc)
    if existencia_inicial > 0:
        await registrar_movimiento(doc, "entrada", existencia_inicial, 0, user, "Alta inicial")
    await log_audit(user, "crear", "producto", doc["id"], f"{codigo} - {doc['descripcion']}")
    return await db.products.find_one({"id": doc["id"]}, {"_id": 0})

def _enriquecer_precios(doc: dict) -> dict:
    """Calcula neto/bruto por lista y utilidad/margen del producto (lista 1).

    Considera el indicador `precio_incluye_iva`:
      True  -> los precios capturados son BRUTOS (con IVA): se extrae el neto.
      False -> los precios capturados son NETOS: se genera el bruto.
    """
    iva_tasa = float(doc.get("iva_tasa", 8.0))
    costo = float(doc.get("costo", 0) or 0)
    incluye = bool(doc.get("precio_incluye_iva", True))
    precios = []
    for p in doc.get("precios", []):
        p = dict(p)
        sin = float(p.get("precio_sin_iva") or 0)
        con = float(p.get("precio_con_iva") or 0)
        util = float(p.get("utilidad_pct") or 0)
        if incluye:
            if con > 0:
                sin = moneycalc.neto_de_precio(con, iva_tasa, True)
            elif sin > 0:
                con = moneycalc.bruto_de_precio(sin, iva_tasa, False)
            else:
                neto_base = costo * (1 + util / 100) if util >= 0 else costo * (1 + (util or 0) / 100)
                sin = round(neto_base, 2)
                con = moneycalc.bruto_de_precio(sin, iva_tasa, False)
        else:
            if sin > 0:
                con = moneycalc.bruto_de_precio(sin, iva_tasa, False)
            elif con > 0:
                sin = moneycalc.neto_de_precio(con, iva_tasa, True)
            else:
                con = round(costo * (1 + util / 100) * (1 + iva_tasa / 100), 2) if util >= 0 else round(costo * (1 + iva_tasa / 100), 2)
                sin = moneycalc.neto_de_precio(con, iva_tasa, True)
        if costo > 0:
            util = round((sin / costo - 1) * 100, 2)
        precios.append({"nombre": p.get("nombre", "Precio"), "utilidad_pct": util,
                        "precio_sin_iva": round(sin, 2), "precio_con_iva": round(con, 2)})
    doc["precios"] = precios
    sin_ok = float(precios[0]["precio_sin_iva"]) if precios else costo
    con_ok = float(precios[0]["precio_con_iva"]) if precios else round(costo * (1 + iva_tasa / 100), 2)
    doc["precio_sin_iva"] = round(sin_ok, 2)
    doc["precio_con_iva"] = round(con_ok, 2)
    util, margen = moneycalc.utilidad_margen(sin_ok, costo)
    doc["utilidad"] = util
    doc["margen"] = margen
    return doc

@api.put("/products/{product_id}")
async def update_product(product_id: str, data: ProductInput, user: dict = Depends(require_permission("producto.editar"))):
    existing = await db.products.find_one({"id": product_id})
    if not existing:
        raise HTTPException(404, "Producto no encontrado")
    doc = data.model_dump()
    doc.pop("existencia", None)  # existencia solo cambia por movimientos
    doc["codigo"] = existing["codigo"]
    # Se respeta exactamente el arreglo enviado; si viene vacío/ausente se
    # conservan los códigos de barras ya registrados (incluye el codigo).
    if not doc.get("codigos_barras"):
        doc["codigos_barras"] = existing.get("codigos_barras", [])
    _enriquecer_precios(doc)
    doc["updated_at"] = iso_now()
    await db.products.update_one({"id": product_id}, {"$set": doc})
    await log_audit(user, "editar", "producto", product_id, existing["codigo"])
    return await db.products.find_one({"id": product_id}, {"_id": 0})

@api.patch("/products/{product_id}/estado")
async def change_estado(product_id: str, estado: str, user: dict = Depends(require_permission("producto.baja"))):
    await db.products.update_one({"id": product_id}, {"$set": {"estado": estado, "updated_at": iso_now()}})
    await log_audit(user, "cambio_estado", "producto", product_id, f"estado={estado}")
    return await db.products.find_one({"id": product_id}, {"_id": 0})

@api.get("/products/{product_id}/movimientos")
async def product_movements(product_id: str, user: dict = Depends(get_current_user)):
    movs = await db.inventory_movements.find({"product_id": product_id}, {"_id": 0}).sort("fecha", 1).to_list(1000)
    return movs

@api.post("/products/{product_id}/ajuste")
async def adjust_inventory(product_id: str, data: InventoryAdjust, user: dict = Depends(require_permission("inventario.ajuste"))):
    p = await db.products.find_one({"id": product_id})
    if not p:
        raise HTTPException(404, "Producto no encontrado")
    entrada = salida = 0.0
    if data.tipo in ("entrada", "devolucion"):
        entrada = abs(data.cantidad)
    elif data.tipo in ("salida", "merma"):
        salida = abs(data.cantidad)
    elif data.tipo in ("ajuste", "correccion"):
        # ajuste/corrección puede ser + o -
        if data.cantidad >= 0:
            entrada = data.cantidad
        else:
            salida = abs(data.cantidad)
    else:
        raise HTTPException(400, "Tipo de movimiento inválido")
    costo = float(data.costo) if (data.costo and float(data.costo) > 0) else float(p.get("costo", 0) or 0)
    nueva = await registrar_movimiento(p, data.tipo, entrada, salida, user, data.documento,
                                       data.concepto, costo=costo, motivo=data.motivo or "",
                                       observaciones=data.observaciones or "")
    await log_audit(user, "ajuste_inventario", "producto", product_id, f"{data.tipo} {data.cantidad}")
    return {"existencia": nueva}

@api.get("/inventory/movements")
async def inventory_movements(tipo: Optional[str] = None, q: Optional[str] = None,
                              desde: Optional[str] = None, hasta: Optional[str] = None,
                              skip: int = 0, limit: int = 200,
                              user: dict = Depends(get_current_user)):
    limit = max(1, min(int(limit), 500))
    skip = max(0, int(skip))
    query: dict = {}
    if tipo and tipo != "all":
        query["tipo"] = tipo
    if q:
        rx = {"$regex": sanitize_search_term(q), "$options": "i"}
        query["$or"] = [{"codigo": rx}, {"descripcion": rx}, {"documento": rx}, {"usuario_nombre": rx}]
    if desde or hasta:
        rango: dict = {}
        if desde:
            rango["$gte"] = desde
        if hasta:
            rango["$lte"] = hasta + "T23:59:59"
        query["fecha"] = rango
    total = await db.inventory_movements.count_documents(query)
    docs = await db.inventory_movements.find(query, {"_id": 0}).sort("fecha", -1).skip(skip).limit(limit).to_list(limit)
    return {"total": total, "movimientos": docs}

# =========================================================================
# CATEGORÍAS
# =========================================================================
class CategoryInput(BaseModel):
    nombre: str
    clave: Optional[str] = ""
    descripcion: Optional[str] = ""
    ficha_tecnica: Optional[str] = ""
    imagen_url: Optional[str] = ""

@api.get("/categories")
async def list_categories(user: dict = Depends(get_current_user)):
    counts = {}
    async for r in db.products.aggregate([
        {"$match": {"clasificacion": {"$nin": ["", None]}}},
        {"$group": {"_id": "$clasificacion", "count": {"$sum": 1}}},
    ]):
        counts[r["_id"]] = r["count"]
    managed = {}
    async for c in db.categories.find({}, {"_id": 0}):
        managed[c["nombre"]] = c
    nombres = set(counts) | set(managed)
    nombres = {n for n in nombres if n}
    out = []
    for n in sorted(nombres):
        m = managed.get(n, {})
        out.append({"nombre": n, "clave": m.get("clave", ""), "descripcion": m.get("descripcion", ""),
                    "ficha_tecnica": m.get("ficha_tecnica", ""), "imagen_url": m.get("imagen_url", ""),
                    "count": counts.get(n, 0)})
    return out

@api.post("/categories")
async def upsert_category(data: CategoryInput, user: dict = Depends(require_permission("producto.editar"))):
    if not data.nombre.strip():
        raise HTTPException(400, "El nombre es obligatorio")
    doc = data.model_dump()
    doc["updated_at"] = iso_now()
    await db.categories.update_one({"nombre": data.nombre}, {"$set": doc}, upsert=True)
    await log_audit(user, "editar", "categoria", data.nombre, "Categoría actualizada")
    return {"ok": True, "nombre": data.nombre}

@api.delete("/categories/{nombre}")
async def delete_category(nombre: str, user: dict = Depends(require_permission("producto.editar"))):
    nombre = nombre.strip()
    if not nombre:
        raise HTTPException(400, "El nombre es obligatorio")
    res = await db.categories.delete_one({"nombre": nombre})
    if res == 0:
        raise HTTPException(404, "Categoría no encontrada")
    await log_audit(user, "eliminar", "categoria", nombre, "Categoría eliminada")
    return {"ok": True, "nombre": nombre}

@api.post("/categories/sync")
async def sync_categories(user: dict = Depends(get_current_user)):
    """Sincroniza la colección categories con las clasificaciones de productos."""
    creadas = 0
    async for r in db.products.aggregate([
        {"$match": {"clasificacion": {"$nin": ["", None]}}},
        {"$group": {"_id": "$clasificacion"}},
    ]):
        await db.categories.update_one({"nombre": r["_id"]}, {"$set": {"nombre": r["_id"]}}, upsert=True)
        creadas += 1
    await log_audit(user, "sincronizar", "categoria", "", f"{creadas} categorías sincronizadas")
    return {"ok": True, "sincronizadas": creadas}

@api.get("/categories/export/excel")
async def export_categories(user: dict = Depends(require_permission("exportar"))):
    cats = []
    async for c in db.categories.find({}, {"_id": 0}):
        cats.append({
            "NOMBRE": c.get("nombre", ""), "CLAVE": c.get("clave", ""),
            "DESCRIPCION": c.get("descripcion", ""), "FICHA_TECNICA": c.get("ficha_tecnica", ""),
            "IMAGEN_URL": c.get("imagen_url", ""),
        })
    df = pd.DataFrame(cats or [{c: None for c in ["NOMBRE", "CLAVE", "DESCRIPCION", "FICHA_TECNICA", "IMAGEN_URL"]}])
    data = df_to_excel_bytes(df)
    return StreamingResponse(io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=categorias.xlsx"})

@api.get("/categories/plantilla/excel")
async def plantilla_categories(user: dict = Depends(get_current_user)):
    df = pd.DataFrame(columns=["NOMBRE", "CLAVE", "DESCRIPCION", "FICHA_TECNICA", "IMAGEN_URL"])
    data = df_to_excel_bytes(df)
    return StreamingResponse(io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_categorias.xlsx"})

@api.post("/categories/import/preview")
async def import_categories_preview(file: UploadFile = File(...), user: dict = Depends(require_permission("importar"))):
    content = await file.read()
    df = read_import_table(content, file.filename or "").fillna("")
    df.columns = [str(c).strip().upper() for c in df.columns]
    rows = df.to_dict("records")
    if not rows:
        raise HTTPException(400, "El archivo no contiene registros.")
    if not any(c == "NOMBRE" for c in df.columns):
        raise HTTPException(400, "El archivo no tiene la columna NOMBRE. Descarga la plantilla de categorías.")
    preview, vistos = [], set()
    total = nuevos = existentes = con_errores = 0
    for i, r in enumerate(rows):
        nombre = str(r.get("NOMBRE", "")).strip()
        errores = []
        if not nombre:
            errores.append({"campo": "NOMBRE", "valor": "", "motivo": "El nombre es obligatorio"})
        elif nombre in vistos:
            errores.append({"campo": "NOMBRE", "valor": nombre, "motivo": "Nombre duplicado en el archivo"})
        vistos.add(nombre)
        existe = bool(await db.categories.find_one({"nombre": nombre}))
        if errores:
            con_errores += 1
        elif existe:
            existentes += 1
        else:
            nuevos += 1
        data = {
            "nombre": nombre,
            "clave": str(r.get("CLAVE", "")).strip(),
            "descripcion": str(r.get("DESCRIPCION", "")).strip(),
            "ficha_tecnica": str(r.get("FICHA_TECNICA", "")).strip(),
            "imagen_url": str(r.get("IMAGEN_URL", "")).strip(),
        }
        preview.append({"fila": i + 2, "codigo": nombre, "nombre": nombre, "descripcion": data["descripcion"],
                        "existe": existe, "errores": errores, "data": data})
        total += 1
    return {"total": total, "nuevos": nuevos, "existentes": existentes, "con_errores": con_errores,
            "columnas": ["NOMBRE", "CLAVE", "DESCRIPCION", "FICHA_TECNICA", "IMAGEN_URL"], "preview": preview}

@api.post("/categories/import/confirm")
async def import_categories_confirm(payload: dict, user: dict = Depends(require_permission("importar"))):
    rows = payload.get("rows", [])
    mode = payload.get("mode", "ambos")
    creados = actualizados = omitidos = 0
    for r in rows:
        if r.get("errores"):
            omitidos += 1; continue
        d = r.get("data", {})
        nombre = str(d.get("nombre", "")).strip()
        if not nombre:
            omitidos += 1; continue
        existing = await db.categories.find_one({"nombre": nombre})
        if existing and mode == "nuevos":
            omitidos += 1; continue
        if not existing and mode == "actualizar":
            omitidos += 1; continue
        doc = {"nombre": nombre, "clave": d.get("clave", ""), "descripcion": d.get("descripcion", ""),
               "ficha_tecnica": d.get("ficha_tecnica", ""), "imagen_url": d.get("imagen_url", ""),
               "updated_at": iso_now()}
        await db.categories.update_one({"nombre": nombre}, {"$set": doc}, upsert=True)
        if existing:
            actualizados += 1
        else:
            creados += 1
    if creados or actualizados:
        await log_audit(user, "importar", "categoria", "", f"{creados} creados, {actualizados} actualizados, {omitidos} omitidos")
    return {"creados": creados, "actualizados": actualizados, "omitidos": omitidos}

# =========================================================================
# CLIENTES
# =========================================================================
# Mapeo de la estructura DBF heredada -> campos internos (nombre, tipo)
# tipo: text | num | int | bool | date
CLIENT_IMPORT_MAP = {
    "CLAVE": ("codigo", "text"), "NOMBRE": ("nombre", "text"),
    "CONTRASENA": ("contrasena", "text"), "REPRESENTA": ("representa", "text"),
    "TELOFICINA": ("tel_oficina", "text"), "TELRESIDEN": ("tel_residencia", "text"),
    "TEL_FAX": ("tel_fax", "text"), "CELULAR": ("celular", "text"),
    "DIRECCION": ("direccion", "text"), "NOINTERIOR": ("numero_interior", "text"),
    "NOEXTERIOR": ("numero_exterior", "text"), "COLONIA": ("colonia", "text"),
    "CIUDADEDO": ("ciudad_edo", "text"), "LOCALIDAD": ("localidad", "text"),
    "REFERENCIA": ("referencias", "text"), "CIUDAD": ("ciudad", "text"),
    "ESTADO": ("estado_geo", "text"), "PAIS": ("pais", "text"),
    "ID_LOCALID": ("id_localidad", "text"), "ID_COLONIA": ("id_colonia", "text"),
    "ID_CIUDAD": ("id_ciudad", "text"), "ID_ESTADO": ("id_estado", "text"),
    "ID_PAIS": ("id_pais", "text"), "RESFISCAL": ("resfiscal", "text"),
    "NREGIDTRIB": ("nregidtrib", "text"), "CODPOSTAL": ("cp", "text"),
    "RFC": ("rfc", "text"), "ALMACEN": ("almacen", "text"),
    "FECHAALTA": ("fecha_alta", "date"), "VENDEDOR": ("vendedor", "text"),
    "PRECIOVTA": ("precio_venta", "int"), "TIPO": ("tipo_clave", "text"),
    "SALDO": ("saldo", "num"), "MENSUAL": ("mensual", "num"), "ANUAL": ("anual", "num"),
    # Alias comunes de plantillas externas (no fallar silenciosamente):
    "SALDO PENDIENTE": ("saldo", "num"), "SALDO_ACTUAL": ("saldo", "num"),
    "LIMCREDITO": ("limite_credito", "num"), "LIMITE_CREDITO": ("limite_credito", "num"),
    "LIMITE DE CREDITO": ("limite_credito", "num"), "LIMITE": ("limite_credito", "num"),
    "CREDITO": ("credito_autorizado", "bool"), "CREDITO_AUTORIZADO": ("credito_autorizado", "bool"),
    "CREDITO AUTORIZADO": ("credito_autorizado", "bool"), "AUTORIZADO": ("credito_autorizado", "bool"),
    "DIASCREDIT": ("dias_credito", "int"), "DIAS_CREDITO": ("dias_credito", "int"),
    "DIAS DE CREDITO": ("dias_credito", "int"),
    "VTACREDITO": ("venta_credito", "num"), "ULTFCOMPRA": ("ult_fecha_compra", "date"),
    "ULTCCOMPRA": ("ult_monto_compra", "num"), "COMENTARIO": ("comentario", "text"),
    "USOCFDI": ("uso_cfdi", "text"), "REGFISCAL": ("reg_fiscal", "text"),
    "STATUS": ("status", "text"), "OFERTAS": ("ofertas", "bool"), "CORREOS": ("correos", "text"),
}
CLIENT_LEGACY_ORDER = list(CLIENT_IMPORT_MAP.keys())
RFC_RE = re.compile(r"^[A-ZÑ&]{3,4}\d{6}[A-Z0-9]{3}$")
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

def _to_bool(v) -> bool:
    if isinstance(v, bool):
        return v
    s = str(v).strip().upper()
    return s in ("1", "TRUE", "T", ".T.", "SI", "SÍ", "S", "X", "Y", "YES", "VERDADERO")

def _to_num(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(str(v).replace(",", "").replace("$", "").strip())
    except Exception:
        return 0.0

def _to_date(v) -> str:
    s = str(v).strip()
    if not s or s.lower() in ("nan", "nat", "none"):
        return ""
    return s[:10] if len(s) >= 10 and s[4] in "-/" else s

def legacy_status_to_estado(s: str) -> str:
    c = (s or "").strip().upper()[:1]
    if c == "S":
        return "suspendido"
    if c in ("B", "I", "C"):
        return "inactivo"
    return "activo"

def normalize_client_doc(doc: dict) -> dict:
    """Sincroniza campos derivados/compat sin romper POS."""
    doc["lista_precios"] = int(doc.get("precio_venta") or doc.get("lista_precios") or 1)
    if not doc.get("telefono"):
        doc["telefono"] = doc.get("tel_oficina") or doc.get("tel_residencia") or ""
    if not doc.get("correo") and doc.get("correos"):
        doc["correo"] = str(doc["correos"]).split(",")[0].split(";")[0].strip()
    # Coordenadas: solo se conservan si son numéricas válidas (mapas/visitas).
    for f in ("latitud", "longitud"):
        v = doc.get(f)
        if v in (None, ""):
            doc[f] = None
            continue
        try:
            doc[f] = float(v)
        except (TypeError, ValueError):
            doc[f] = None
    if doc.get("vendedor_id") in (None, ""):
        doc.pop("vendedor_id", None)
    return doc

def parse_client_row(row: dict):
    """Convierte una fila (headers legacy en mayúsculas) a doc interno + errores."""
    data, errores = {}, []
    for legacy, (field, kind) in CLIENT_IMPORT_MAP.items():
        if legacy not in row:
            continue
        raw = row.get(legacy, "")
        if kind == "bool":
            data[field] = _to_bool(raw)
        elif kind == "num":
            data[field] = round(_to_num(raw), 4)
        elif kind == "int":
            data[field] = int(_to_num(raw))
        elif kind == "date":
            data[field] = _to_date(raw)
        else:
            data[field] = str(raw).strip()
    if data.get("status"):
        data["estado"] = legacy_status_to_estado(data["status"])
    clave = data.get("codigo", "").strip()
    if not clave:
        errores.append({"campo": "CLAVE", "valor": "", "motivo": "CLAVE es obligatoria"})
    if not data.get("nombre", "").strip():
        errores.append({"campo": "NOMBRE", "valor": "", "motivo": "NOMBRE es obligatorio"})
    rfc = data.get("rfc", "").strip().upper()
    if rfc and not RFC_RE.match(rfc):
        errores.append({"campo": "RFC", "valor": rfc, "motivo": "RFC con formato inválido"})
    correo = data.get("correo") or (str(data.get("correos", "")).split(",")[0].strip())
    if correo and not EMAIL_RE.match(correo):
        errores.append({"campo": "CORREOS", "valor": correo, "motivo": "Correo con formato inválido"})
    for f in ("limite_credito", "saldo", "dias_credito"):
        if data.get(f, 0) < 0:
            errores.append({"campo": f, "valor": data.get(f), "motivo": "No puede ser negativo"})
    return data, errores


@api.get("/clients")
async def list_clients(q: Optional[str] = None, estado: Optional[str] = None,
                       tipo: Optional[str] = None, filtro: Optional[str] = None,
                       ciudad: Optional[str] = None, vendedor: Optional[str] = None,
                       rfc: Optional[str] = None, telefono: Optional[str] = None,
                       fecha_desde: Optional[str] = None, fecha_hasta: Optional[str] = None,
                       user: dict = Depends(get_current_user)):
    query = _client_export_query(q=q, estado=estado, tipo=tipo, filtro=filtro, ciudad=ciudad,
                                 vendedor=vendedor, rfc=rfc, telefono=telefono,
                                 fecha_desde=fecha_desde, fecha_hasta=fecha_hasta)
    clients = await db.clients.find(query, {"_id": 0}).sort("nombre", 1).to_list(5000)
    now = now_utc()
    mes = now.strftime("%Y-%m")
    anio = str(now.year)
    sales = await db.sales.find({"estado": "confirmada"}, {"_id": 0, "cliente_id": 1, "total": 1, "fecha": 1}).to_list(20000)
    mes_map, anio_map = {}, {}
    for s in sales:
        cid = s.get("cliente_id")
        if not cid:
            continue
        f = s.get("fecha", "")
        if f[:7] == mes:
            mes_map[cid] = mes_map.get(cid, 0) + s["total"]
        if f[:4] == anio:
            anio_map[cid] = anio_map.get(cid, 0) + s["total"]
    for c in clients:
        c["compras_mes"] = round(mes_map.get(c["id"], 0), 2)
        c["compras_anio"] = round(anio_map.get(c["id"], 0), 2)
        c["credito_disponible"] = round(float(c.get("limite_credito", 0)) - float(c.get("saldo", 0)), 2)
    return clients

@api.post("/clients")
async def create_client(data: ClientInput, user: dict = Depends(require_permission("cliente.crear"))):
    codigo = (data.codigo or "").strip() or await next_counter("cliente", "C", 5)
    if await db.clients.find_one({"codigo": codigo}):
        raise HTTPException(400, f"La CLAVE '{codigo}' ya existe")
    doc = normalize_client_doc(data.model_dump())
    # Un vendedor registra el cliente bajo su propia cartera en campo.
    if not doc.get("vendedor_id") and user.get("role") in ("vendedor", "supervisor"):
        doc["vendedor_id"] = user["id"]
    if not doc.get("vendedor") and doc.get("vendedor_id"):
        doc["vendedor"] = user.get("name", "")
    doc["codigo"] = codigo
    doc["id"] = uid()
    doc["saldo"] = float(data.saldo or 0.0)
    doc["created_at"] = iso_now()
    await db.clients.insert_one(doc)
    await log_audit(user, "crear", "cliente", doc["id"], doc["nombre"])
    return await db.clients.find_one({"id": doc["id"]}, {"_id": 0})

@api.put("/clients/{client_id}")
async def update_client(client_id: str, data: ClientInput, user: dict = Depends(require_permission("cliente.editar"))):
    existing = await db.clients.find_one({"id": client_id})
    if not existing:
        raise HTTPException(404, "Cliente no encontrado")
    doc = normalize_client_doc(data.model_dump())
    doc["codigo"] = existing["codigo"]
    doc["saldo"] = existing.get("saldo", 0.0)  # saldo lo gobiernan las ventas, no la edición
    await db.clients.update_one({"id": client_id}, {"$set": doc})
    await log_audit(user, "editar", "cliente", client_id)
    return await db.clients.find_one({"id": client_id}, {"_id": 0})

@api.patch("/clients/{client_id}/estado")
async def client_estado(client_id: str, estado: str, user: dict = Depends(require_permission("cliente.baja"))):
    await db.clients.update_one({"id": client_id}, {"$set": {"estado": estado}})
    await log_audit(user, "cambio_estado", "cliente", client_id, f"estado={estado}")
    return await db.clients.find_one({"id": client_id}, {"_id": 0})

@api.patch("/clients/{client_id}/credito-toggle")
async def toggle_credito(client_id: str, data: QuickToggle, user: dict = Depends(require_permission("credito.autorizar"))):
    c = await db.clients.find_one({"id": client_id})
    if not c:
        raise HTTPException(404, "Cliente no encontrado")
    await db.clients.update_one({"id": client_id}, {"$set": {"credito_autorizado": data.valor}})
    await log_audit(user, "editar", "cliente", client_id, f"crédito {'habilitado' if data.valor else 'deshabilitado'}")
    return await db.clients.find_one({"id": client_id}, {"_id": 0})

@api.patch("/clients/{client_id}/credito")
async def set_credito(client_id: str, data: CreditInput, user: dict = Depends(require_permission("credito.autorizar"))):
    c = await db.clients.find_one({"id": client_id})
    if not c:
        raise HTTPException(404, "Cliente no encontrado")
    await db.clients.update_one({"id": client_id}, {"$set": {
        "credito_autorizado": data.credito_autorizado, "limite_credito": data.limite_credito}})
    await log_audit(user, "editar", "cliente", client_id, f"crédito autorizado={data.credito_autorizado} límite={data.limite_credito}")
    return await db.clients.find_one({"id": client_id}, {"_id": 0})

# =========================================================================
# CAJA
# =========================================================================
async def _default_sucursal_id() -> Optional[str]:
    """Devuelve la sucursal activa por defecto (la primera). Crea "Matriz" si no hay."""
    s = await db.sucursales.find_one({"activa": True}, {"_id": 0})
    if s:
        return s["id"]
    # Compatibilidad con la semilla cosmética previa en settings.sucursales.
    st = await db.settings.find_one({"_id": "app"}, {"_id": 0})
    nombre = "Matriz"
    for su in (st or {}).get("sucursales", []):
        if su.get("activa"):
            nombre = su.get("nombre", "Matriz")
            break
    doc = {"id": uid(), "codigo": "MATRIZ", "nombre": nombre, "activa": True,
           "direccion": "", "ciudad": "", "estado": "", "cp": "", "telefono": "",
           "created_at": iso_now()}
    await db.sucursales.insert_one(doc)
    return doc["id"]

async def caja_abierta_de(user_id: str):
    return await db.cajas.find_one({"usuario_id": user_id, "estado": "abierta"}, {"_id": 0})

async def asignar_caja_numero(target: dict) -> int:
    """Asigna una sola vez un número de caja único por usuario (sin repetir).
    El rol administrador/propietario nunca recibe el número 1."""
    num = target.get("caja_numero")
    if not num:
        used = set()
        async for u in db.users.find({"active": True, "caja_numero": {"$exists": True}},
                                     {"id": 1, "caja_numero": 1}):
            if u["id"] != target.get("id"):
                used.add(u["caja_numero"])
        num = 1
        while True:
            if num in used:
                num += 1
                continue
            if target.get("role") in ADMIN_SYSTEM_ROLES and num == 1:
                num += 1
                continue
            break
        await db.users.update_one({"id": target["id"]}, {"$set": {"caja_numero": num}})
    return int(num)

@api.get("/caja/actual")
async def caja_actual(user: dict = Depends(get_current_user)):
    caja = await caja_abierta_de(user["id"])
    if not caja:
        return {"caja": None}
    movs = await db.caja_movimientos.find({"caja_id": caja["id"]}, {"_id": 0}).sort("fecha", 1).to_list(1000)
    return {"caja": caja, "movimientos": movs, "resumen": resumen_caja(caja, movs)}

def resumen_caja(caja: dict, movs: List[dict]) -> dict:
    ventas_efectivo = sum(m["monto"] for m in movs if m["tipo"] == "venta")
    entradas = sum(m["monto"] for m in movs if m["tipo"] == "entrada")
    retiros = sum(m["monto"] for m in movs if m["tipo"] in ("retiro", "gasto"))
    devoluciones = sum(m["monto"] for m in movs if m["tipo"] == "devolucion")
    esperado = round(caja["fondo_inicial"] + ventas_efectivo + entradas - retiros - devoluciones, 2)
    return {"fondo_inicial": caja["fondo_inicial"], "ventas_efectivo": round(ventas_efectivo, 2),
            "entradas": round(entradas, 2), "retiros": round(retiros, 2),
            "devoluciones": round(devoluciones, 2), "efectivo_esperado": esperado}

@api.post("/caja/abrir")
async def abrir_caja(data: CajaOpen, user: dict = Depends(require_permission("caja.abrir"))):
    if await caja_abierta_de(user["id"]):
        raise HTTPException(400, "Ya tienes una caja abierta")
    caja_numero = user.get("caja_numero")
    if (data.caja_nombre or "").strip():
        nombre = data.caja_nombre.strip()
    else:
        caja_numero = await asignar_caja_numero(user)
        nombre = f"Caja {caja_numero}"
    doc = {"id": uid(), "usuario_id": user["id"], "usuario_nombre": user["name"],
           "caja_nombre": nombre, "caja_numero": caja_numero, "fondo_inicial": data.fondo_inicial,
           "denominaciones": data.denominaciones, "metodo": data.metodo or "denominaciones",
           "sucursal_id": user.get("sucursal_id") or await _default_sucursal_id(),
           "estado": "abierta", "fecha_apertura": iso_now(), "fecha_cierre": None}
    await db.cajas.insert_one(doc)
    await log_audit(user, "abrir_caja", "caja", doc["id"], f"fondo {data.fondo_inicial}")
    return await db.cajas.find_one({"id": doc["id"]}, {"_id": 0})

@api.post("/caja/abrir-por-usuario")
async def abrir_caja_por_usuario(data: CajaOpenPorUsuario, user: dict = Depends(get_current_user)):
    if not es_admin_sistema(user):
        raise HTTPException(403, "Solo administradores/propietario pueden abrir cajas de otros usuarios")
    target = await db.users.find_one({"id": data.usuario_id})
    if not target or not target.get("active", True):
        raise HTTPException(404, "Usuario no encontrado o desactivado")
    if await caja_abierta_de(data.usuario_id):
        raise HTTPException(400, f"{target['name']} ya tiene una caja abierta")
    caja_numero = target.get("caja_numero")
    if (data.caja_nombre or "").strip():
        nombre = data.caja_nombre.strip()
    else:
        caja_numero = await asignar_caja_numero(target)
        nombre = f"Caja {caja_numero}"
    doc = {"id": uid(), "usuario_id": data.usuario_id, "usuario_nombre": target["name"],
           "caja_nombre": nombre, "caja_numero": caja_numero, "fondo_inicial": data.fondo_inicial,
           "denominaciones": data.denominaciones, "metodo": data.metodo or "denominaciones",
           "sucursal_id": target.get("sucursal_id") or await _default_sucursal_id(),
           "estado": "abierta", "fecha_apertura": iso_now(), "fecha_cierre": None}
    await db.cajas.insert_one(doc)
    await log_audit(user, "abrir_caja_por_usuario", "caja", doc["id"],
                    f"{target['name']} fondo {data.fondo_inicial}")
    return await db.cajas.find_one({"id": doc["id"]}, {"_id": 0})

@api.post("/caja/movimiento")
async def caja_movimiento(data: CajaMovimiento, user: dict = Depends(require_permission("caja.entrada"))):
    caja = await caja_abierta_de(user["id"])
    if not caja:
        raise HTTPException(400, "No tienes caja abierta")
    # Validación de retiro (§3.3): no entregar más efectivo del disponible
    # (fondo + ventas efectivo + entradas − retiros/gastos/devoluciones).
    # Forzar el retiro pese al excedente queda reservado a encargado/admin.
    if data.tipo in ("retiro", "gasto"):
        movs_previos = await db.caja_movimientos.find({"caja_id": caja["id"]}, {"_id": 0}).to_list(2000)
        res_previo = resumen_caja(caja, movs_previos)
        disponible = float(res_previo.get("efectivo_esperado", 0))
        puede_forzar = (
            es_admin_sistema(user)
            or user.get("role") == "encargado"
            or user_has_permission(user, "caja.retiro_forzado")
        )
        if abs(data.monto) > disponible + 0.01 and not (data.forzar and puede_forzar):
            raise HTTPException(status_code=409, detail={
                "mensaje": f"El monto excede el efectivo disponible (${disponible:,.2f}).",
                "disponible": disponible,
                "requiere_autorizacion": True,
                "puede_forzar": puede_forzar,
            })
    # Las entregas de efectivo (retiros) llevan folio propio RET-xxxxxx para
    # trazabilidad e impresión del ticket de entrega.
    folio = ""
    if data.tipo == "retiro":
        folio = await next_counter("retiro", "RET", 6)
    doc = {"id": uid(), "caja_id": caja["id"], "tipo": data.tipo,
           "folio": folio, "concepto": data.concepto,
           "monto": abs(data.monto), "referencia": data.referencia,
           "usuario_id": user["id"], "usuario_nombre": user["name"], "fecha": iso_now()}
    if data.evidencia_url:
        doc["evidencia_url"] = data.evidencia_url
        doc["evidencia_estado"] = "pendiente"
    await db.caja_movimientos.insert_one(doc)
    await log_audit(user, "caja_movimiento", "caja", caja["id"],
                    f"{data.tipo} {data.monto}" + (" (forzado)" if data.forzar else ""))
    return {"ok": True, "movimiento": doc}


@api.get("/caja/evidencias")
async def caja_evidencias(estado: str = "pendiente",
                          user: dict = Depends(require_permission("auditoria.ver"))):
    """Bandeja de evidencias fotográficas de gastos/retiros (solo admin)."""
    if not es_admin_sistema(user):
        raise HTTPException(403, "Solo administradores/propietario")
    flt = {"evidencia_url": {"$ne": None}}
    if estado in ("pendiente", "revisado"):
        flt["evidencia_estado"] = estado
    movs = await db.caja_movimientos.find(flt, {"_id": 0}).sort("fecha", -1).to_list(500)
    cajas = {c["id"]: c for c in await db.cajas.find({}, {"_id": 0}).to_list(1000)}
    out = []
    for m in movs:
        c = cajas.get(m.get("caja_id"), {})
        out.append({**m, "caja_nombre": c.get("caja_nombre"), "fecha_corte": c.get("fecha_cierre")})
    return out


@api.patch("/caja/movimientos/{mov_id}/revisar")
async def caja_evidencia_revisar(mov_id: str, user: dict = Depends(get_current_user)):
    """Marca una evidencia como revisada (solo admin/propietario)."""
    if not es_admin_sistema(user):
        raise HTTPException(403, "Solo administradores/propietario")
    res = await db.caja_movimientos.update_one(
        {"id": mov_id},
        {"$set": {"evidencia_estado": "revisado",
                  "revisado_por": user.get("name"), "revisado_fecha": iso_now()}})
    if not res:
        raise HTTPException(404, "Movimiento no encontrado")
    await log_audit(user, "revisar_evidencia", "caja_movimiento", mov_id)
    return {"ok": True}


@api.get("/catalogo")
async def catalogo_consulta(q: Optional[str] = None, categoria: Optional[str] = None,
                            con_existencia: Optional[bool] = None,
                            user: dict = Depends(get_current_user)):
    """Catálogo por categoría SOLO consulta: imagen, nombre, categoría,
    precio al público y existencia disponible. Sin costos ni utilidad."""
    flt = {"estado": "activo"}
    if categoria:
        flt["$or"] = [{"clasificacion": categoria}, {"categoria": categoria}]
    if q:
        rx = {"$regex": re.escape(q), "$options": "i"}
        flt["$and"] = [{"$or": [{"descripcion": rx}, {"codigo": rx}, {"codigos_barras": rx}]}]
    if con_existencia:
        flt["existencia"] = {"$gt": 0}
    docs = await db.products.find(
        flt, {"_id": 0, "id": 1, "codigo": 1, "descripcion": 1,
              "clasificacion": 1, "categoria": 1, "imagen": 1, "IMAGEN": 1,
              "precios": 1, "existencia": 1}).sort("descripcion", 1).to_list(5000)
    out = []
    for d in docs:
        precio = ""
        try:
            precio = float(((d.get("precios") or [{}])[0]).get("precio_con_iva") or 0)
        except Exception:
            precio = 0.0
        out.append({
            "id": d.get("id"),
            "nombre": d.get("descripcion") or "",
            "codigo": d.get("codigo") or "",
            "categoria": d.get("clasificacion") or d.get("categoria") or "",
            "imagen": d.get("imagen") or d.get("IMAGEN") or "",
            "precio_publico": round(precio, 2),
            "existencia": round(float(d.get("existencia") or 0), 3),
        })
    return out


@api.get("/sales/mi-reporte-hoy")
async def mi_reporte_hoy(user: dict = Depends(get_current_user)):
    """Reporte rápido PROPIO del día: tickets, total vendido y desglose por
    forma de pago. Nunca incluye utilidad/margen (§3.2)."""
    hoy = now_utc().date().isoformat()
    ventas = await db.sales.find(
        {"usuario_id": user["id"], "estado": "confirmada", "fecha": {"$regex": "^" + hoy}},
        {"_id": 0, "folio": 1, "hora": 1, "total": 1, "pagos": 1, "condicion": 1}
    ).sort("fecha", -1).to_list(2000)
    metodos = {}
    total = 0.0
    tickets = []
    for s in ventas:
        t = round(float(s.get("total", 0) or 0), 2)
        total = round(total + t, 2)
        restante = t
        metas_txt = []
        for p in (s.get("pagos") or []):
            met = p.get("metodo") or "otros"
            aplicar = min(round(float(p.get("monto", 0) or 0), 2), restante)
            metodos[met] = round(metodos.get(met, 0) + aplicar, 2)
            restante = round(restante - aplicar, 2)
            metas_txt.append(met)
        if not metas_txt:
            metas_txt = ["credito" if s.get("condicion") == "credito" else "contado"]
            metodos[metas_txt[0]] = round(metodos.get(metas_txt[0], 0) + t, 2)
        tickets.append({"folio": s.get("folio"), "hora": s.get("hora"),
                        "total": t, "metodo": " + ".join(metas_txt)})
    return {"fecha": hoy, "num_ventas": len(ventas), "total": total,
            "por_metodo": metodos, "tickets": tickets}


@api.post("/caja/movimientos/{mov_id}/comprobante")
async def caja_mov_comprobante(mov_id: str, user: dict = Depends(get_current_user)):
    """Ticket (80 mm) de un movimiento de caja: entrega de efectivo, gasto,
    entrada o devolución. Descuenta del efectivo esperado al registrarse; este
    comprobante solo documenta la entrega con folio y firmas."""
    mov = await db.caja_movimientos.find_one({"id": mov_id}, {"_id": 0})
    if not mov:
        raise HTTPException(404, "Movimiento no encontrado")
    if not es_admin_sistema(user) and mov.get("usuario_id") != user["id"]:
        raise HTTPException(403, "Solo puedes imprimir tus propios movimientos")
    caja = await db.cajas.find_one({"id": mov.get("caja_id")}, {"_id": 0}) or {}
    movs = await db.caja_movimientos.find({"caja_id": mov.get("caja_id")}, {"_id": 0}).to_list(3000)
    res = resumen_caja(caja, movs)
    settings_doc = await db.settings.find_one({"_id": "app"}, {"_id": 0}) or {}
    try:
        pdf_bytes = storage.build_entrega_pdf(mov, caja, settings_doc,
                                              efectivo_en_caja=res.get("efectivo_esperado"))
        nombre_base = (mov.get("folio") or mov_id[:8]).replace("/", "-")
        path = f"caja/comprobante-{nombre_base}-{uid()[:8]}.pdf"
        result = storage.put_object(path, pdf_bytes, "application/pdf")
    except Exception as e:
        logger.error("Comprobante de movimiento falló: %s", str(e)[:200])
        raise HTTPException(502, "No se pudo generar el ticket de entrega.")
    stored = result.get("path", path)
    await db.files.insert_one({
        "id": uid(), "storage_path": stored,
        "original_filename": f"RYSA_Entrega_{mov.get('folio') or mov_id[:8]}.pdf",
        "content_type": "application/pdf", "size": result.get("size", len(pdf_bytes)),
        "movimiento_id": mov_id, "caja_id": mov.get("caja_id"),
        "is_deleted": False, "created_at": iso_now()})
    return {"path": stored, "url": f"/api/files/{stored}",
            "filename": f"RYSA_Entrega_{mov.get('folio') or mov_id[:8]}.pdf"}

@api.post("/caja/cerrar")
async def cerrar_caja(data: CajaClose, user: dict = Depends(require_permission("caja.cerrar"))):
    if data.caja_id:
        if not es_admin_sistema(user):
            raise HTTPException(403, "Solo administradores/propietario pueden cerrar cajas de otros usuarios")
        caja = await db.cajas.find_one({"id": data.caja_id})
        if not caja:
            raise HTTPException(404, "Caja no encontrada")
        if caja.get("estado") != "abierta":
            raise HTTPException(400, "La caja ya está cerrada")
    else:
        caja = await caja_abierta_de(user["id"])
        if not caja:
            raise HTTPException(400, "No tienes caja abierta")
    movs = await db.caja_movimientos.find({"caja_id": caja["id"]}, {"_id": 0}).to_list(1000)
    res = resumen_caja(caja, movs)
    diferencia = round(data.efectivo_contado - res["efectivo_esperado"], 2)
    cierre = {**res, "efectivo_contado": data.efectivo_contado, "diferencia": diferencia}
    # Guard de carrera: solo cierra si SIGUE abierta (evita doble cierre
    # concurrente que sobrescriba el primer cierre con números ya obsoletos).
    cerradas = await db.cajas.update_one(
        {"id": caja["id"], "estado": "abierta"},
        {"$set": {"estado": "cerrada", "fecha_cierre": iso_now(), "cierre": cierre}})
    if not cerradas:
        raise HTTPException(409, "La caja acaba de ser cerrada por otro usuario")
    await log_audit(user, "cerrar_caja", "caja", caja["id"], f"diferencia {diferencia}")
    # Se devuelven los movimientos completos y el DESGLOSE de ventas del turno
    # para que la UI muestre el reporte de verificación sin más peticiones.
    desglose = await _desglose_ventas_caja(caja["id"])
    return {"cierre": cierre, "movimientos": movs, "desglose": desglose,
            "caja": {"id": caja.get("id"), "caja_nombre": caja.get("caja_nombre"),
                     "usuario_nombre": caja.get("usuario_nombre"),
                     "fondo_inicial": caja.get("fondo_inicial"),
                     "fecha_apertura": caja.get("fecha_apertura"),
                     "fecha_cierre": iso_now()}}


async def _desglose_ventas_caja(caja_id: str) -> dict:
    """Ventas confirmadas del turno agrupadas por método de pago.
    Cada pago cuenta solo hasta cubrir el total de su venta (el CAMBIO no es
    dinero que quede en caja) — mismo criterio que el movimiento de caja."""
    vsales = await db.sales.find(
        {"caja_id": caja_id, "estado": "confirmada"},
        {"_id": 0, "pagos": 1, "total": 1}).to_list(20000)
    metodos = {}
    total_vendido = 0.0
    for s in vsales:
        total_s = round(float(s.get("total", 0) or 0), 2)
        total_vendido = round(total_vendido + total_s, 2)
        restante = total_s
        for p in (s.get("pagos") or []):
            if restante <= 0:
                break
            aplicar = min(round(float(p.get("monto", 0) or 0), 2), restante)
            met = p.get("metodo") or "otros"
            metodos[met] = round(metodos.get(met, 0) + aplicar, 2)
            restante = round(restante - aplicar, 2)
    return {"num_ventas": len(vsales), "total_vendido": total_vendido,
            "metodos": metodos}


@api.get("/caja/{caja_id}/desglose")
async def caja_desglose(caja_id: str, user: dict = Depends(require_permission("caja.ver"))):
    """Desglose de ventas del corte (para verificar el reporte en cualquier
    momento, con la caja abierta o cerrada). Solo el dueño o administración."""
    caja = await db.cajas.find_one({"id": caja_id}, {"_id": 0})
    if not caja:
        raise HTTPException(404, "Corte de caja no encontrado")
    if not es_admin_sistema(user) and caja.get("usuario_id") != user["id"]:
        raise HTTPException(403, "Solo puedes consultar tus propios cortes")
    movs = await db.caja_movimientos.find({"caja_id": caja_id}, {"_id": 0}).sort("fecha", 1).to_list(5000)
    desglose = await _desglose_ventas_caja(caja_id)
    return {"caja": {"id": caja.get("id"), "caja_nombre": caja.get("caja_nombre"),
                     "usuario_nombre": caja.get("usuario_nombre"),
                     "fondo_inicial": caja.get("fondo_inicial"),
                     "fecha_apertura": caja.get("fecha_apertura"),
                     "fecha_cierre": caja.get("fecha_cierre"),
                     "cierre": caja.get("cierre")},
            "movimientos": movs,
            "resumen": resumen_caja(caja, movs),
            "desglose": desglose}

@api.get("/caja/operadores")
async def caja_operadores(user: dict = Depends(require_permission("caja.ver"))):
    if not es_admin_sistema(user):
        raise HTTPException(403, "Solo administradores/propietario pueden ver el estado global de cajas")
    usuarios = await db.users.find({"active": True}, {"_id": 0, "password_hash": 0}).to_list(1000)
    abiertas_por_usuario = {}
    cursor = db.cajas.find({"estado": "abierta"}, {"_id": 0})
    async for c in cursor:
        abiertas_por_usuario.setdefault(c["usuario_id"], []).append(c)
    out = []
    for u in usuarios:
        abiertas = abiertas_por_usuario.get(u["id"], [])
        if abiertas:
            caja = sorted(abiertas, key=lambda c: c.get("fecha_apertura", ""))[0]
            movs = await db.caja_movimientos.find({"caja_id": caja["id"]}, {"_id": 0}).to_list(1000)
            out.append({
                "usuario_id": u["id"], "usuario_nombre": u["name"], "role": u.get("role"),
                "caja_numero": u.get("caja_numero"),
                "estado": "abierta", "caja": caja, "resumen": resumen_caja(caja, movs)})
        else:
            ultima = await db.cajas.find({"usuario_id": u["id"], "estado": "cerrada"},
                                         {"_id": 0}).sort("fecha_cierre", -1).to_list(1)
            out.append({
                "usuario_id": u["id"], "usuario_nombre": u["name"], "role": u.get("role"),
                "caja_numero": u.get("caja_numero"),
                "estado": "cerrada", "caja": ultima[0] if ultima else None, "resumen": None})
    return out

@api.get("/caja/historial")
async def caja_historial(desde: Optional[str] = None, hasta: Optional[str] = None,
                         estado: Optional[str] = None, usuario_id: Optional[str] = None,
                         user: dict = Depends(require_permission("caja.ver"))):
    query = {}
    if not es_admin_sistema(user):
        query["usuario_id"] = user["id"]
    elif usuario_id:
        query["usuario_id"] = usuario_id
    if estado in ("abierta", "cerrada"):
        query["estado"] = estado
    cajas = await db.cajas.find(query, {"_id": 0}).sort("fecha_apertura", -1).to_list(500)
    d = desde[:10] if desde else None
    h = hasta[:10] if hasta else None
    if d or h:
        cajas = [c for c in cajas if (not d or (c.get("fecha_apertura", "")[:10] >= d)) and (not h or (c.get("fecha_apertura", "")[:10] <= h))]
    # Movimientos de cada corte (ledger detallado del día).
    for c in cajas:
        movs = await db.caja_movimientos.find({"caja_id": c["id"]}, {"_id": 0}).sort("fecha", 1).to_list(1000)
        c["movimientos"] = movs
    return cajas


def _caja_reporte_data(caja: dict, movs: List[dict]) -> dict:
    """Construye los datos del reporte de cierre: ledger con saldo corrido,
    totales por tipo y desglose de ventas del turno por método de pago."""
    fondo = float(caja.get("fondo_inicial", 0) or 0)
    saldo = fondo
    filas = []
    totales = {}
    for m in sorted(movs or [], key=lambda x: (x.get("fecha") or "")):
        monto = float(m.get("monto", 0) or 0)
        tipo = m.get("tipo") or ""
        entrada = round(monto, 2) if tipo in ("venta", "entrada") else 0.0
        salida = round(monto, 2) if tipo in ("retiro", "gasto", "devolucion") else 0.0
        saldo = round(saldo + entrada - salida, 2)
        totales[tipo] = round(totales.get(tipo, 0) + monto, 2)
        fecha = (m.get("fecha") or "")
        filas.append({
            "hora": fecha[11:16] if len(fecha) >= 16 else fecha[:10],
            "tipo": tipo, "concepto": m.get("concepto") or "",
            "referencia": m.get("referencia") or "",
            "usuario": m.get("usuario_nombre") or "",
            "entrada": entrada, "salida": salida, "saldo": saldo,
        })
    # Ventas del turno agrupadas por método de pago.
    ventas_metodo = {}
    num_ventas = 0
    return {"fondo": fondo, "filas": filas, "totales": totales,
            "saldo_final": saldo, "ventas_metodo": ventas_metodo,
            "num_ventas": num_ventas}


async def _caja_reporte_payload(caja_id: str, user: dict) -> dict:
    """Reúne caja + movimientos + resumen para el reporte exportable."""
    caja = await db.cajas.find_one({"id": caja_id}, {"_id": 0})
    if not caja:
        raise HTTPException(404, "Corte de caja no encontrado")
    if not es_admin_sistema(user) and caja.get("usuario_id") != user["id"]:
        raise HTTPException(403, "Solo puedes consultar tus propios cortes")
    movs = await db.caja_movimientos.find({"caja_id": caja_id}, {"_id": 0}).sort("fecha", 1).to_list(5000)
    rep = _caja_reporte_data(caja, movs)
    # Desglose de ventas del turno por método de pago.
    vsales = await db.sales.find(
        {"caja_id": caja_id, "estado": "confirmada"},
        {"_id": 0, "pagos": 1, "total": 1}).to_list(20000)
    metodos = {}
    for s in vsales:
        for p in (s.get("pagos") or []):
            met = p.get("metodo") or "otros"
            metodos[met] = round(metodos.get(met, 0) + float(p.get("monto", 0) or 0), 2)
    settings_doc = await db.settings.find_one({"_id": "app"}, {"_id": 0}) or {}
    return {"caja": caja, **rep, "metodos": metodos, "settings": settings_doc,
            "user_name": user.get("name")}


@api.get("/caja/{caja_id}/ventas.xlsx")
async def caja_ventas_xlsx(caja_id: str, ambito: str = "turno",
                           user: dict = Depends(require_permission("caja.ver"))):
    """Descarga el listado de ventas del TURNO (esta caja) o del DÍA
    (fecha de apertura del corte). Solo dueño de la caja o administración."""
    caja = await db.cajas.find_one({"id": caja_id}, {"_id": 0})
    if not caja:
        raise HTTPException(404, "Corte de caja no encontrado")
    if not es_admin_sistema(user) and caja.get("usuario_id") != user["id"]:
        raise HTTPException(403, "Solo puedes consultar tus propios cortes")

    if ambito == "dia":
        dia = (caja.get("fecha_apertura") or iso_now())[:10]
        flt = {"estado": "confirmada", "fecha": {"$regex": "^" + dia}}
        if not es_admin_sistema(user):
            flt["$or"] = [{"usuario_id": user["id"]}, {"vendedor_id": user["id"]}]
        titulo = f"VENTAS DEL DÍA {dia}"
    else:
        flt = {"estado": "confirmada", "caja_id": caja_id}
        titulo = "VENTAS DEL TURNO"

    ventas = await db.sales.find(flt, {"_id": 0}).sort("fecha", 1).to_list(20000)
    metodos = {}
    total = 0.0
    headers = ["Folio", "Hora", "Cliente", "Vendedor", "Método(s)", "Artículos", "Total"]
    rows = []
    for s in ventas:
        pagos = s.get("pagos") or []
        met_txt = " + ".join((p.get("metodo") or "otros") for p in pagos) or (
            "crédito" if s.get("condicion") == "credito" else "contado")
        for p in pagos:
            m = p.get("metodo") or "otros"
            metodos[m] = round(metodos.get(m, 0) + float(p.get("monto", 0) or 0), 2)
        total = round(total + float(s.get("total", 0) or 0), 2)
        arts = sum(int(float(i.get("cantidad", 0) or 0)) for i in (s.get("items") or []))
        rows.append({
            "Folio": s.get("folio") or "",
            "Hora": (s.get("fecha") or "")[11:16],
            "Cliente": s.get("cliente_nombre") or "",
            "Vendedor": s.get("vendedor_nombre") or s.get("usuario_nombre") or "",
            "Método(s)": met_txt,
            "Artículos": arts,
            "Total": round(float(s.get("total", 0) or 0), 2),
        })
    rows.append({})
    rows.append({"Folio": "TOTAL", "Cliente": f"{len(ventas)} ventas", "Total": round(total, 2)})
    for met, monto in sorted(metodos.items()):
        rows.append({"Folio": "", "Cliente": f"  {met}", "Total": round(monto, 2)})

    filtros = (f"Caja: {caja.get('caja_nombre') or 'Caja'} · Cajero: {caja.get('usuario_nombre') or '—'} · "
               f"Apertura: {(caja.get('fecha_apertura') or '')[:16].replace('T', ' ')} · Ámbito: {'día' if ambito == 'dia' else 'turno'}")
    data = exports.excel_bytes(rows, headers, sheet_name="Ventas",
                               title=f"{titulo} - GRUPO RYSA")
    stamp = (caja.get("fecha_apertura") or iso_now())[:10]
    return StreamingResponse(io.BytesIO(data),
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=ventas_{ambito}_{stamp}.xlsx"})


@api.get("/caja/{caja_id}/reporte.xlsx")
async def caja_reporte_xlsx(caja_id: str, user: dict = Depends(require_permission("caja.ver"))):
    d = await _caja_reporte_payload(caja_id, user)
    caja = d["caja"]
    headers = ["Hora", "Tipo", "Concepto", "Referencia", "Usuario", "Entrada", "Salida", "Saldo"]
    rows = [{"Hora": f["hora"], "Tipo": f["tipo"], "Concepto": f["concepto"],
             "Referencia": f["referencia"], "Usuario": f["usuario"],
             "Entrada": f["entrada"] if f["entrada"] else "",
             "Salida": f["salida"] if f["salida"] else "",
             "Saldo": f["saldo"]} for f in d["filas"]]
    rows.append({})
    rows.append({"Concepto": "Fondo inicial", "Saldo": "", "Entrada": round(d["fondo"], 2)})
    for tipo, monto in sorted(d["totales"].items()):
        rows.append({"Concepto": f"Total {tipo}", "Entrada": round(monto, 2)})
    rows.append({"Concepto": "EFECTIVO ESPERADO", "Entrada": round(float((caja.get("cierre") or {}).get("efectivo_esperado", d["saldo_final"])), 2)})
    if caja.get("cierre"):
        rows.append({"Concepto": "Efectivo contado", "Entrada": round(float(caja["cierre"].get("efectivo_contado", 0)), 2)})
        rows.append({"Concepto": "DIFERENCIA", "Entrada": round(float(caja["cierre"].get("diferencia", 0)), 2)})
    if d["metodos"]:
        rows.append({})
        rows.append({"Concepto": "Ventas por método de pago"})
        for met, monto in sorted(d["metodos"].items()):
            rows.append({"Concepto": f"  {met}", "Entrada": round(monto, 2)})
    filtros = f"Caja: {caja.get('caja_nombre') or 'Caja'} · Cajero: {caja.get('usuario_nombre') or '—'} · Apertura: {(caja.get('fecha_apertura') or '')[:16].replace('T',' ')}"
    data = exports.excel_bytes(rows, headers, sheet_name="Corte de caja",
                               title="REPORTE DE CORTE DE CAJA - GRUPO RYSA")
    stamp = (caja.get("fecha_apertura") or iso_now())[:10]
    return StreamingResponse(io.BytesIO(data),
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=caja_{stamp}.xlsx"})


@api.get("/caja/{caja_id}/reporte.pdf")
async def caja_reporte_pdf(caja_id: str, user: dict = Depends(require_permission("caja.ver"))):
    d = await _caja_reporte_payload(caja_id, user)
    caja = d["caja"]
    cierre = caja.get("cierre") or {}
    mx = lambda v: f"${float(v or 0):,.2f}"  # noqa: E731
    num_ventas = await db.sales.count_documents({"caja_id": caja_id, "estado": "confirmada"})
    headers = ["Hora", "Tipo", "Concepto", "Ref.", "Entrada", "Salida", "Saldo"]
    rows = [[f["hora"], f["tipo"], f["concepto"], f["referencia"],
             round(f["entrada"], 2) if f["entrada"] else "",
             round(f["salida"], 2) if f["salida"] else "",
             round(f["saldo"], 2)] for f in d["filas"]]
    filtros = (f"Caja: {caja.get('caja_nombre') or 'Caja'} · Cajero: {caja.get('usuario_nombre') or '—'} · "
               f"Apertura: {(caja.get('fecha_apertura') or '')[:16].replace('T', ' ')} · "
               + ("CERRADA" if caja.get("estado") == "cerrada" else "ABIERTA") +
               f" · Fondo inicial: {mx(d['fondo'])} · Ventas del turno: {num_ventas}")
    if cierre:
        filtros += (f" · Esperado: {mx(cierre.get('efectivo_esperado'))} · "
                    f"Contado: {mx(cierre.get('efectivo_contado'))} · Diferencia: {mx(cierre.get('diferencia'))}")
    if d["metodos"]:
        filtros += " · Métodos: " + ", ".join(f"{k} {mx(v)}" for k, v in sorted(d["metodos"].items()))
    data = exports.pdf_bytes("REPORTE DE CORTE DE CAJA", headers, rows,
                             settings=d["settings"], user_name=d["user_name"],
                             filtros=filtros, col_weights=[1, 1.4, 3.2, 1.6, 1.4, 1.4, 1.6])
    stamp = (caja.get("fecha_apertura") or iso_now())[:10]
    return StreamingResponse(io.BytesIO(data), media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename=caja_{stamp}.pdf"})

# =========================================================================
# VENTAS / POS
# =========================================================================
def calcular_venta(items: List[dict], descuento_global: float, precios_incluyen_iva: bool = True):
    # Convención (ver moneycalc): True = los precios enviados incluyen IVA (brutos,
    # se extrae el neto); False = precios netos (se suma el IVA). El TOTAL siempre
    # es el bruto. Devuelve además `detalle` con las líneas enriquecidas
    # (importe_neto/bruto, iva_linea, precios unitarios neto/bruto).
    res = moneycalc.calcular_venta(items, descuento_global, bool(precios_incluyen_iva))
    return {
        "subtotal": res["subtotal"],
        "iva_total": res["iva_total"],
        "descuento_total": res["descuento_total"],
        "total": res["total"],
        "detalle": res["detalle"],
    }

@api.get("/sales")
async def list_sales(rango: Optional[str] = None, estado: Optional[str] = None,
                     desde: Optional[str] = None, hasta: Optional[str] = None,
                     vendedor_id: Optional[str] = None, q: Optional[str] = None,
                     user: dict = Depends(get_current_user)):
    if vendedor_id and not ver_todas_ventas(user):
        raise HTTPException(403, "No tienes permiso para filtrar ventas de otros operadores")
    query = {}
    if estado:
        query["estado"] = estado
    if not ver_todas_ventas(user):
        # Usuario normal: solo ve SUS ventas (las realiza él mismo).
        query["vendedor_id"] = user["id"]
    elif vendedor_id:
        query["vendedor_id"] = vendedor_id
    if q:
        rx = {"$regex": sanitize_search_term(q), "$options": "i"}
        query["$or"] = [{"folio": rx}, {"cliente_nombre": rx}]
    sales = await db.sales.find(query, {"_id": 0}).sort("fecha", -1).to_list(3000)
    now = now_utc()
    d = desde[:10] if desde else None
    h = hasta[:10] if hasta else None
    if rango and rango not in ("all", "rango"):
        if rango == "hoy":
            d = h = now.date().isoformat()
        elif rango == "semana":
            d = (now - timedelta(days=now.weekday())).date().isoformat(); h = now.date().isoformat()
        elif rango == "mes":
            d = now.strftime("%Y-%m-01"); h = now.date().isoformat()
        elif rango == "mes_anterior":
            first_this = now.replace(day=1)
            last_prev = first_this - timedelta(days=1)
            d = last_prev.strftime("%Y-%m-01"); h = last_prev.date().isoformat()
        elif rango == "anio":
            d = f"{now.year}-01-01"; h = now.date().isoformat()
    if d or h:
        sales = [s for s in sales if (not d or s.get("fecha", "")[:10] >= d) and (not h or s.get("fecha", "")[:10] <= h)]
    return sales

@api.get("/sales/por-folio")
async def sale_por_folio(folio: str, user: dict = Depends(get_current_user)):
    """Busca una venta por folio (exacto o parcial) SIN limitar por fecha ni por
    el tope de 3000 registros, para poder reimprimir/reenviar tickets históricos."""
    query = {"$or": [{"folio": {"$regex": sanitize_search_term(folio), "$options": "i"}},
                     {"cliente_nombre": {"$regex": sanitize_search_term(folio), "$options": "i"}}]}
    if not ver_todas_ventas(user):
        query["vendedor_id"] = user["id"]
    docs = await db.sales.find(query, {"_id": 0}).sort("fecha", -1).to_list(200)
    return docs

@api.get("/sales/{sale_id}")
async def get_sale(sale_id: str, user: dict = Depends(get_current_user)):
    s = await db.sales.find_one({"id": sale_id}, {"_id": 0})
    if not s:
        raise HTTPException(404, "Venta no encontrada")
    # Visibilidad: solo roles autorizados ven ventas ajenas.
    if not ver_todas_ventas(user) and s.get("vendedor_id") != user["id"]:
        raise HTTPException(403, "No tienes permiso para ver esta venta")
    return s

@api.put("/sales/{sale_id}/cliente")
async def set_sale_cliente(sale_id: str, payload: dict, user: dict = Depends(require_permission("venta.crear"))):
    s = await db.sales.find_one({"id": sale_id})
    if not s:
        raise HTTPException(404, "Venta no encontrada")
    if not ver_todas_ventas(user) and s.get("vendedor_id") != user["id"]:
        raise HTTPException(403, "No tienes permiso para modificar esta venta")
    if s.get("facturado"):
        raise HTTPException(400, "La venta ya fue facturada")
    await db.sales.update_one({"id": sale_id}, {"$set": {
        "cliente_id": payload.get("cliente_id"), "cliente_nombre": payload.get("cliente_nombre", "")}})
    return {"ok": True}

@api.post("/sales")
async def create_sale(data: SaleInput, user: dict = Depends(require_permission("venta.crear"))):
    """Crea una venta o cotización. Lógica completa en `_crear_venta` (reutilizada por
    la conversión de cotizaciones y pedidos a venta)."""
    return await _crear_venta(user, data)


async def _crear_venta(user: dict, data: SaleInput):
    if not data.items:
        raise HTTPException(400, "La venta no tiene productos")
    cliente = None
    if data.cliente_id:
        cliente = await db.clients.find_one({"id": data.cliente_id}, {"_id": 0})
    cliente_nombre = cliente["nombre"] if cliente else "Público General"
    # Vendedor: automático = usuario autenticado. Solo roles autorizados con
    # `venta.cambiar_operador` pueden registrar la venta a nombre de otro operador.
    # El resto SIEMPRE usa su propio id (imposible suplantar operador).
    puede_cambiar_operador = user_has_permission(user, "venta.cambiar_operador")
    vendedor_id = data.vendedor_id if (puede_cambiar_operador and data.vendedor_id) else user["id"]
    vendedor_nombre = user["name"]
    if vendedor_id != user["id"]:
        v = await db.users.find_one({"id": vendedor_id}, {"_id": 0})
        if v:
            vendedor_nombre = v["name"]
    items = [it.model_dump() for it in data.items]
    es_cotizacion = data.tipo_venta == "cotizacion"
    # Override de inventario negativo: solo roles con el permiso dedicado pueden
    # vender con existencia insuficiente (queda inventario negativo).
    override_inv = None
    if data.allow_negative_inventory and not es_cotizacion:
        if not user_has_permission(user, "inventario.autorizar_negativo"):
            raise HTTPException(403, "No tienes permiso para vender sin inventario suficiente")
        override_inv = {
            "allow_negative_inventory": True,
            "override_user_id": user["id"],
            "override_user_nombre": user["name"],
            "override_reason": (data.override_reason or "").strip(),
            "override_timestamp": iso_now(),
        }
    # Caja obligatoria: ninguna venta (no cotización) puede finalizar sin caja abierta.
    if not es_cotizacion:
        caja = await caja_abierta_de(user["id"])
        if not caja:
            raise HTTPException(409, "No hay caja abierta. Abre una caja antes de vender.")
        if caja.get("estado") != "abierta":
            raise HTTPException(409, "La caja no está abierta")
    else:
        caja = None
    for it in items:
        if not it["product_id"]:
            continue  # línea sin inventario (p. ej. recarga remitida)
        p = await db.products.find_one({"id": it["product_id"]})
        if not p:
            raise HTTPException(400, f"Producto {it['codigo']} no existe")
        if p.get("estado") != "activo":
            raise HTTPException(400, f"Producto {p['codigo']} no está activo")
        if not es_cotizacion:
            controles = p.get("controles", {}) or {}
            controlar = controles.get("controlar_inventario", True)
            permitir_neg = controles.get("permitir_inventario_negativo", False)
            if controlar and not permitir_neg and not override_inv and float(p.get("existencia", 0)) < it["cantidad"]:
                raise HTTPException(400, f"Existencia insuficiente de {p['codigo']} (disp: {p.get('existencia',0)})")
        # Snapshot histórico de la venta: costo y clasificación se congelan aquí
        # para que los reportes históricos no cambien si el precio/costo se edita.
        it["costo"] = float(p.get("costo") or 0)
        it["clasificacion"] = p.get("clasificacion") or ""
        it["linea"] = p.get("linea") or ""
        it["precio_incluye_iva"] = bool(p.get("precio_incluye_iva", True))
    totales = calcular_venta(items, data.descuento_global, data.precios_incluyen_iva)
    total = totales["total"]
    # Líneas enriquecidas con importe_neto/bruto e IVA por línea (snapshot).
    items = totales.pop("detalle", items)
    pagos = [p.model_dump() for p in data.pagos]
    pagado = sum(p["monto"] for p in pagos)
    cambio = 0.0
    saldo = 0.0
    now = now_utc()

    # Idempotencia POS: un reintento con la misma key devuelve la venta existente
    # sin generar un folio nuevo ni una segunda venta.
    if data.idempotency_key:
        existing = await _pgpos._existing_by_key(data.idempotency_key)
        if existing:
            return existing

    if es_cotizacion:
        folio = await next_counter("cotizacion", "COT", 6)
        estado = "cotizacion"
        # Regla de vencimiento de cotizaciones: SIEMPRE emisión + 2 días
        # (calculado en el servidor; no editable desde el POS).
        fecha_vencimiento = (now_utc() + timedelta(days=2)).strftime("%Y-%m-%d")
    else:
        fecha_vencimiento = ""
        if data.condicion == "contado":
            if round(pagado, 2) + 0.01 < round(total, 2):
                raise HTTPException(400, f"El pago ({round(pagado,2)}) es menor al total ({round(total,2)})")
            cambio = max(0.0, round(pagado - total, 2))
        else:
            if not cliente:
                raise HTTPException(400, "Selecciona un cliente para venta a crédito")
            if not cliente.get("credito_autorizado", False):
                raise HTTPException(400, "El cliente no tiene crédito autorizado")
            disponible = float(cliente.get("limite_credito", 0)) - float(cliente.get("saldo", 0))
            if total > disponible + 0.01:
                raise HTTPException(400, f"Excede el crédito disponible ({round(disponible, 2)})")
            saldo = total
        folio = await next_counter("venta", "V", 6)
        estado = "confirmada"
    sale = {
        "id": uid(), "folio": folio, "fecha": iso_now(),
        "hora": now.strftime("%H:%M"), "usuario_id": user["id"], "usuario_nombre": user["name"],
        "vendedor_id": vendedor_id, "vendedor_nombre": vendedor_nombre,
        "cliente_id": data.cliente_id, "cliente_nombre": cliente_nombre,
        "items": items, **totales, "tipo_venta": data.tipo_venta, "condicion": data.condicion,
        "pagos": pagos, "cambio": cambio, "saldo": saldo, "estado": estado,
        "factura": False, "caja_id": caja["id"] if (caja and not es_cotizacion) else None,
        "sucursal_id": (caja or {}).get("sucursal_id") or user.get("sucursal_id"),
        "lista_precios": data.lista_precios,
        "fecha_vencimiento": fecha_vencimiento,
    }
    if override_inv:
        sale["inventario_override"] = override_inv

    # --- Persistencia atómica (PostgreSQL): una sola transacción ---
    try:
        creada = await _pgpos.crear_venta_pg(
            user=user, sale=sale, items=items, pagos=pagos, total=total,
            es_cotizacion=es_cotizacion, caja=caja, condicion=data.condicion,
            cliente=cliente, folio=folio, idempotency_key=data.idempotency_key,
            override_inv=override_inv)
    except _pgpos.VentaError as e:
        raise HTTPException(status_code=e.status, detail=e.message)

    # Generador ÚNICO de documentos: pre-genera ticket + carta UNA vez aquí;
    # vista previa / descarga / impresión / WhatsApp / correo usarán SIEMPRE
    # estos mismos archivos. Best-effort: si falla, se generan bajo demanda.
    if not es_cotizacion:
        try:
            import documentos as _docs
            await _docs.asegurar_documentos(creada["id"])
        except Exception as e:
            logger.warning("Pre-generación de documentos pendiente %s: %s",
                           folio, str(e)[:120])
    return creada

# =========================================================================
# COTIZACIONES (listado, detalle y conversión a venta)
# =========================================================================
@api.get("/cotizaciones")
async def cotizaciones_list(estado: Optional[str] = None, vendedor_id: Optional[str] = None,
                            cliente_id: Optional[str] = None, desde: Optional[str] = None,
                            hasta: Optional[str] = None, q: Optional[str] = None,
                            user: dict = Depends(get_current_user)):
    flt = {"tipo_venta": "cotizacion"}
    if estado and estado != "todos":
        flt["estado"] = estado
    if vendedor_id:
        flt["vendedor_id"] = vendedor_id
    if cliente_id:
        flt["cliente_id"] = cliente_id
    docs = await db.sales.find(flt, {"_id": 0}).sort("fecha", -1).to_list(100000)
    if desde:
        docs = [d for d in docs if (d.get("fecha") or "")[:10] >= desde]
    if hasta:
        docs = [d for d in docs if (d.get("fecha") or "")[:10] <= hasta]
    if q:
        ql = q.lower().strip()
        docs = [d for d in docs if ql in " ".join(str(d.get(k) or "") for k in ("folio", "cliente_nombre", "vendedor_nombre")).lower()]
    return docs


@api.get("/cotizaciones/{cot_id}")
async def cotizacion_detail(cot_id: str, user: dict = Depends(get_current_user)):
    doc = await db.sales.find_one({"id": cot_id, "tipo_venta": "cotizacion"}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Cotización no encontrada")
    return doc


@api.post("/cotizaciones/{cot_id}/convertir")
async def cotizacion_convertir(cot_id: str, data: CotizacionConvertInput,
                               user: dict = Depends(require_permission("venta.crear"))):
    """Convierte una cotización guardada en venta real. Reutiliza la pipeline
    completa de `_crear_venta` (caja obligatoria, inventario, crédito, caja pg)."""
    cot = await db.sales.find_one({"id": cot_id, "tipo_venta": "cotizacion"}, {"_id": 0})
    if not cot:
        raise HTTPException(404, "Cotización no encontrada")
    if cot.get("estado") != "cotizacion":
        raise HTTPException(409, f"Esta cotización ya no es convertible (estado: {cot.get('estado')})")
    items = []
    for it in cot.get("items", []):
        items.append(SaleItem(
            product_id=it.get("product_id"), codigo=it.get("codigo") or "",
            descripcion=it.get("descripcion") or "", cantidad=float(it.get("cantidad") or 0),
            unidad=it.get("unidad") or "PZA",
            precio=float(it.get("precio") or it.get("precio_bruto") or 0),
            iva_tasa=float(it.get("iva_tasa") or 8), descuento=float(it.get("descuento") or 0),
            comentario=it.get("comentario") or ""))
    payload = SaleInput(
        cliente_id=cot.get("cliente_id"), items=items,
        descuento_global=float(cot.get("descuento_global") or 0),
        condicion=data.condicion, pagos=data.pagos,
        lista_precios=int(cot.get("lista_precios") or 1),
        tipo_venta="directa", vendedor_id=data.vendedor_id, precios_incluyen_iva=True)
    sale = await _crear_venta(user, payload)
    await db.sales.update_one({"id": cot_id}, {"$set": {
        "estado": "convertida", "convertida_a": sale["id"], "convertida_folio": sale["folio"],
        "convertida_en": iso_now()}})
    await log_audit(user, "cotizacion_convertir", "venta", cot_id,
                    f"{cot.get('folio')} -> venta {sale['folio']}")
    return sale


# =========================================================================
# COMPROBANTES DE PAGO POR QR (cotizaciones) — §1-31
# Enlace público por token seguro + recepción de evidencia + revisión.
# ENVIAR COMPROBANTE ≠ PAGO VALIDADO: la aprobación es manual (§8).
# =========================================================================
class RevisionInput(BaseModel):
    comentario: str = ""


def _qr_cfg(settings: dict) -> dict:
    c = settings.get("qr_comprobante") or {}
    return {
        "activo": bool(c.get("activo", True)),
        "vigencia_dias": max(1, int(c.get("vigencia_dias", 30) or 30)),
        "max_mb": max(1, int(c.get("max_mb", 10) or 10)),
        "max_archivos": max(1, int(c.get("max_archivos", 6) or 6)),
    }


async def _link_activo(cotizacion_id: str):
    return await db.cot_pago_tokens.find_one(
        {"cotizacion_id": cotizacion_id, "estado": "activo"}, {"_id": 0})


@api.post("/sales/{sale_id}/pago-link")
async def crear_pago_link(sale_id: str, regenerar: bool = False,
                          request: Request = None,
                          user: dict = Depends(require_permission("cxc.abono"))):
    """Crea/reutiliza el enlace público de comprobante de la cotización.
    regenerar=true revoca el enlace anterior (los QR viejos mueren, §19)."""
    sale = await db.sales.find_one({"id": sale_id}, {"_id": 0})
    if not sale or sale.get("tipo_venta") != "cotizacion":
        raise HTTPException(404, "Cotización no encontrada")
    settings = await db.settings.find_one({"_id": "app"}, {"_id": 0}) or {}
    cfg = _qr_cfg(settings)
    if not cfg["activo"]:
        raise HTTPException(400, "El QR de comprobantes está desactivado en Configuración")
    import pago_qr as pq
    if regenerar:
        rev = await db.cot_pago_tokens.update_many(
            {"cotizacion_id": sale_id, "estado": "activo"},
            {"$set": {"estado": "revocado", "revocado_en": iso_now(),
                      "revocado_por": user["name"]}})
        if getattr(rev, "modified_count", 0):
            await log_audit(user, "pago_link_revocado", "cotizacion", sale_id,
                            f"{rev.modified_count} enlace(s) revocado(s)")
    link = await _link_activo(sale_id)
    creado = False
    if not link:
        token = pq.nuevo_token()
        from datetime import timedelta
        link = {
            "id": uid(), "cotizacion_id": sale_id, "folio": sale.get("folio"),
            "token": token, "token_hash": pq.hash_token(token),
            "estado": "activo", "creado_por": user["name"],
            "created_at": iso_now(),
            "expires_at": (now_utc() + timedelta(days=cfg["vigencia_dias"])).isoformat(),
        }
        await db.cot_pago_tokens.insert_one(dict(link))
        creado = True
        await log_audit(user, "pago_link_creado", "cotizacion", sale_id,
                        f"vigencia {cfg['vigencia_dias']} días")
    base = os.environ.get("PUBLIC_BASE_URL", "") or (str(request.base_url).rstrip("/") if request else "")
    url = f"{base.rstrip('/')}/pago/comprobante/{link['token']}" if base else ""
    return {"ok": True, "url": url, "expires_at": link.get("expires_at"),
            "nuevo": creado}


async def _validar_token_publico(token: str) -> dict:
    """Hash-lookup + vigencia. Misma respuesta para inválido/expirado/revocado."""
    import pago_qr as _pq
    tok = await db.cot_pago_tokens.find_one(
        {"token_hash": _pq.hash_token(token)}, {"_id": 0})
    if not tok or tok.get("estado") != "activo":
        raise HTTPException(410, "Enlace no disponible")
    exp = str(tok.get("expires_at") or "")
    if exp and exp < now_utc().isoformat():
        raise HTTPException(410, "Enlace no disponible")
    return tok


@api.get("/public/pago-comprobante/{token}")
async def public_pago_info(request: Request, token: str):
    """Página pública: SOLO datos estrictamente necesarios (§5/§15)."""
    import pago_qr as _pq
    ip = request.client.host if request.client else "?"
    if not _pq.permitir(f"info:{ip}:{token[:6]}", 60, 3600):
        raise HTTPException(429, "Demasiadas consultas; intenta más tarde")
    try:
        tok = await _validar_token_publico(token)
    except HTTPException:
        raise HTTPException(410, "Enlace no disponible")
    sale = await db.sales.find_one({"id": tok["cotizacion_id"]},
                                   {"_id": 0, "folio": 1, "cliente_nombre": 1,
                                    "total": 1, "fecha_vencimiento": 1, "estado": 1})
    if not sale:
        raise HTTPException(410, "Enlace no disponible")
    settings = await db.settings.find_one({"_id": "app"},
                                          {"_id": 0, "empresa_nombre": 1,
                                           "moneda": 1, "whatsapp_empresa": 1}) or {}
    return {
        "folio": sale.get("folio"), "cliente": sale.get("cliente_nombre") or "",
        "importe": float(sale.get("total") or 0),
        "moneda": settings.get("moneda", "MXN"),
        "vence": str(sale.get("fecha_vencimiento") or "")[:10],
        "empresa": settings.get("empresa_nombre", "Grupo RYSA"),
        "whatsapp_empresa": settings.get("whatsapp_empresa", ""),
        "metodos": ["transferencia", "deposito", "tarjeta", "otros"],
    }


@api.post("/public/pago-comprobante/{token}")
async def public_pago_upload(request: Request, token: str,
                             comprobante: UploadFile = File(...),
                             metodo: str = Form("transferencia"),
                             referencia: str = Form(""),
                             comentarios: str = Form("")):
    """Recepción PÚBLICA del comprobante. Queda PENDIENTE de revisión;
    NUNCA registra pago ni toca CxC por sí solo (§8)."""
    import pago_qr as _pq
    ip = request.client.host if request.client else "?"
    if not _pq.permitir(f"up:{ip}:{token[:6]}", 5, 3600):
        raise HTTPException(429, "Límite de envíos alcanzado; intenta más tarde")
    try:
        tok = await _validar_token_publico(token)
    except HTTPException:
        raise HTTPException(410, "Enlace no disponible")
    sale = await db.sales.find_one(
        {"id": tok["cotizacion_id"]},
        {"_id": 0, "id": 1, "folio": 1, "cliente_id": 1, "cliente_nombre": 1,
         "total": 1, "estado": 1})
    if not sale or sale.get("estado") not in ("cotizacion", "convertida"):
        raise HTTPException(410, "Enlace no disponible")
    settings = await db.settings.find_one({"_id": "app"}, {"_id": 0}) or {}
    cfg = _qr_cfg(settings)

    data = await comprobante.read()
    ok, msg = _pq.validar_archivo(comprobante.filename, data, cfg["max_mb"])
    if not ok:
        raise HTTPException(400, msg)
    metodo_n = (metodo or "").strip().lower()
    if metodo_n not in ("transferencia", "deposito", "tarjeta", "otros"):
        raise HTTPException(400, "Método de pago inválido")

    # Anti-duplicados (§25): misma referencia+metodo viva en esta cotización.
    ref = (referencia or "").strip()
    if ref:
        dup = await db.payment_evidence.find_one(
            {"cotizacion_id": sale["id"], "metodo": metodo_n, "referencia": ref,
             "estado": {"$ne": "rechazado"}})
        if dup:
            raise HTTPException(409, "Ya existe un comprobante con esa referencia")

    total_ev = await db.payment_evidence.count_documents(
        {"cotizacion_id": sale["id"], "estado": {"$ne": "rechazado"}})
    if total_ev >= cfg["max_archivos"]:
        raise HTTPException(409, "Ya se recibieron los comprobantes permitidos")

    ext = os.path.splitext(_pq.sanitizar_nombre(comprobante.filename))[1].lower() or ".bin"
    path = f"comprobantes/{sale.get('folio')}-{int(now_utc().timestamp())}-{_pq.nuevo_token()[:8]}{ext}"
    storage.put_object(path, data, comprobante.content_type or "application/octet-stream")
    ev = {
        "id": uid(), "cotizacion_id": sale["id"], "folio_cot": sale.get("folio"),
        "link_id": tok["id"], "storage_path": path,
        "original_filename": _pq.sanitizar_nombre(comprobante.filename),
        "mime_type": storage.detect_mime_type(data), "file_size": len(data),
        "metodo": metodo_n, "referencia": ref,
        "comentarios": (comentarios or "").strip()[:500],
        "estado": "pendiente", "ip": ip,
        "created_at": iso_now(),
        "reviewed_at": "", "reviewed_by": "", "review_comentario": "",
        "abono_folio": "", "abono_id": "",
    }
    await db.payment_evidence.insert_one(dict(ev))
    await log_audit({"id": "publico", "name": f"anon:{ip}"}, "comprobante_recibido",
                    "payment_evidence", ev["id"], f"{sale.get('folio')} · {metodo_n}")
    wa_txt = _pq.mensaje_wa(sale.get("folio", ""), sale.get("cliente_nombre", ""),
                            money_fmt(sale.get("total")))
    return {"ok": True, "mensaje": "Comprobante recibido correctamente.",
            "whatsapp_empresa": settings.get("whatsapp_empresa", ""),
            "wa_texto": wa_txt}


def money_fmt(v) -> str:
    try:
        return f"${float(v):,.2f} MXN"
    except Exception:
        return f"{v} MXN"


@api.get("/sales/{sale_id}/comprobantes")
async def listar_comprobantes(sale_id: str, request: Request = None,
                              user: dict = Depends(require_permission("cxc.abono"))):
    """Historial de evidencias + enlace vigente (§12)."""
    sale = await db.sales.find_one({"id": sale_id}, {"_id": 0})
    if not sale:
        raise HTTPException(404, "Cotización no encontrada")
    import pago_qr as _pq
    link = await _link_activo(sale_id)
    url = ""
    if link and link.get("token"):
        base = os.environ.get("PUBLIC_BASE_URL", "") or (str(request.base_url).rstrip("/") if request else "")
        url = f"{base.rstrip('/')}/pago/comprobante/{link['token']}" if base else ""
    evids = await db.payment_evidence.find(
        {"cotizacion_id": sale_id}, {"_id": 0}).sort("created_at", -1).to_list(200)
    for e in evids:
        e.pop("storage_path", None)  # ruta interna nunca sale del backend
    return {"link": {"url": url, "estado": (link or {}).get("estado", ""),
                     "expires_at": (link or {}).get("expires_at", "")},
            "evidencias": evids}


@api.get("/comprobantes-pago/{evid_id}/archivo")
async def ver_comprobante_archivo(evid_id: str,
                                  user: dict = Depends(require_permission("cxc.abono"))):
    """Descarga/vista del archivo SOLO para personal autorizado."""
    ev = await db.payment_evidence.find_one({"id": evid_id}, {"_id": 0})
    if not ev or not ev.get("storage_path"):
        raise HTTPException(404, "Comprobante no encontrado")
    try:
        data, ctype = storage.get_object(ev["storage_path"])
    except Exception:
        raise HTTPException(404, "Archivo no disponible")
    from fastapi.responses import Response as FastResponse
    return FastResponse(content=data, media_type=ev.get("mime_type") or ctype)


@api.post("/comprobantes-pago/{evid_id}/aprobar")
async def aprobar_comprobante(evid_id: str, data: RevisionInput,
                              user: dict = Depends(require_permission("cxc.abono"))):
    """Aprobación manual (§14): registra el pago vía abonar_pg si la
    cotización ya fue convertida a venta; lock optimista anti-doble (§25)."""
    ev = await db.payment_evidence.find_one({"id": evid_id}, {"_id": 0})
    if not ev:
        raise HTTPException(404, "Comprobante no encontrado")
    if ev["estado"] == "aprobando":
        raise HTTPException(409, "La revisión está en proceso")
    # PGCollection.update_one devuelve el nº de documentos modificados.
    locked = await db.payment_evidence.update_one(
        {"id": evid_id, "estado": "pendiente"},
        {"$set": {"estado": "aprobando"}})
    if not locked:
        raise HTTPException(409, "Este comprobante ya fue procesado")

    def _revertir():
        db_sync = db.payment_evidence.update_one(
            {"id": evid_id, "estado": "aprobando"},
            {"$set": {"estado": "pendiente"}})

    sale = await db.sales.find_one({"id": ev["cotizacion_id"]}, {"_id": 0}) or {}
    abono_info = ""
    try:
        if sale.get("convertida_a") and sale.get("cliente_id"):
            caja = await caja_abierta_de(user["id"])
            folio_ab = await next_counter("abono", "AB", 6)
            metodo = ev["metodo"] if ev["metodo"] in ("efectivo", "tarjeta", "transferencia", "deposito") else "otros"
            res = await _pgcxc.abonar_pg(
                client_id=sale["cliente_id"], monto=round(float(sale.get("total") or 0), 2),
                metodo=metodo, referencia=f"EVID-{ev['id'][:8]} {ev.get('referencia','')}".strip(),
                nota=f"Comprobante QR cotización {ev.get('folio_cot','')}",
                user=user, caja=caja, folio=folio_ab)
            abono_info = res.get("abono", {}).get("folio", folio_ab)
    except Exception as e:
        await _revertir()
        raise HTTPException(400, f"No se pudo registrar el pago: {str(e)[:140]}")

    upd = await db.payment_evidence.update_one(
        {"id": evid_id, "estado": "aprobando"},
        {"$set": {"estado": "aprobado", "reviewed_by": user["name"],
                  "reviewed_at": iso_now(), "review_comentario": data.comentario[:400],
                  "abono_folio": abono_info}})
    if not upd:  # carrera extrema: verificar estado manualmente
        raise HTTPException(409, "Conflicto de concurrencia; verifica el estado")
    await log_audit(user, "comprobante_aprobado", "payment_evidence", evid_id,
                    f"{ev.get('folio_cot')} · abono {abono_info or 'N/A (sin venta convertida)'}")
    return {"ok": True, "abono_folio": abono_info}


@api.post("/comprobantes-pago/{evid_id}/rechazar")
async def rechazar_comprobante(evid_id: str, data: RevisionInput,
                               user: dict = Depends(require_permission("cxc.abono"))):
    upd = await db.payment_evidence.update_one(
        {"id": evid_id, "estado": "pendiente"},
        {"$set": {"estado": "rechazado", "reviewed_by": user["name"],
                  "reviewed_at": iso_now(), "review_comentario": data.comentario[:400]}})
    if not upd:
        raise HTTPException(409, "Este comprobante ya fue procesado o está en proceso")
    await log_audit(user, "comprobante_rechazado", "payment_evidence", evid_id,
                    data.comentario[:120])
    return {"ok": True}


# =========================================================================
# PEDIDOS (CRUD + estados + conversión a venta)
# =========================================================================
@api.get("/pedidos")
async def pedidos_list(estado: Optional[str] = None, cliente_id: Optional[str] = None,
                       vendedor_id: Optional[str] = None, q: Optional[str] = None,
                       user: dict = Depends(get_current_user)):
    flt = {}
    if estado and estado != "todos":
        flt["estado"] = estado
    if cliente_id:
        flt["cliente_id"] = cliente_id
    if vendedor_id:
        flt["vendedor_id"] = vendedor_id
    docs = await db.pedidos.find(flt, {"_id": 0}).sort("fecha_pedido", -1).to_list(100000)
    if q:
        ql = q.lower().strip()
        docs = [d for d in docs if ql in " ".join(str(d.get(k) or "") for k in ("folio", "cliente_nombre", "vendedor_nombre")).lower()]
    return docs


@api.get("/pedidos/{ped_id}")
async def pedido_detail(ped_id: str, user: dict = Depends(get_current_user)):
    doc = await db.pedidos.find_one({"id": ped_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Pedido no encontrado")
    return doc


@api.post("/pedidos")
async def pedido_create(data: PedidoInput, user: dict = Depends(require_permission("pedido.gestionar"))):
    if not data.items:
        raise HTTPException(400, "Agrega al menos un producto al pedido")
    cliente = await db.clients.find_one({"id": data.cliente_id}, {"_id": 0}) if data.cliente_id else None
    cliente_nombre = cliente["nombre"] if cliente else "Público General"
    items = []
    subtotal = 0.0
    for it in data.items:
        if float(it.solicitado) <= 0:
            continue
        precio = float(it.precio or 0)
        if precio <= 0 and it.product_id:
            p = await db.products.find_one({"id": it.product_id}, {"_id": 0})
            if p:
                precio = float(p.get("precio") or (p.get("precios") or [{}])[0].get("precio_con_iva") or 0)
        importe = round(float(it.solicitado) * precio, 2)
        subtotal += importe
        items.append({
            "product_id": it.product_id, "codigo": it.codigo, "descripcion": it.descripcion,
            "unidad": it.unidad or "PZA", "solicitado": float(it.solicitado),
            "surtido": 0.0, "pendiente": float(it.solicitado),
            "precio": round(precio, 2), "iva_tasa": float(it.iva_tasa or 8), "importe": importe,
        })
    if not items:
        raise HTTPException(400, "No hay cantidades válidas en el pedido")
    folio = await next_counter("pedido", "PDO", 6)
    iva = round(sum(i["importe"] * (i["iva_tasa"] / 100) for i in items), 2)
    doc = {
        "id": uid(), "folio": folio, "cliente_id": data.cliente_id, "cliente_nombre": cliente_nombre,
        "vendedor_id": data.vendedor_id or user["id"], "vendedor_nombre": user["name"],
        "fecha_pedido": data.fecha_pedido or iso_now()[:10], "fecha_entrega": data.fecha_entrega,
        "notas": data.notas or "", "items": items,
        "subtotal": round(subtotal, 2), "iva": iva, "total": round(subtotal + iva, 2),
        "estado": "borrador", "sucursal_id": data.sucursal_id or user.get("sucursal_id"),
        "usuario_id": user["id"], "usuario_nombre": user["name"],
        "creado_en": iso_now(), "actualizado_en": iso_now(),
    }
    await db.pedidos.insert_one(doc)
    await log_audit(user, "pedido_crear", "pedido", doc["id"], f"{folio} {cliente_nombre}")
    return doc


@api.put("/pedidos/{ped_id}")
async def pedido_update(ped_id: str, data: PedidoInput, user: dict = Depends(require_permission("pedido.gestionar"))):
    ex = await db.pedidos.find_one({"id": ped_id})
    if not ex:
        raise HTTPException(404, "Pedido no encontrado")
    if ex.get("estado") in ("convertido", "cancelado"):
        raise HTTPException(409, f"Un pedido {ex.get('estado')} no se puede editar")
    cliente = await db.clients.find_one({"id": data.cliente_id}, {"_id": 0}) if data.cliente_id else None
    cliente_nombre = cliente["nombre"] if cliente else ex.get("cliente_nombre") or "Público General"
    items = []
    subtotal = 0.0
    for it in data.items:
        if float(it.solicitado) <= 0:
            continue
        prev = next((x for x in ex.get("items", []) if x.get("product_id") == it.product_id), None)
        surtido = float((prev or {}).get("surtido", 0) or 0)
        if float(it.solicitado) < surtido:
            raise HTTPException(400, f"{it.descripcion}: no puedes bajar la cantidad por debajo de lo ya surtido ({surtido})")
        precio = float(it.precio or (prev or {}).get("precio") or 0)
        importe = round(float(it.solicitado) * precio, 2)
        subtotal += importe
        items.append({
            "product_id": it.product_id, "codigo": it.codigo, "descripcion": it.descripcion,
            "unidad": it.unidad or "PZA", "solicitado": float(it.solicitado),
            "surtido": surtido, "pendiente": round(float(it.solicitado) - surtido, 3),
            "precio": round(precio, 2), "iva_tasa": float(it.iva_tasa or 8), "importe": importe,
        })
    if not items:
        raise HTTPException(400, "No hay cantidades válidas en el pedido")
    iva = round(sum(i["importe"] * (i["iva_tasa"] / 100) for i in items), 2)
    up = {**ex,
          "cliente_id": data.cliente_id, "cliente_nombre": cliente_nombre,
          "vendedor_id": data.vendedor_id or ex.get("vendedor_id"),
          "fecha_pedido": data.fecha_pedido or ex.get("fecha_pedido"),
          "fecha_entrega": data.fecha_entrega, "notas": data.notas or "",
          "items": items, "subtotal": round(subtotal, 2), "iva": iva,
          "total": round(subtotal + iva, 2),
          "sucursal_id": data.sucursal_id or ex.get("sucursal_id"),
          "actualizado_en": iso_now()}
    await db.pedidos.update_one({"id": ped_id}, {"$set": up})
    await log_audit(user, "pedido_editar", "pedido", ped_id, ex.get("folio"))
    return up


@api.post("/pedidos/{ped_id}/estado")
async def pedido_estado(ped_id: str, data: PedidoEstadoInput, user: dict = Depends(require_permission("pedido.gestionar"))):
    ex = await db.pedidos.find_one({"id": ped_id})
    if not ex:
        raise HTTPException(404, "Pedido no encontrado")
    if data.estado not in ("borrador", "confirmado", "surtido", "cancelado"):
        raise HTTPException(400, "Estado no válido")
    if ex.get("estado") == "convertido":
        raise HTTPException(409, "Un pedido convertido a venta no puede cambiar de estado")
    await db.pedidos.update_one({"id": ped_id},
                                {"$set": {"estado": data.estado, "actualizado_en": iso_now()}})
    await log_audit(user, "pedido_estado", "pedido", ped_id, f"{ex.get('folio')} -> {data.estado}")
    return {"ok": True, "estado": data.estado}


@api.post("/pedidos/{ped_id}/convertir")
async def pedido_convertir(ped_id: str, data: PedidoConvertInput,
                           user: dict = Depends(require_permission("venta.crear"))):
    """Convierte un pedido en venta real (usa la pipeline completa `_crear_venta`)."""
    ped = await db.pedidos.find_one({"id": ped_id}, {"_id": 0})
    if not ped:
        raise HTTPException(404, "Pedido no encontrado")
    if ped.get("estado") in ("convertido", "cancelado"):
        raise HTTPException(409, f"Un pedido {ped.get('estado')} no se puede convertir")
    items = []
    for it in ped.get("items", []):
        cantidad = float(it.get("surtido") or 0) or float(it.get("solicitado") or 0)
        if cantidad <= 0:
            continue
        items.append(SaleItem(
            product_id=it.get("product_id"), codigo=it.get("codigo") or "",
            descripcion=it.get("descripcion") or "", cantidad=cantidad,
            unidad=it.get("unidad") or "PZA", precio=float(it.get("precio") or 0),
            iva_tasa=float(it.get("iva_tasa") or 8), descuento=0.0))
    if not items:
        raise HTTPException(400, "El pedido no tiene cantidades para surtir")
    payload = SaleInput(
        cliente_id=ped.get("cliente_id"), items=items, descuento_global=0.0,
        condicion=data.condicion, pagos=data.pagos, lista_precios=1,
        tipo_venta="directa", vendedor_id=data.vendedor_id or ped.get("vendedor_id"),
        precios_incluyen_iva=True)
    sale = await _crear_venta(user, payload)
    await db.pedidos.update_one({"id": ped_id}, {"$set": {
        "estado": "convertido", "convertida_a": sale["id"], "convertida_folio": sale["folio"],
        "actualizado_en": iso_now()}})
    await log_audit(user, "pedido_convertir", "pedido", ped_id,
                    f"{ped.get('folio')} -> venta {sale['folio']}")
    return sale


@api.post("/recargas")
async def crear_recarga(data: RecargaInput, user: dict = Depends(require_permission("venta.crear"))):
    if data.monto <= 0:
        raise HTTPException(400, "El monto debe ser mayor a 0")
    if not data.telefono.strip():
        raise HTTPException(400, "Captura el número de teléfono")
    now = now_utc()
    folio = await next_counter("recarga", "R", 6)
    descripcion = f"Recarga {data.compania} · {data.telefono}".strip()
    total = round(float(data.monto), 2)
    items = [{"product_id": None, "codigo": "RECARGA", "descripcion": descripcion, "cantidad": 1,
              "unidad": "SERV", "precio": total, "iva_tasa": 0, "descuento": 0, "importe": total}]
    # Caja obligatoria para recargas (cobro en mostrador).
    caja = await caja_abierta_de(user["id"])
    if not caja:
        raise HTTPException(409, "No hay caja abierta. Abre una caja antes de registrar la recarga.")
    sale = {
        "id": uid(), "folio": folio, "fecha": iso_now(), "hora": now.strftime("%H:%M"),
        "usuario_id": user["id"], "usuario_nombre": user["name"],
        "vendedor_id": user["id"], "vendedor_nombre": user["name"],
        "cliente_id": None, "cliente_nombre": "Público General",
        "items": items, "subtotal": total, "iva_total": 0.0, "descuento_global": 0.0, "total": total,
        "tipo_venta": "recarga", "condicion": "contado",
        "pagos": [{"metodo": data.metodo, "monto": total}], "cambio": 0.0, "saldo": 0.0,
        "estado": "confirmada", "factura": False, "caja_id": caja["id"] if caja else None,
        "sucursal_id": caja.get("sucursal_id") or user.get("sucursal_id"),
        "lista_precios": 1, "compania": data.compania, "telefono": data.telefono,
        "referencia_tae": (data.referencia_tae or "").strip(),
        "comision": round(float(data.comision or 0), 2),
    }
    await db.sales.insert_one(sale)
    if caja and data.metodo == "efectivo":
        await db.caja_movimientos.insert_one({
            "id": uid(), "caja_id": caja["id"], "tipo": "venta", "concepto": f"Recarga {folio}",
            "monto": total, "referencia": folio, "usuario_id": user["id"],
            "usuario_nombre": user["name"], "fecha": iso_now()})
    await log_audit(user, "crear", "recarga", sale["id"], f"{folio} {descripcion} {total}")
    return await db.sales.find_one({"id": sale["id"]}, {"_id": 0})

@api.post("/sales/{sale_id}/cancelar")
async def cancel_sale(sale_id: str, data: CancelInput, user: dict = Depends(require_permission("venta.cancelar"))):
    sale = await db.sales.find_one({"id": sale_id}, {"_id": 0})
    if not sale:
        raise HTTPException(404, "Venta no encontrada")
    if not ver_todas_ventas(user) and sale.get("vendedor_id") != user["id"]:
        raise HTTPException(403, "No tienes permiso para cancelar esta venta")
    if not (data.motivo or "").strip():
        raise HTTPException(422, "El motivo de cancelación es obligatorio")
    try:
        return await _pgpos.cancela_venta_pg(user=user, sale_id=sale_id,
                                             motivo=(data.motivo or "").strip())
    except _pgpos.VentaError as e:
        raise HTTPException(status_code=e.status, detail=e.message)

# Ventas suspendidas
@api.post("/sales/suspend")
async def suspend_sale(data: SaleInput, user: dict = Depends(require_permission("venta.crear"))):
    doc = {"id": uid(), "usuario_id": user["id"], "fecha": iso_now(),
           "payload": data.model_dump(), "estado": "suspendida"}
    await db.suspended_sales.insert_one(doc)
    return {"id": doc["id"]}

@api.get("/sales-suspended")
async def list_suspended(user: dict = Depends(get_current_user)):
    return await db.suspended_sales.find({"usuario_id": user["id"]}, {"_id": 0}).sort("fecha", -1).to_list(100)

@api.delete("/sales-suspended/{sid}")
async def delete_suspended(sid: str, user: dict = Depends(require_permission("venta.crear"))):
    await db.suspended_sales.delete_one({"id": sid})
    return {"ok": True}

# =========================================================================
# CUENTAS POR COBRAR (CxC)
# =========================================================================
AGING_KEYS = ["corriente", "b1_30", "b31_60", "b61_90", "b90"]

def _parse_date(s):
    try:
        return date.fromisoformat(str(s)[:10])
    except Exception:
        return None

def _dias_vencido(fecha_iso, dias_credito, hoy):
    d = _parse_date(fecha_iso)
    if not d:
        return 0, None
    vence = d + timedelta(days=int(dias_credito or 0))
    return (hoy - vence).days, vence.isoformat()

def _bucket(dv):
    if dv <= 0:
        return "corriente"
    if dv <= 30:
        return "b1_30"
    if dv <= 60:
        return "b31_60"
    if dv <= 90:
        return "b61_90"
    return "b90"

@api.get("/cxc")
async def cxc_list(q: Optional[str] = None, solo_vencidos: Optional[bool] = False,
                   estado: Optional[str] = None, vendedor_id: Optional[str] = None,
                   facturada: Optional[str] = None,
                   user: dict = Depends(require_permission("cxc.ver"))):
    hoy = now_utc().date()
    clientes = await db.clients.find({"saldo": {"$gt": 0}}, {"_id": 0}).to_list(20000)
    cmap = {c["id"]: c for c in clientes}
    sf = {"condicion": "credito", "estado": "confirmada", "saldo": {"$gt": 0}}
    if vendedor_id:
        sf["vendedor_id"] = vendedor_id
    sales = await db.sales.find(sf, {"_id": 0, "cliente_id": 1, "fecha": 1, "saldo": 1,
                                     "total": 1, "vendedor_id": 1, "facturado": 1, "folio": 1}).to_list(100000)
    agg = {}
    for s in sales:
        cid = s.get("cliente_id")
        if not cid or cid not in cmap:
            continue
        cli = cmap[cid]
        dv, _ = _dias_vencido(s["fecha"], cli.get("dias_credito", 0), hoy)
        a = agg.get(cid)
        if not a:
            a = {k: 0.0 for k in AGING_KEYS}
            a.update({"vencido": 0.0, "max_dias": 0, "n": 0, "monto_original": 0.0, "con_abonos": 0, "sin_abonos": 0, "facturadas": 0, "no_facturadas": 0, "folios": []})
            agg[cid] = a
        a[_bucket(dv)] += s["saldo"]
        a["monto_original"] += float(s.get("total", 0))
        a["n"] += 1
        a["folios"].append(s.get("folio", ""))
        if float(s.get("saldo", 0)) < float(s.get("total", 0)) - 0.01:
            a["con_abonos"] += 1
        else:
            a["sin_abonos"] += 1
        if s.get("facturado"):
            a["facturadas"] += 1
        else:
            a["no_facturadas"] += 1
        if dv > 0:
            a["vencido"] += s["saldo"]
            a["max_dias"] = max(a["max_dias"], dv)
    rows = []
    ql = (q or "").lower().strip()
    for cid, cli in cmap.items():
        a = agg.get(cid)
        if not a:
            # Cliente con saldo pero sin ventas a crédito individuales
            # (p. ej. saldo inicial importado). Se exhibe como vigente.
            a = {k: 0.0 for k in AGING_KEYS}
            a["corriente"] = round(float(cli.get("saldo", 0) or 0), 2)
            a.update({"vencido": 0.0, "max_dias": 0, "n": 0,
                      "monto_original": round(float(cli.get("saldo", 0) or 0), 2),
                      "con_abonos": 0, "sin_abonos": 0,
                      "facturadas": 0, "no_facturadas": 0, "folios": []})
        pendiente_total = sum(a[k] for k in AGING_KEYS)
        abonado_parcial = a["monto_original"] > pendiente_total
        if a["vencido"] > 0:
            st = "vencida"
        elif pendiente_total <= 0:
            st = "liquidada"
        elif abonado_parcial:
            st = "parcialmente_pagada"
        else:
            st = "pendiente"
        item = {
            "cliente_id": cid, "codigo": cli.get("codigo"), "nombre": cli.get("nombre"),
            "telefono": cli.get("telefono"), "celular": cli.get("celular"),
            "limite_credito": round(float(cli.get("limite_credito", 0)), 2),
            "dias_credito": cli.get("dias_credito", 0),
            "saldo": round(float(cli.get("saldo", 0)), 2),
            "monto_original": round(a["monto_original"], 2),
            "vencido": round(a["vencido"], 2),
            "max_dias": a["max_dias"],
            "ventas_pendientes": a["n"],
            "estado": st,
            "con_abonos": a["con_abonos"], "sin_abonos": a["sin_abonos"],
            "facturadas": a["facturadas"], "no_facturadas": a["no_facturadas"],
            "aging": {k: round(a[k], 2) for k in AGING_KEYS},
        }
        if solo_vencidos and item["vencido"] <= 0:
            continue
        if estado and estado != "todos" and st != estado:
            continue
        if facturada == "si" and item["facturadas"] == 0:
            continue
        if facturada == "no" and item["no_facturadas"] == 0:
            continue
        if ql:
            haystack = " ".join(str(v or "") for v in [
                cli.get("nombre", ""), cli.get("codigo", ""), cli.get("rfc", ""),
                cli.get("telefono", ""), cli.get("celular", ""),
            ]).lower()
            if ql not in haystack:
                continue
        rows.append(item)
    rows.sort(key=lambda r: r["vencido"] * 1e9 + r["saldo"], reverse=True)
    tot = {"cartera": round(sum(r["saldo"] for r in rows), 2),
           "vencido": round(sum(r["vencido"] for r in rows), 2),
           "clientes": len(rows)}
    tot["por_vencer"] = round(tot["cartera"] - tot["vencido"], 2)
    for k in AGING_KEYS:
        tot[k] = round(sum(r["aging"][k] for r in rows), 2)
    return {"totales": tot, "clientes": rows}

@api.get("/cxc/{client_id}")
async def cxc_detail(client_id: str, user: dict = Depends(require_permission("cxc.ver"))):
    cli = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not cli:
        raise HTTPException(404, "Cliente no encontrado")
    hoy = now_utc().date()
    sales = await db.sales.find({"cliente_id": client_id, "condicion": "credito", "estado": "confirmada"},
                                {"_id": 0}).sort("fecha", 1).to_list(20000)
    ventas = []
    for s in sales:
        dv, vence = _dias_vencido(s["fecha"], cli.get("dias_credito", 0), hoy)
        saldo = round(float(s.get("saldo", 0)), 2)
        ventas.append({"id": s["id"], "folio": s["folio"], "fecha": s["fecha"], "total": s["total"],
                       "saldo": saldo, "vence": vence, "dias_vencido": max(dv, 0),
                       "pagada": saldo <= 0.001})
    abonos = await db.abonos.find({"cliente_id": client_id}, {"_id": 0}).sort("fecha", -1).to_list(5000)
    return {
        "cliente": {"id": cli["id"], "codigo": cli.get("codigo"), "nombre": cli.get("nombre"),
                    "telefono": cli.get("telefono"), "celular": cli.get("celular"),
                    "limite_credito": round(float(cli.get("limite_credito", 0)), 2),
                    "dias_credito": cli.get("dias_credito", 0),
                    "saldo": round(float(cli.get("saldo", 0)), 2)},
        "ventas": ventas, "abonos": abonos,
    }

@api.get("/cxc/{client_id}/adeudo-pdf")
async def cxc_adeudo_pdf(client_id: str, user: dict = Depends(require_permission("cxc.ver"))):
    """PDF detallado de adeudo para imprimir y entregar al cliente."""
    cli = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not cli:
        raise HTTPException(404, "Cliente no encontrado")
    hoy = now_utc().date()
    sales = await db.sales.find({"cliente_id": client_id, "condicion": "credito", "estado": "confirmada"},
                                {"_id": 0}).sort("fecha", 1).to_list(20000)
    abonos = await db.abonos.find({"cliente_id": client_id}, {"_id": 0}).to_list(5000)
    total_vendido = round(sum(float(s.get("total", 0)) for s in sales), 2)
    total_abonos = round(sum(float(a.get("monto", 0)) for a in abonos), 2)
    saldo = round(float(cli.get("saldo", 0)), 2)

    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from io import BytesIO

    buf = BytesIO()
    st = getSampleStyleSheet()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    flow = []
    flow.append(Paragraph(f"<b>ESTADO DE CUENTA / ADEUDO</b>", st['Title']))
    flow.append(Spacer(1, 4 * mm))
    flow.append(Paragraph(f"<b>Cliente:</b> {cli.get('nombre')} ({cli.get('codigo')})", st['Normal']))
    flow.append(Paragraph(f"<b>RFC:</b> {cli.get('rfc') or '—'} &nbsp;&nbsp; <b>Fecha:</b> {hoy.isoformat()}", st['Normal']))
    flow.append(Spacer(1, 3 * mm))
    rows = [["Venta", "Fecha", "Producto", "Descripción", "Cant", "Precio", "Total"]]
    total_items = 0.0
    for s in sales:
        for it in s.get("items", []):
            cant = it.get("cantidad", 0); precio = float(it.get("precio", 0))
            imp = float(it.get("importe", cant * precio))
            total_items += imp
            rows.append([s.get("folio", ""), (s.get("fecha", "") or "")[:10],
                         str(it.get("codigo", "") or ""), str(it.get("descripcion", ""))[:28],
                         cant, f"${precio:,.2f}", f"${imp:,.2f}"])
    t = Table(rows, repeatRows=1)
    t.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#B95A3A')),
                           ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                           ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                           ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
                           ('FONTSIZE', (0, 0), (-1, -1), 7.5)]))
    flow.append(t)
    flow.append(Spacer(1, 3 * mm))
    flow.append(Paragraph(f"<b>Detalle vendido:</b> ${total_items:,.2f}", st['Normal']))
    flow.append(Paragraph(f"<b>Abonos:</b> ${total_abonos:,.2f}", st['Normal']))
    flow.append(Paragraph(f"<b>Saldo pendiente:</b> ${saldo:,.2f}", st['Heading3']))
    doc.build(flow)
    buf.seek(0)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'inline; filename="{cli.get("codigo")}-adeudo.pdf"'})

@api.post("/cxc/{client_id}/recordatorio")
async def cxc_recordatorio(client_id: str, user: dict = Depends(require_permission("cxc.ver"))):
    """Genera el recordatorio de saldo pendiente (WhatsApp/correo) con plantilla."""
    cli = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not cli:
        raise HTTPException(404, "Cliente no encontrado")
    saldo = round(float(cli.get("saldo", 0)), 2)
    if saldo <= 0:
        raise HTTPException(400, "El cliente no tiene saldo pendiente")
    wt = (await db.settings.find_one({"_id": "app"}, {"_id": 0}) or {}).get("ticket_config", {}) or {}
    plantilla = (await db.settings.find_one({"_id": "app"}, {"_id": 0}) or {}).get("plantilla_recordatorio") or \
        "Estimado {cliente}, le recordamos que tiene un saldo pendiente de ${saldo}. Si ya lo cubrió, ignore este mensaje."
    texto = plantilla.replace("{cliente}", cli.get("nombre", "")).replace("{saldo}", f"{saldo:,.2f}")
    telefono = (cli.get("whatsapp") or cli.get("celular") or cli.get("telefono") or "").strip()
    digits = "".join(ch for ch in telefono if ch.isdigit())
    wa_url = ""
    if len(digits) >= 10:
        phone = digits if len(digits) == 12 else ("52" + digits if len(digits) == 10 else digits)
        wa_url = f"https://wa.me/{phone}?text={__import__('urllib.parse', fromlist=['quote']).quote(texto)}"
    # Registrar historial de mensaje
    msg = {"id": uid(), "cliente_id": client_id, "cliente_nombre": cli.get("nombre", ""),
           "canal": "whatsapp", "tipo": "recordatorio_cxc", "contenido": texto,
           "destinatario": telefono, "estado": "generado", "usuario_id": user["id"],
           "usuario_nombre": user["name"], "fecha": iso_now(), "wa_url": wa_url}
    await db.mensajes.insert_one(msg)
    await log_audit(user, "recordatorio_cxc", "cliente", client_id,
                    f"Recordatorio saldo {saldo}")
    return {"texto": texto, "telefono": telefono, "wa_url": wa_url, "historico_id": msg["id"]}

@api.post("/cxc/{client_id}/abono")
async def cxc_abono(client_id: str, data: AbonoInput, user: dict = Depends(require_permission("cxc.abono"))):
    """Abono FIFO ATÓMICO (pgstore.cxc.abonar_pg): cliente y ventas se bloquean
    con FOR UPDATE en una sola transacción; imposible dejar saldo negativo o
    aplicaciones parciales."""
    cli = await db.clients.find_one({"id": client_id})
    if not cli:
        raise HTTPException(404, "Cliente no encontrado")
    monto = round(float(data.monto), 2)
    if monto <= 0:
        raise HTTPException(400, "El monto debe ser mayor a cero")
    caja = await caja_abierta_de(user["id"])
    folio = await next_counter("abono", "AB", 6)
    try:
        return await _pgcxc.abonar_pg(
            client_id=client_id, monto=monto, metodo=data.metodo,
            referencia=data.referencia or "", nota=data.nota or "",
            user=user, caja=caja, folio=folio)
    except _pgcxc.CxcError as e:
        raise HTTPException(e.status, e.message)

# =========================================================================
# ABONOS: historial, comprobante PDF y cancelación auditada
# =========================================================================
@api.get("/abonos")
async def abonos_list(desde: Optional[str] = None, hasta: Optional[str] = None,
                      cliente: Optional[str] = None, usuario: Optional[str] = None,
                      metodo: Optional[str] = None, folio: Optional[str] = None,
                      q: Optional[str] = None,
                      user: dict = Depends(get_current_user)):
    docs = await db.abonos.find({}, {"_id": 0}).to_list(100000)
    if desde:
        docs = [d for d in docs if (d.get("fecha") or "")[:10] >= desde]
    if hasta:
        docs = [d for d in docs if (d.get("fecha") or "")[:10] <= hasta]
    if cliente:
        docs = [d for d in docs if d.get("cliente_id") == cliente or cliente.lower() in (d.get("cliente_nombre") or "").lower()]
    if usuario:
        docs = [d for d in docs if d.get("usuario_id") == usuario]
    if metodo and metodo != "todos":
        docs = [d for d in docs if d.get("metodo") == metodo]
    if folio:
        docs = [d for d in docs if folio.lower() in (d.get("folio") or "").lower()]
    if q:
        ql = q.lower().strip()
        docs = [d for d in docs if ql in " ".join(str(d.get(k) or "") for k in ("folio", "cliente_nombre", "referencia")).lower()]
    docs.sort(key=lambda d: d.get("fecha") or "", reverse=True)
    return docs

@api.post("/abonos/{abono_id}/pdf")
async def abono_pdf(abono_id: str, user: dict = Depends(get_current_user)):
    abono = await db.abonos.find_one({"id": abono_id}, {"_id": 0})
    if not abono:
        raise HTTPException(404, "Abono no encontrado")
    settings = await db.settings.find_one({"_id": "app"}, {"_id": 0}) or {}
    cliente = None
    if abono.get("cliente_id"):
        c = await db.clients.find_one({"id": abono["cliente_id"]}, {"_id": 0})
        if c:
            cliente = {"nombre": c.get("nombre"), "rfc": c.get("rfc"),
                       "telefono": c.get("telefono") or c.get("celular") or c.get("whatsapp"),
                       "correo": c.get("correo") or c.get("correos")}
    try:
        pdf_bytes = storage.build_abono_pdf(abono, settings, cliente)
        folio_clean = "".join(c for c in abono.get('folio', 'abono') if c.isalnum())
        path = f"abonos/{folio_clean}-{uid()[:8]}.pdf"
        result = storage.put_object(path, pdf_bytes, "application/pdf")
    except Exception as e:
        logger.error("Comprobante de abono PDF falló: %s", str(e)[:200])
        raise HTTPException(502, "No se pudo generar el comprobante de abono.")
    stored = result.get("path", path)
    await db.files.insert_one({
        "id": uid(), "storage_path": stored,
        "original_filename": f"RYSA_Comprobante_Abono_{abono.get('folio')}.pdf",
        "content_type": "application/pdf", "size": result.get("size", len(pdf_bytes)),
        "abono_id": abono_id, "cliente_id": abono.get("cliente_id"),
        "is_deleted": False, "created_at": iso_now()})
    return {"path": stored, "url": f"/api/files/{stored}", "filename": f"RYSA_Comprobante_Abono_{abono.get('folio')}.pdf"}

@api.post("/abonos/{abono_id}/cancelar")
async def abono_cancelar(abono_id: str, motivo: str = "Cancelación",
                         user: dict = Depends(require_permission("caja.entrada"))):
    """Cancela un abono confirmado de forma ATÓMICA (pgstore.cxc.cancelar_abono_pg):
    recompone saldos de ventas y cliente en una sola transacción. Requiere el
    mismo permiso que crear un abono (antes cualquier usuario autenticado
    podía cancelarlo)."""
    try:
        return await _pgcxc.cancelar_abono_pg(abono_id=abono_id,
                                              motivo=motivo or "Cancelación", user=user)
    except _pgcxc.CxcError as e:
        raise HTTPException(e.status, e.message)

# =========================================================================
# PROVEEDORES
# =========================================================================
@api.get("/proveedores")
async def proveedores_list(q: Optional[str] = None,
                           user: dict = Depends(get_current_user)):
    flt = {}
    docs = await db.proveedores.find(flt, {"_id": 0}).to_list(20000)
    if q:
        ql = q.lower().strip()
        docs = [d for d in docs if ql in " ".join(str(d.get(k) or "") for k in
                ("nombre", "razon_social", "rfc", "contacto", "telefono", "email")).lower()]
    docs.sort(key=lambda d: (d.get("nombre") or "").lower())
    return docs

@api.post("/proveedores")
async def proveedor_create(data: ProveedorInput, user: dict = Depends(require_permission("proveedor.crear"))):
    pid = uid()
    codigo = f"PRV{pid[:6].upper()}"
    doc = {"id": pid, "codigo": codigo, **data.model_dump(),
           "estado": "activo" if data.activo else "inactivo",
           "created_at": iso_now(), "updated_at": iso_now(),
           "usuario_id": user["id"], "usuario_nombre": user["name"]}
    await db.proveedores.insert_one(doc)
    await log_audit(user, "proveedor_crear", "proveedor", pid, data.nombre)
    return {k: doc[k] for k in doc if k != "usuario_id"}

@api.put("/proveedores/{proveedor_id}")
async def proveedor_update(proveedor_id: str, data: ProveedorInput,
                           user: dict = Depends(require_permission("proveedor.editar"))):
    ex = await db.proveedores.find_one({"id": proveedor_id})
    if not ex:
        raise HTTPException(404, "Proveedor no encontrado")
    up = data.model_dump()
    up.update({"id": proveedor_id, "codigo": ex.get("codigo"),
               "estado": "activo" if data.activo else "inactivo",
               "updated_at": iso_now()})
    await db.proveedores.update_one({"id": proveedor_id}, {"$set": up})
    await log_audit(user, "proveedor_editar", "proveedor", proveedor_id, data.nombre)
    return {k: up[k] for k in up if k != "usuario_id"}

@api.patch("/proveedores/{proveedor_id}/estado")
async def proveedor_estado(proveedor_id: str, activo: bool,
                           user: dict = Depends(require_permission("proveedor.editar"))):
    await db.proveedores.update_one({"id": proveedor_id},
                                    {"$set": {"activo": activo, "estado": "activo" if activo else "inactivo",
                                              "updated_at": iso_now()}})
    await log_audit(user, "proveedor_estado", "proveedor", proveedor_id,
                    "activo" if activo else "inactivo")
    return {"ok": True}

@api.get("/proveedores/{proveedor_id}/ficha")
async def proveedor_ficha(proveedor_id: str, user: dict = Depends(get_current_user)):
    """Ficha del proveedor: datos + compras/facturas/pendientes/historial."""
    prov = await db.proveedores.find_one({"id": proveedor_id}, {"_id": 0})
    if not prov:
        raise HTTPException(404, "Proveedor no encontrado")
    compras = await db.compras.find({"proveedor_id": proveedor_id, "estado": "confirmada"},
                                    {"_id": 0}).to_list(20000)
    total_compras = round(sum(float(c.get("total", 0) or 0) for c in compras), 2)
    pendiente = round(sum(float(c.get("saldo_pendiente", 0) or 0) for c in compras), 2)
    facturas = [c for c in compras if c.get("factura_numero")]
    ultima = None
    if compras:
        ultima = sorted(compras, key=lambda c: c.get("fecha_recepcion") or "", reverse=True)[0]
    # Costos por producto de este proveedor.
    costos = await db.costos_historial.find({"proveedor_id": proveedor_id}, {"_id": 0}).to_list(5000)
    por_producto = {}
    for h in costos:
        e = por_producto.setdefault(h["product_id"], {"product_id": h["product_id"],
                                                      "codigo": h.get("codigo"),
                                                      "descripcion": h.get("descripcion"),
                                                      "historial": []})
        e["historial"].append({"fecha": h.get("fecha"), "factura": h.get("factura"),
                               "compra": h.get("folio"), "cantidad": h.get("cantidad"),
                               "costo": h.get("costo")})
    for e in por_producto.values():
        e["historial"].sort(key=lambda x: x.get("fecha") or "", reverse=True)
        e["ultimo_costo"] = e["historial"][0]["costo"] if e["historial"] else None
    return {"proveedor": prov, "resumen": {
                "compras_total": total_compras, "facturas": len(facturas),
                "pendiente": pendiente,
                "ultima_compra": str(ultima.get("fecha_recepcion") or "")[:10] if ultima else "",
            },
            "compras": sorted(compras, key=lambda c: c.get("fecha_recepcion") or "", reverse=True)[:50],
            "productos": list(por_producto.values())}

# =========================================================================
# CUENTAS BANCARIAS
# =========================================================================
@api.get("/cuentas-bancarias")
async def cuentas_bancarias_list(user: dict = Depends(get_current_user)):
    docs = await db.cuentas_bancarias.find({}, {"_id": 0}).to_list(5000)
    docs.sort(key=lambda d: (0 if d.get("predeterminada") else 1, (d.get("banco") or "").lower()))
    return docs


# --- Catálogo de bancos + logos (selector UI y PDF de cotizaciones) ----------
@api.get("/catalogo-bancos")
async def catalogo_bancos(user: dict = Depends(get_current_user)):
    """Catálogo fijo de bancos con logo_url; única fuente para el selector."""
    import bancos as _bancos
    return [{"nombre": b["nombre"], "aliases": b.get("aliases", []),
             "color": b.get("color", "#C1401E"),
             "logo_url": "/api/bancos-logo/%s" % b["logo"]} for b in _bancos.BANCOS]


@api.get("/bancos-logo/{archivo}")
def bancos_logo_get(archivo: str):
    """Asset estático del logo del banco (público: solo imágenes, sin datos).
    Público a propósito: las etiquetas <img> del frontend no envían Authorization."""
    import bancos as _bancos
    from fastapi.responses import FileResponse
    try:
        return _bancos.servir_logo(archivo)
    except FileNotFoundError:
        raise HTTPException(404, "Logo no encontrado")

@api.post("/cuentas-bancarias")
async def cuenta_bancaria_create(data: CuentaBancariaInput,
                                 user: dict = Depends(require_permission("cuentas.editar"))):
    cid = uid()
    if data.predeterminada:
        for cta in await db.cuentas_bancarias.find({}, {"_id": 0}).to_list(5000):
            await db.cuentas_bancarias.update_one({"id": cta["id"]}, {"$set": {"predeterminada": False}})
    doc = {"id": cid, **data.model_dump(), "created_at": iso_now(),
           "usuario_id": user["id"], "usuario_nombre": user["name"]}
    await db.cuentas_bancarias.insert_one(doc)
    await log_audit(user, "cuenta_bancaria_crear", "cuenta_bancaria", cid,
                    f"{data.banco} {data.numero_cuenta}")
    return doc

@api.put("/cuentas-bancarias/{cuenta_id}")
async def cuenta_bancaria_update(cuenta_id: str, data: CuentaBancariaInput,
                                 user: dict = Depends(require_permission("cuentas.editar"))):
    if not await db.cuentas_bancarias.find_one({"id": cuenta_id}):
        raise HTTPException(404, "Cuenta no encontrada")
    if data.predeterminada:
        for cta in await db.cuentas_bancarias.find({}, {"_id": 0}).to_list(5000):
            await db.cuentas_bancarias.update_one({"id": cta["id"]}, {"$set": {"predeterminada": False}})
    up = data.model_dump()
    await db.cuentas_bancarias.update_one({"id": cuenta_id}, {"$set": up})
    await log_audit(user, "cuenta_bancaria_editar", "cuenta_bancaria", cuenta_id,
                    f"{data.banco} {data.numero_cuenta}")
    return await db.cuentas_bancarias.find_one({"id": cuenta_id}, {"_id": 0})

@api.patch("/cuentas-bancarias/{cuenta_id}/pagar")
async def cuenta_bancaria_pagar(cuenta_id: str, user: dict = Depends(require_permission("cuentas.editar"))):
    """Activa/desactiva una cuenta (nunca se elimina físicamente)."""
    cta = await db.cuentas_bancarias.find_one({"id": cuenta_id})
    if not cta:
        raise HTTPException(404, "Cuenta no encontrada")
    nueva = not bool(cta.get("activa", True))
    await db.cuentas_bancarias.update_one({"id": cuenta_id}, {"$set": {"activa": nueva}})
    await log_audit(user, "cuenta_bancaria_estado", "cuenta_bancaria", cuenta_id,
                    "activa" if nueva else "inactiva")
    return {"ok": True, "activa": nueva}

# =========================================================================
# COMPRAS / GASTOS
# =========================================================================
@api.post("/compras")
async def compras_create(data: CompraInput, user: dict = Depends(get_current_user)):
    """Registra y confirma una compra/gasto. Afecta inventario de forma
    transaccional SOLO para los items con `afecta_inventario`."""
    proveedor_nombre = data.proveedor_nombre or ""
    if data.proveedor_id:
        prov = await db.proveedores.find_one({"id": data.proveedor_id}, {"_id": 0})
        if prov:
            proveedor_nombre = prov.get("nombre") or proveedor_nombre
    tipo = data.tipo
    items = []
    for it in data.items:
        importe = it.importe
        if importe is None:
            importe = round(float(it.cantidad) * float(it.costo) - float(it.descuento), 2)
        items.append({"product_id": it.product_id, "codigo": it.codigo,
                      "descripcion": it.descripcion, "unidad": it.unidad,
                      "cantidad": float(it.cantidad), "costo": float(it.costo),
                      "iva_tasa": float(it.iva_tasa), "descuento": float(it.descuento),
                      "afecta_inventario": bool(it.afecta_inventario),
                      "importe": round(float(importe), 2)})
    if tipo not in ("compra", "gasto", "mixto"):
        tipo = "compra" if any(i["afecta_inventario"] for i in items) else "gasto"
    cid = uid()
    folio = await next_counter("compra", "CMP", 6) if tipo != "gasto" else await next_counter("gasto", "GST", 6)
    saldo_pendiente = 0.0
    if data.forma_pago == "credito" or data.metodo_pago == "credito":
        saldo_pendiente = round(float(data.total), 2)
    caja = await caja_abierta_de(user["id"]) if data.metodo_pago == "efectivo" else None
    costos_adicionales = round(float(data.flete) + float(data.seguro) +
                               float(data.maniobras) + float(data.transporte) +
                               float(data.otros_costos), 2)
    doc = {
        "id": cid, "folio": folio, "tipo": tipo,
        "proveedor_id": data.proveedor_id, "proveedor_nombre": proveedor_nombre,
        "factura_numero": data.factura_numero, "fecha_factura": data.fecha_factura,
        "fecha_recepcion": data.fecha_recepcion or iso_now(),
        "fecha_vencimiento": data.fecha_vencimiento, "concepto": data.concepto,
        "categoria": data.categoria, "subtotal": round(float(data.subtotal), 2),
        "descuento": round(float(data.descuento), 2), "iva": round(float(data.iva), 2),
        "otros_impuestos": round(float(data.otros_impuestos), 2),
        "total": round(float(data.total), 2), "metodo_pago": data.metodo_pago,
        "forma_pago": data.forma_pago or ("credito" if data.metodo_pago == "credito" else "contado"),
        "cuenta_bancaria_id": data.cuenta_bancaria_id,
        "observaciones": data.observaciones, "items": items,
        "documentos": data.documentos or [],
        # Evolución Compras y Gastos
        "orden_id": data.orden_id, "recepcion_id": data.recepcion_id,
        "centro_costo_id": data.centro_costo_id, "centro_costo_nombre": data.centro_costo_nombre,
        "sucursal_id": data.sucursal_id or user.get("sucursal_id"),
        "costos_adicionales": {
            "flete": round(float(data.flete), 2), "seguro": round(float(data.seguro), 2),
            "maniobras": round(float(data.maniobras), 2),
            "transporte": round(float(data.transporte), 2),
            "otros": round(float(data.otros_costos), 2),
            "total": costos_adicionales,
        },
        "costo_total_mercancia": round(float(data.total) + costos_adicionales, 2),
        "pagos": [],
        "abonado": 0.0, "saldo_pendiente": round(saldo_pendiente, 2),
        "estado": "confirmada", "caja_id": caja["id"] if caja else None,
        "usuario_id": user["id"], "usuario_nombre": user["name"],
        "fecha": iso_now(), "created_at": iso_now(),
    }
    try:
        result = await _pgcompras.registrar_compra_pg(user=user, doc=doc)
    except Exception as e:
        if getattr(e, "status", 0) and getattr(e, "message", None):
            raise HTTPException(e.status, e.message)
        raise
    # Efectivo al contado entra a caja.
    if caja and data.metodo_pago == "efectivo" and data.forma_pago != "credito":
        await db.caja_movimientos.insert_one({
            "id": uid(), "caja_id": caja["id"], "tipo": "entrada",
            "concepto": (f"Compra {folio}" if tipo != "gasto" else f"Gasto {folio}") + " · " + (proveedor_nombre or "S/D"),
            "monto": -round(float(data.total), 2), "referencia": folio,
            "usuario_id": user["id"], "usuario_nombre": user["name"], "fecha": iso_now()})
    return result

@api.post("/compras/{compra_id}/cancelar")
async def compras_cancelar(compra_id: str, motivo: str = "Cancelación",
                           user: dict = Depends(require_permission("compra.cancelar"))):
    try:
        result = await _pgcompras.cancela_compra_pg(user=user, compra_id=compra_id, motivo=motivo)
    except Exception as e:
        if getattr(e, "status", 0) and getattr(e, "message", None):
            raise HTTPException(e.status, e.message)
        raise
    return result

@api.get("/compras")
async def compras_list(desde: Optional[str] = None, hasta: Optional[str] = None,
                       tipo: Optional[str] = None, proveedor: Optional[str] = None,
                       rfc: Optional[str] = None, factura: Optional[str] = None,
                       categoria: Optional[str] = None, producto: Optional[str] = None,
                       sucursal: Optional[str] = None, usuario: Optional[str] = None,
                       metodo_pago: Optional[str] = None, estado: Optional[str] = None,
                       afecta_inventario: Optional[bool] = None,
                       con_documento: Optional[bool] = None,
                       pagada: Optional[str] = None,
                       q: Optional[str] = None,
                       user: dict = Depends(get_current_user)):
    flt = {"estado": {"$ne": "borrador"}} if estado != "todos" else {}
    docs = await db.compras.find(flt, {"_id": 0}).to_list(100000)
    if desde:
        docs = [d for d in docs if (d.get("fecha_recepcion") or "")[:10] >= desde]
    if hasta:
        docs = [d for d in docs if (d.get("fecha_recepcion") or "")[:10] <= hasta]
    if tipo and tipo != "todos":
        docs = [d for d in docs if d.get("tipo") == tipo]
    if proveedor:
        docs = [d for d in docs if d.get("proveedor_id") == proveedor or proveedor.lower() in (d.get("proveedor_nombre") or "").lower()]
    if rfc:
        docs = [d for d in docs if rfc.lower() in (d.get("proveedor_rfc") or "").lower()]
    if factura:
        docs = [d for d in docs if factura.lower() in (d.get("factura_numero") or "").lower()]
    if categoria and categoria != "todos":
        docs = [d for d in docs if d.get("categoria") == categoria]
    if producto:
        docs = [d for d in docs if any(producto.lower() in (it.get("descripcion") or "").lower() or producto.lower() == str(it.get("codigo") or "").lower() for it in d.get("items", []))]
    if sucursal:
        docs = [d for d in docs if d.get("sucursal_id") == sucursal]
    if usuario:
        docs = [d for d in docs if d.get("usuario_id") == usuario]
    if metodo_pago and metodo_pago != "todos":
        docs = [d for d in docs if d.get("metodo_pago") == metodo_pago]
    if estado and estado != "todos":
        docs = [d for d in docs if d.get("estado") == estado]
    if afecta_inventario is not None:
        docs = [d for d in docs if any(it.get("afecta_inventario") for it in d.get("items", [])) == afecta_inventario] if afecta_inventario else \
               [d for d in docs if not any(it.get("afecta_inventario") for it in d.get("items", []))]
    if con_documento is not None:
        docs = [d for d in docs if bool(d.get("documentos")) == con_documento]
    if pagada and pagada != "todos":
        if pagada == "pagada":
            docs = [d for d in docs if float(d.get("saldo_pendiente", 0) or 0) <= 0]
        elif pagada == "pendiente":
            docs = [d for d in docs if float(d.get("saldo_pendiente", 0) or 0) > 0]
        elif pagada == "vencida":
            docs = [d for d in docs if float(d.get("saldo_pendiente", 0) or 0) > 0 and (d.get("fecha_vencimiento") or "")[:10] and (d.get("fecha_vencimiento") or "")[:10] < now_utc().date().isoformat()]
    if q:
        ql = q.lower().strip()
        docs = [d for d in docs if ql in " ".join(str(d.get(k) or "") for k in ("folio", "factura_numero", "concepto", "proveedor_nombre")).lower()]
    docs.sort(key=lambda d: d.get("fecha_recepcion") or "", reverse=True)
    return docs

@api.get("/compras/resumen")
async def compras_resumen(desde: Optional[str] = None, hasta: Optional[str] = None,
                          user: dict = Depends(get_current_user)):
    docs = await db.compras.find({"estado": "confirmada"}, {"_id": 0}).to_list(100000)
    if desde:
        docs = [d for d in docs if (d.get("fecha_recepcion") or "")[:10] >= desde]
    if hasta:
        docs = [d for d in docs if (d.get("fecha_recepcion") or "")[:10] <= hasta]
    total_compras = round(sum(float(d.get("total", 0) or 0) for d in docs if d.get("tipo") != "gasto"), 2)
    total_gastos = round(sum(float(d.get("total", 0) or 0) for d in docs if d.get("tipo") == "gasto"), 2)
    compras_pend = round(sum(float(d.get("saldo_pendiente", 0) or 0) for d in docs if d.get("tipo") != "gasto"), 2)
    gastos_pend = round(sum(float(d.get("saldo_pendiente", 0) or 0) for d in docs if d.get("tipo") == "gasto"), 2)
    # Proveedor con mayor volumen, categoría mayor y productos más comprados.
    por_prov = {}
    por_cat = {}
    prod_count = {}
    for d in docs:
        por_prov[d.get("proveedor_nombre") or "S/D"] = round(por_prov.get(d.get("proveedor_nombre") or "S/D", 0) + float(d.get("total", 0) or 0), 2)
        cat = d.get("categoria") or "Sin categoría"
        por_cat[cat] = round(por_cat.get(cat, 0) + float(d.get("total", 0) or 0), 2)
        for it in d.get("items", []):
            if it.get("afecta_inventario"):
                k = it.get("product_id") or it.get("codigo")
                e = prod_count.setdefault(k, {"product_id": it.get("product_id"), "codigo": it.get("codigo"),
                                              "descripcion": it.get("descripcion"), "cantidad": 0.0})
                e["cantidad"] = round(e["cantidad"] + float(it.get("cantidad", 0) or 0), 2)
    mes_actual = now_utc().strftime("%Y-%m")
    docs_mes = [d for d in docs if (d.get("fecha_recepcion") or "")[:7] == mes_actual]
    return {
        "compras_periodo": total_compras, "gastos_periodo": total_gastos,
        "total_periodo": round(total_compras + total_gastos, 2),
        "compras_pendientes": compras_pend, "gastos_pendientes": gastos_pend,
        "facturas": len([d for d in docs if d.get("factura_numero")]),
        "compras_mes": round(sum(float(d.get("total", 0) or 0) for d in docs_mes if d.get("tipo") != "gasto"), 2),
        "gastos_mes": round(sum(float(d.get("total", 0) or 0) for d in docs_mes if d.get("tipo") == "gasto"), 2),
        "mejor_proveedor": max(por_prov, key=por_prov.get, default="—"),
        "mejor_categoria": max(por_cat, key=por_cat.get, default="—"),
        "productos_mas_comprados": sorted(prod_count.values(), key=lambda x: x["cantidad"], reverse=True)[:10],
    }

@api.post("/compras/ocr")
async def compras_ocr(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """Lee una factura (foto o PDF) con OCR y devuelve proveedor, montos y
    conceptos. Los conceptos se cruzan contra el catálogo de productos para
    poder ajustar el inventario automáticamente al confirmar la compra."""
    data = await file.read()
    if len(data) > 15 * 1024 * 1024:
        raise HTTPException(400, "El archivo no debe superar 15 MB.")
    filename = file.filename or ""
    real_mime = storage.detect_mime_type(data)
    if real_mime not in ("application/pdf", "image/jpeg", "image/png", "image/webp", "image/gif"):
        raise HTTPException(400, "Formato no permitido. Usa una imagen (JPG/PNG/WEBP) o un PDF de la factura.")
    products = await db.products.find(
        {"estado": "activo"},
        {"_id": 0, "id": 1, "codigo": 1, "descripcion": 1, "costo": 1,
         "iva_tasa": 1, "unidad_medida": 1}).to_list(100000)
    try:
        # OCR pesado (Tesseract + render PDF): corre en threadpool para NO
        # bloquear el event loop (antes congelaba la API decenas de segundos).
        result = await run_in_threadpool(_ocr_invoice.process_factura, data, filename, products)
    except HTTPException:
        raise
    except Exception as e:
        logger.warning("OCR factura falló: %s", str(e)[:200])
        raise HTTPException(422, str(e))
    return result

# =========================================================================
# ÓRDENES DE COMPRA
# =========================================================================
@api.get("/compras/ordenes")
async def ordenes_list(estado: Optional[str] = None, proveedor: Optional[str] = None,
                       pendientes: Optional[bool] = None, q: Optional[str] = None,
                       user: dict = Depends(get_current_user)):
    docs = await db.compras_ordenes.find({}, {"_id": 0}).to_list(100000)
    if estado and estado != "todos":
        docs = [d for d in docs if d.get("estado") == estado]
    if proveedor:
        docs = [d for d in docs if d.get("proveedor_id") == proveedor or
                proveedor.lower() in (d.get("proveedor_nombre") or "").lower()]
    if pendientes is True:
        docs = [d for d in docs if d.get("estado") in ("enviada", "parcialmente_recibida")]
    if q:
        ql = q.lower().strip()
        docs = [d for d in docs if ql in " ".join(str(d.get(k) or "") for k in ("folio", "proveedor_nombre", "notas")).lower()]
    docs.sort(key=lambda d: d.get("fecha_orden") or "", reverse=True)
    return docs

@api.get("/compras/ordenes/{orden_id}")
async def orden_detail(orden_id: str, user: dict = Depends(get_current_user)):
    doc = await db.compras_ordenes.find_one({"id": orden_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Orden no encontrada")
    return doc

@api.post("/compras/ordenes")
async def orden_create(data: OrdenCompraInput, user: dict = Depends(get_current_user)):
    proveedor_nombre = data.proveedor_nombre or ""
    if data.proveedor_id:
        prov = await db.proveedores.find_one({"id": data.proveedor_id}, {"_id": 0})
        if prov:
            proveedor_nombre = prov.get("nombre") or proveedor_nombre
    if not data.items:
        raise HTTPException(400, "Agrega al menos un producto a la orden")
    oid = uid()
    folio = await next_counter("orden", "OC", 6)
    items = []
    subtotal = 0.0
    for it in data.items:
        if float(it.solicitado) <= 0:
            continue
        subtotal += round(float(it.solicitado) * float(it.costo), 2)
        items.append({"product_id": it.product_id, "codigo": it.codigo,
                      "descripcion": it.descripcion, "unidad": it.unidad or "PZA",
                      "solicitado": float(it.solicitado), "recibido": 0.0,
                      "pendiente": float(it.solicitado), "costo": float(it.costo),
                      "iva_tasa": float(it.iva_tasa)})
    if not items:
        raise HTTPException(400, "No hay cantidades válidas en la orden")
    iva = round(sum(float(i["solicitado"]) * float(i["costo"]) * (float(i["iva_tasa"]) / 100) for i in items), 2)
    estado = "enviada" if data.estado == "enviada" else "borrador"
    doc = {
        "id": oid, "folio": folio, "proveedor_id": data.proveedor_id,
        "proveedor_nombre": proveedor_nombre,
        "fecha_orden": data.fecha_orden or iso_now()[:10],
        "fecha_estimada": data.fecha_estimada, "estado": estado,
        "notas": data.notas or "", "items": items,
        "subtotal": round(subtotal, 2), "iva": iva, "total": round(subtotal + iva, 2),
        "sucursal_id": data.sucursal_id or user.get("sucursal_id"),
        "cuenta_bancaria_id": data.cuenta_bancaria_id,
        "usuario_id": user["id"], "usuario_nombre": user["name"],
        "creado_en": iso_now(), "actualizado_en": iso_now(),
    }
    await db.compras_ordenes.insert_one(doc)
    await log_audit(user, "orden_crear", "orden_compra", oid, f"{folio} {proveedor_nombre}")
    return doc

@api.put("/compras/ordenes/{orden_id}")
async def orden_update(orden_id: str, data: OrdenCompraInput,
                       user: dict = Depends(get_current_user)):
    ex = await db.compras_ordenes.find_one({"id": orden_id})
    if not ex:
        raise HTTPException(404, "Orden no encontrada")
    if ex.get("estado") in ("recibida", "cancelada"):
        raise HTTPException(409, f"Una orden {ex['estado']} no se puede editar")
    proveedor_nombre = data.proveedor_nombre or ex.get("proveedor_nombre") or ""
    if data.proveedor_id:
        prov = await db.proveedores.find_one({"id": data.proveedor_id}, {"_id": 0})
        if prov:
            proveedor_nombre = prov.get("nombre") or proveedor_nombre
    items = []
    subtotal = 0.0
    for it in data.items:
        if float(it.solicitado) <= 0:
            continue
        prev = next((x for x in ex.get("items", []) if x.get("product_id") == it.product_id), None)
        recibido = float((prev or {}).get("recibido", 0) or 0)
        if float(it.solicitado) < recibido:
            raise HTTPException(400, f"{it.descripcion}: no puedes bajar la cantidad por debajo de lo ya recibido ({recibido})")
        subtotal += round(float(it.solicitado) * float(it.costo), 2)
        items.append({"product_id": it.product_id, "codigo": it.codigo,
                      "descripcion": it.descripcion, "unidad": it.unidad or "PZA",
                      "solicitado": float(it.solicitado), "recibido": recibido,
                      "pendiente": round(float(it.solicitado) - recibido, 3),
                      "costo": float(it.costo), "iva_tasa": float(it.iva_tasa)})
    iva = round(sum(float(i["solicitado"]) * float(i["costo"]) * (float(i["iva_tasa"]) / 100) for i in items), 2)
    up = {"id": orden_id, "folio": ex.get("folio"), "proveedor_id": data.proveedor_id,
          "proveedor_nombre": proveedor_nombre,
          "fecha_orden": data.fecha_orden or ex.get("fecha_orden"),
          "fecha_estimada": data.fecha_estimada, "estado": ex.get("estado"),
          "notas": data.notas or "", "items": items,
          "subtotal": round(subtotal, 2), "iva": iva, "total": round(subtotal + iva, 2),
          "sucursal_id": data.sucursal_id or ex.get("sucursal_id"),
          "cuenta_bancaria_id": data.cuenta_bancaria_id,
          "usuario_id": ex.get("usuario_id"), "usuario_nombre": ex.get("usuario_nombre"),
          "creado_en": ex.get("creado_en"), "actualizado_en": iso_now()}
    await db.compras_ordenes.update_one({"id": orden_id}, {"$set": up})
    await log_audit(user, "orden_editar", "orden_compra", orden_id, ex.get("folio"))
    return up

@api.post("/compras/ordenes/{orden_id}/estado")
async def orden_estado(orden_id: str, data: OrdenEstadoInput,
                       user: dict = Depends(get_current_user)):
    ex = await db.compras_ordenes.find_one({"id": orden_id})
    if not ex:
        raise HTTPException(404, "Orden no encontrada")
    if data.estado not in ("borrador", "enviada", "cancelada"):
        raise HTTPException(400, "Estado no válido")
    if ex.get("estado") == "recibida" and data.estado != "cancelada":
        raise HTTPException(409, "Una orden recibida no puede cambiar de estado")
    await db.compras_ordenes.update_one({"id": orden_id},
                                        {"$set": {"estado": data.estado, "actualizado_en": iso_now()}})
    await log_audit(user, "orden_estado", "orden_compra", orden_id,
                    f"{ex.get('folio')} -> {data.estado}")
    return {"ok": True, "estado": data.estado}

# =========================================================================
# RECEPCIONES DE MERCANCÍA
# =========================================================================
@api.post("/compras/recepciones")
async def recepcion_create(data: RecepcionInput, user: dict = Depends(get_current_user)):
    orden = await db.compras_ordenes.find_one({"id": data.orden_id}, {"_id": 0})
    if not orden:
        raise HTTPException(404, "Orden de compra no encontrada")
    if orden.get("estado") == "cancelada":
        raise HTTPException(409, "No se puede recibir una orden cancelada")
    # Mapa de productos de la orden.
    orden_items = {i.get("product_id"): i for i in orden.get("items", []) if i.get("product_id")}
    rec_items = []
    for it in data.items:
        pid = it.product_id
        oi = orden_items.get(pid)
        if oi is None:
            raise HTTPException(400, f"{it.descripcion or it.codigo} no está en la orden")
        pendiente = float(oi.get("pendiente", 0) or 0)
        cantidad = float(it.cantidad)
        if cantidad <= 0:
            cantidad = pendiente
        if cantidad > pendiente + 1e-9:
            raise HTTPException(400, f"{oi.get('descripcion')}: recibes {cantidad} pero solo quedan {pendiente} pendientes")
        if cantidad <= 0:
            continue
        rec_items.append({"product_id": pid, "codigo": it.codigo or oi.get("codigo"),
                          "descripcion": it.descripcion or oi.get("descripcion"),
                          "unidad": it.unidad or oi.get("unidad") or "PZA",
                          "cantidad": round(cantidad, 3),
                          "costo": float(it.costo) if it.costo else float(oi.get("costo", 0) or 0),
                          "iva_tasa": float(it.iva_tasa) if it.iva_tasa else float(oi.get("iva_tasa", 8) or 8)})
    if not rec_items:
        raise HTTPException(400, "No hay cantidades recibidas que registrar")

    rid = uid()
    folio_rcp = await next_counter("recepcion", "RCP", 6)
    folio_cmp = await next_counter("compra", "CMP", 6)
    subtotal = round(sum(float(i["cantidad"]) * float(i["costo"]) for i in rec_items), 2)
    iva = round(sum(float(i["cantidad"]) * float(i["costo"]) * (float(i["iva_tasa"]) / 100) for i in rec_items), 2)
    total = round(subtotal + iva, 2)
    cid = uid()
    proveedor_nombre = orden.get("proveedor_nombre") or ""
    fecha = data.fecha or iso_now()[:10]
    saldo_pendiente = round(total, 2) if (data.forma_pago == "credito" or data.metodo_pago == "credito") else 0.0

    recepcion = {
        "id": rid, "folio": folio_rcp, "orden_id": data.orden_id,
        "orden_folio": orden.get("folio"), "proveedor_id": orden.get("proveedor_id"),
        "proveedor_nombre": proveedor_nombre, "fecha": fecha,
        "factura_numero": data.factura_numero, "fecha_factura": data.fecha_factura,
        "metodo_pago": data.metodo_pago,
        "forma_pago": data.forma_pago or ("credito" if data.metodo_pago == "credito" else "contado"),
        "cuenta_bancaria_id": data.cuenta_bancaria_id,
        "fecha_vencimiento": data.fecha_vencimiento, "observaciones": data.observaciones or "",
        "items": rec_items, "subtotal": subtotal, "iva": iva, "total": total,
        "saldo_pendiente": saldo_pendiente,
        "documentos": data.documentos or [], "estado": "confirmada",
        "usuario_id": user["id"], "usuario_nombre": user["name"],
        "sucursal_id": orden.get("sucursal_id") or user.get("sucursal_id"),
        "compra_id": cid, "created_at": iso_now(),
    }
    compra = {
        "id": cid, "folio": folio_cmp, "tipo": "compra",
        "proveedor_id": orden.get("proveedor_id"), "proveedor_nombre": proveedor_nombre,
        "factura_numero": data.factura_numero, "fecha_factura": data.fecha_factura,
        "fecha_recepcion": fecha, "fecha_vencimiento": data.fecha_vencimiento,
        "concepto": f"Recepción {folio_rcp} / Orden {orden.get('folio')}",
        "categoria": "", "subtotal": subtotal, "descuento": 0.0, "iva": iva,
        "otros_impuestos": 0.0, "total": total, "metodo_pago": data.metodo_pago,
        "forma_pago": recepcion["forma_pago"], "cuenta_bancaria_id": data.cuenta_bancaria_id,
        "observaciones": data.observaciones or "",
        "items": [{**i, "descuento": 0.0, "afecta_inventario": False,
                   "importe": round(float(i["cantidad"]) * float(i["costo"]), 2)} for i in rec_items],
        "documentos": data.documentos or [],
        "orden_id": data.orden_id, "recepcion_id": rid,
        "centro_costo_id": None, "centro_costo_nombre": "",
        "sucursal_id": recepcion["sucursal_id"],
        "costos_adicionales": {"flete": 0.0, "seguro": 0.0, "maniobras": 0.0,
                               "transporte": 0.0, "otros": 0.0, "total": 0.0},
        "costo_total_mercancia": total, "pagos": [],
        "abonado": 0.0, "saldo_pendiente": saldo_pendiente,
        "estado": "confirmada", "caja_id": None,
        "usuario_id": user["id"], "usuario_nombre": user["name"],
        "fecha": iso_now(), "created_at": iso_now(),
    }
    try:
        result = await _pgcompras.recibir_orden_pg(user=user, orden=orden,
                                                   recepcion=recepcion, compra=compra)
    except Exception as e:
        if getattr(e, "status", 0) and getattr(e, "message", None):
            raise HTTPException(e.status, e.message)
        raise
    # Efectivo al contado sale de caja.
    caja = await caja_abierta_de(user["id"]) if data.metodo_pago == "efectivo" else None
    if caja and recepcion["forma_pago"] != "credito":
        await db.caja_movimientos.insert_one({
            "id": uid(), "caja_id": caja["id"], "tipo": "gasto",
            "concepto": f"Recepción {folio_rcp} · {proveedor_nombre or 'S/D'}",
            "monto": round(total, 2), "referencia": folio_rcp,
            "usuario_id": user["id"], "usuario_nombre": user["name"], "fecha": iso_now()})
    await log_audit(user, "recepcion_crear", "recepcion", rid,
                    f"{folio_rcp} orden {orden.get('folio')} total {total}")
    return result

@api.get("/compras/recepciones")
async def recepciones_list(desde: Optional[str] = None, hasta: Optional[str] = None,
                           proveedor: Optional[str] = None, orden: Optional[str] = None,
                           user: dict = Depends(get_current_user)):
    docs = await db.compras_recepciones.find({}, {"_id": 0}).to_list(100000)
    if desde:
        docs = [d for d in docs if (d.get("fecha") or "")[:10] >= desde]
    if hasta:
        docs = [d for d in docs if (d.get("fecha") or "")[:10] <= hasta]
    if proveedor:
        docs = [d for d in docs if d.get("proveedor_id") == proveedor or
                proveedor.lower() in (d.get("proveedor_nombre") or "").lower()]
    if orden:
        docs = [d for d in docs if d.get("orden_id") == orden or
                orden.lower() in (d.get("orden_folio") or "").lower()]
    docs.sort(key=lambda d: d.get("fecha") or "", reverse=True)
    return docs

@api.get("/compras/recepciones/{recepcion_id}")
async def recepcion_detail(recepcion_id: str, user: dict = Depends(get_current_user)):
    doc = await db.compras_recepciones.find_one({"id": recepcion_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Recepción no encontrada")
    return doc

# =========================================================================
# CENTROS DE COSTO
# =========================================================================
@api.get("/centros-costo")
async def centros_costo_list(user: dict = Depends(get_current_user)):
    docs = await db.centros_costo.find({}, {"_id": 0}).to_list(5000)
    docs.sort(key=lambda d: (0 if d.get("activo", True) else 1, (d.get("nombre") or "").lower()))
    return docs

@api.post("/centros-costo")
async def centro_costo_create(data: CentroCostoInput, user: dict = Depends(get_current_user)):
    cid = uid()
    doc = {"id": cid, "codigo": data.codigo or f"CC{cid[:4].upper()}",
           "nombre": data.nombre, "descripcion": data.descripcion or "",
           "activo": data.activo, "created_at": iso_now(),
           "usuario_id": user["id"], "usuario_nombre": user["name"]}
    await db.centros_costo.insert_one(doc)
    await log_audit(user, "centro_costo_crear", "centro_costo", cid, data.nombre)
    return doc

@api.put("/centros-costo/{centro_id}")
async def centro_costo_update(centro_id: str, data: CentroCostoInput,
                              user: dict = Depends(get_current_user)):
    if not await db.centros_costo.find_one({"id": centro_id}):
        raise HTTPException(404, "Centro de costo no encontrado")
    up = {"id": centro_id, "codigo": data.codigo, "nombre": data.nombre,
          "descripcion": data.descripcion or "", "activo": data.activo,
          "updated_at": iso_now()}
    await db.centros_costo.update_one({"id": centro_id}, {"$set": up})
    await log_audit(user, "centro_costo_editar", "centro_costo", centro_id, data.nombre)
    return up

@api.patch("/centros-costo/{centro_id}/estado")
async def centro_costo_estado(centro_id: str, activo: bool,
                              user: dict = Depends(get_current_user)):
    await db.centros_costo.update_one({"id": centro_id},
                                      {"$set": {"activo": activo, "updated_at": iso_now()}})
    return {"ok": True}

# =========================================================================
# PRESUPUESTOS
# =========================================================================
@api.get("/presupuestos")
async def presupuestos_list(periodo: Optional[str] = None, categoria: Optional[str] = None,
                            user: dict = Depends(get_current_user)):
    docs = await db.presupuestos.find({}, {"_id": 0}).to_list(5000)
    if periodo:
        docs = [d for d in docs if d.get("periodo") == periodo]
    if categoria:
        docs = [d for d in docs if d.get("categoria") == categoria]
    docs.sort(key=lambda d: (d.get("periodo") or "", d.get("categoria") or ""))
    return docs

@api.get("/presupuestos/resumen")
async def presupuestos_resumen(periodo: Optional[str] = None,
                               user: dict = Depends(get_current_user)):
    docs = await db.presupuestos.find({}, {"_id": 0}).to_list(5000)
    if periodo:
        docs = [d for d in docs if d.get("periodo") == periodo]
    gastos = await db.compras.find({"estado": "confirmada", "tipo": "gasto"}, {"_id": 0}).to_list(100000)
    mes = periodo or now_utc().strftime("%Y-%m")
    gastos_mes = [d for d in gastos if (d.get("fecha_recepcion") or "")[:7] == mes]
    def _gasto_clave(d):
        return (d.get("categoria") or "Sin categoría", d.get("centro_costo_nombre") or "")
    por_clave = {}
    for d in gastos_mes:
        k = _gasto_clave(d)
        por_clave[k] = round(por_clave.get(k, 0) + float(d.get("total", 0) or 0), 2)
    result = []
    for p in docs:
        monto = float(p.get("monto", 0) or 0)
        gastado = por_clave.get((p.get("categoria") or "Sin categoría", p.get("centro_costo_nombre") or ""), 0.0)
        disponible = round(monto - gastado, 2)
        result.append({
            "id": p["id"], "categoria": p.get("categoria"), "periodo": p.get("periodo"),
            "centro_costo_id": p.get("centro_costo_id"), "centro_costo_nombre": p.get("centro_costo_nombre"),
            "sucursal_id": p.get("sucursal_id"), "monto": monto, "gastado": gastado,
            "disponible": disponible, "excedido": disponible < 0, "notas": p.get("notas"),
        })
    result.sort(key=lambda x: (x["excedido"], -x["gastado"]))
    return {"periodo": mes, "presupuestos": result,
            "total_presupuestado": round(sum(x["monto"] for x in result), 2),
            "total_gastado": round(sum(x["gastado"] for x in result), 2),
            "total_disponible": round(sum(x["disponible"] for x in result), 2)}

@api.post("/presupuestos")
async def presupuesto_create(data: PresupuestoInput, user: dict = Depends(get_current_user)):
    if not data.periodo:
        raise HTTPException(400, "Indica el periodo (YYYY-MM)")
    pid = uid()
    doc = {"id": pid, "categoria": data.categoria or "Sin categoría",
           "sucursal_id": data.sucursal_id, "centro_costo_id": data.centro_costo_id,
           "centro_costo_nombre": data.centro_costo_nombre, "periodo": data.periodo,
           "monto": round(float(data.monto), 2), "notas": data.notas or "",
           "created_at": iso_now(), "usuario_id": user["id"], "usuario_nombre": user["name"]}
    await db.presupuestos.insert_one(doc)
    await log_audit(user, "presupuesto_crear", "presupuesto", pid,
                    f"{data.categoria} {data.periodo} {data.monto}")
    return doc

@api.put("/presupuestos/{presupuesto_id}")
async def presupuesto_update(presupuesto_id: str, data: PresupuestoInput,
                             user: dict = Depends(get_current_user)):
    if not await db.presupuestos.find_one({"id": presupuesto_id}):
        raise HTTPException(404, "Presupuesto no encontrado")
    up = {"id": presupuesto_id, "categoria": data.categoria or "Sin categoría",
          "sucursal_id": data.sucursal_id, "centro_costo_id": data.centro_costo_id,
          "centro_costo_nombre": data.centro_costo_nombre, "periodo": data.periodo,
          "monto": round(float(data.monto), 2), "notas": data.notas or "",
          "updated_at": iso_now()}
    await db.presupuestos.update_one({"id": presupuesto_id}, {"$set": up})
    await log_audit(user, "presupuesto_editar", "presupuesto", presupuesto_id, data.categoria)
    return up

# =========================================================================
# GASTOS / COMPRAS RECURRENTES
# =========================================================================
@api.get("/recurrentes")
async def recurrentes_list(activos: Optional[bool] = None,
                           user: dict = Depends(get_current_user)):
    docs = await db.recurrentes.find({}, {"_id": 0}).to_list(5000)
    if activos is not None:
        docs = [d for d in docs if d.get("activo") == activos]
    docs.sort(key=lambda d: (d.get("categoria") or "", d.get("concepto") or ""))
    return docs

@api.post("/recurrentes")
async def recurrente_create(data: RecurrenteInput, user: dict = Depends(get_current_user)):
    rid = uid()
    proveedor_nombre = data.proveedor_nombre or ""
    if data.proveedor_id:
        prov = await db.proveedores.find_one({"id": data.proveedor_id}, {"_id": 0})
        if prov:
            proveedor_nombre = prov.get("nombre") or proveedor_nombre
    doc = {"id": rid, "tipo": data.tipo, "proveedor_id": data.proveedor_id,
           "proveedor_nombre": proveedor_nombre, "concepto": data.concepto or "",
           "categoria": data.categoria or "", "importe": round(float(data.importe), 2),
           "frecuencia": data.frecuencia, "dia": max(1, min(31, int(data.dia or 1))),
           "cuenta_bancaria_id": data.cuenta_bancaria_id, "recordatorio": data.recordatorio,
           "sucursal_id": data.sucursal_id, "centro_costo_id": data.centro_costo_id,
           "centro_costo_nombre": data.centro_costo_nombre, "activo": data.activo,
           "notas": data.notas or "", "created_at": iso_now(),
           "usuario_id": user["id"], "usuario_nombre": user["name"]}
    await db.recurrentes.insert_one(doc)
    await log_audit(user, "recurrente_crear", "recurrente", rid, data.concepto)
    return doc

@api.put("/recurrentes/{recurrente_id}")
async def recurrente_update(recurrente_id: str, data: RecurrenteInput,
                            user: dict = Depends(get_current_user)):
    if not await db.recurrentes.find_one({"id": recurrente_id}):
        raise HTTPException(404, "Recurrente no encontrado")
    proveedor_nombre = data.proveedor_nombre or ""
    if data.proveedor_id:
        prov = await db.proveedores.find_one({"id": data.proveedor_id}, {"_id": 0})
        if prov:
            proveedor_nombre = prov.get("nombre") or proveedor_nombre
    up = {"id": recurrente_id, "tipo": data.tipo, "proveedor_id": data.proveedor_id,
          "proveedor_nombre": proveedor_nombre, "concepto": data.concepto or "",
          "categoria": data.categoria or "", "importe": round(float(data.importe), 2),
          "frecuencia": data.frecuencia, "dia": max(1, min(31, int(data.dia or 1))),
          "cuenta_bancaria_id": data.cuenta_bancaria_id, "recordatorio": data.recordatorio,
          "sucursal_id": data.sucursal_id, "centro_costo_id": data.centro_costo_id,
          "centro_costo_nombre": data.centro_costo_nombre, "activo": data.activo,
          "notas": data.notas or "", "updated_at": iso_now()}
    await db.recurrentes.update_one({"id": recurrente_id}, {"$set": up})
    await log_audit(user, "recurrente_editar", "recurrente", recurrente_id, data.concepto)
    return up

@api.patch("/recurrentes/{recurrente_id}/estado")
async def recurrente_estado(recurrente_id: str, activo: bool,
                            user: dict = Depends(get_current_user)):
    await db.recurrentes.update_one({"id": recurrente_id},
                                    {"$set": {"activo": activo, "updated_at": iso_now()}})
    return {"ok": True}

# =========================================================================
# CUENTAS POR PAGAR (CxP)
# =========================================================================
@api.get("/compras/cxp")
async def compras_cxp(desde: Optional[str] = None, hasta: Optional[str] = None,
                      user: dict = Depends(get_current_user)):
    docs = await db.compras.find({"estado": {"$in": ["confirmada", "pagada"]}}, {"_id": 0}).to_list(100000)
    if desde:
        docs = [d for d in docs if (d.get("fecha_recepcion") or "")[:10] >= desde]
    if hasta:
        docs = [d for d in docs if (d.get("fecha_recepcion") or "")[:10] <= hasta]
    hoy = now_utc().date().isoformat()
    dentro7 = (now_utc().date() + timedelta(days=7)).isoformat()
    pendientes = [d for d in docs if float(d.get("saldo_pendiente", 0) or 0) > 0]
    vencidas = [d for d in pendientes if (d.get("fecha_vencimiento") or "")[:10] and (d.get("fecha_vencimiento") or "")[:10] < hoy]
    proximas = [d for d in pendientes if (d.get("fecha_vencimiento") or "")[:10] and
                hoy <= (d.get("fecha_vencimiento") or "")[:10] <= dentro7]
    saldo_total = round(sum(float(d.get("saldo_pendiente", 0) or 0) for d in pendientes), 2)
    # Pagos realizados en el periodo.
    pagos = []
    for d in docs:
        for p in d.get("pagos", []):
            if desde and (p.get("fecha") or "")[:10] < desde:
                continue
            if hasta and (p.get("fecha") or "")[:10] > hasta:
                continue
            pagos.append({"compra_id": d["id"], "folio": d.get("folio"),
                          "proveedor_nombre": d.get("proveedor_nombre"),
                          "fecha": p.get("fecha"), "monto": p.get("monto"),
                          "metodo_pago": p.get("metodo_pago"), "referencia": p.get("referencia")})
    pagos.sort(key=lambda x: x.get("fecha") or "", reverse=True)
    pendientes.sort(key=lambda d: (d.get("fecha_vencimiento") or "")[:10] or "9999", reverse=False)
    return {
        "saldo_total": saldo_total,
        "facturas_pendientes": len(pendientes),
        "vencidas": [{"id": d["id"], "folio": d.get("folio"), "proveedor_nombre": d.get("proveedor_nombre"),
                      "factura_numero": d.get("factura_numero"), "fecha_vencimiento": d.get("fecha_vencimiento"),
                      "saldo": float(d.get("saldo_pendiente", 0) or 0)} for d in vencidas],
        "vencidas_total": round(sum(float(d.get("saldo_pendiente", 0) or 0) for d in vencidas), 2),
        "proximas_vencer": [{"id": d["id"], "folio": d.get("folio"), "proveedor_nombre": d.get("proveedor_nombre"),
                             "factura_numero": d.get("factura_numero"), "fecha_vencimiento": d.get("fecha_vencimiento"),
                             "saldo": float(d.get("saldo_pendiente", 0) or 0)} for d in proximas],
        "pagos": pagos,
        "facturas": [{"id": d["id"], "folio": d.get("folio"), "tipo": d.get("tipo"),
                      "proveedor_id": d.get("proveedor_id"), "proveedor_nombre": d.get("proveedor_nombre"),
                      "factura_numero": d.get("factura_numero"), "fecha_recepcion": d.get("fecha_recepcion"),
                      "fecha_vencimiento": d.get("fecha_vencimiento"), "total": float(d.get("total", 0) or 0),
                      "abonado": float(d.get("abonado", 0) or 0), "saldo": float(d.get("saldo_pendiente", 0) or 0),
                      "metodo_pago": d.get("metodo_pago"), "sucursal_id": d.get("sucursal_id")} for d in pendientes],
    }

# Alias section under Cuenta por pagar (CxP) - nueva navegación
@api.get("/cxp")
async def cxp_list(desde: Optional[str] = None, hasta: Optional[str] = None,
                   user: dict = Depends(get_current_user)):
    """Vista simplificada de Cuentas por pagar para el módulo CxP."""
    return await compras_cxp(desde=desde, hasta=hasta, user=user)


# ---- Exportación CxP (Excel + PDF) ----
CXP_EXPORT_HEADERS = ["Folio", "Proveedor", "Factura", "Vencimiento",
                      "Total", "Abonado", "Saldo", "Estado"]


def _cxp_export_rows(facturas) -> list:
    rows = []
    for d in facturas:
        saldo = float(d.get("saldo", 0) or 0)
        total = float(d.get("total", 0) or 0)
        abonado = float(d.get("abonado", 0) or 0)
        rows.append({
            "Folio": d.get("folio"),
            "Proveedor": d.get("proveedor_nombre"),
            "Factura": d.get("factura_numero") or "—",
            "Vencimiento": (d.get("fecha_vencimiento") or "")[:10],
            "Total": round(total, 2),
            "Abonado": round(abonado, 2),
            "Saldo": round(saldo, 2),
            "Estado": "Pendiente" if saldo > 0 else "Pagada",
        })
    return rows


def _cxp_filtros_dict(desde, hasta):
    vals = {"Período desde": desde, "Período hasta": hasta}
    return {k: v for k, v in vals.items() if v not in (None, "", "all")}


@api.get("/cxp/exportar.xlsx")
async def cxp_export_excel(desde: Optional[str] = None, hasta: Optional[str] = None,
                           user: dict = Depends(require_permission("exportar"))):
    res = await compras_cxp(desde=desde, hasta=hasta, user=user)
    rows = _cxp_export_rows(res.get("facturas", []))
    data = exports.excel_bytes(rows, CXP_EXPORT_HEADERS,
                               sheet_name="Cuentas por pagar", title="CUENTAS POR PAGAR - GRUPO RYSA")
    return StreamingResponse(io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=cxp.xlsx"})


@api.get("/cxp/exportar.pdf")
async def cxp_export_pdf(desde: Optional[str] = None, hasta: Optional[str] = None,
                         user: dict = Depends(require_permission("exportar"))):
    res = await compras_cxp(desde=desde, hasta=hasta, user=user)
    rows = _cxp_export_rows(res.get("facturas", []))
    sett = await db.settings.find_one({"_id": "app"}) or {}
    filt = _cxp_filtros_dict(desde, hasta)
    data = exports.pdf_bytes("REPORTE DE CUENTAS POR PAGAR", CXP_EXPORT_HEADERS,
                             [[r[h] for h in CXP_EXPORT_HEADERS] for r in rows],
                             settings=sett, user_name=user.get("name"), filtros=filt)
    return StreamingResponse(io.BytesIO(data), media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=cxp.pdf"})

@api.post("/cxp/{compra_id}/pagar")
async def cxp_pagar(compra_id: str, data: CompraPagoInput,
                    user: dict = Depends(require_permission("cxp.pagar"))):
    """Pago de factura vía Cuenta por pagar."""
    return await compras_pagar(compra_id, data, user)

@api.post("/compras/{compra_id}/pagar")
async def compras_pagar(compra_id: str, data: CompraPagoInput,
                        user: dict = Depends(require_permission("cxp.pagar"))):
    compra = await db.compras.find_one({"id": compra_id})
    if not compra:
        raise HTTPException(404, "Compra no encontrada")
    pend = float(compra.get("saldo_pendiente", 0) or 0)
    if pend <= 0:
        raise HTTPException(409, "La factura ya está pagada")
    if data.monto <= 0:
        raise HTTPException(400, "El monto debe ser mayor a cero")
    if data.monto > pend + 0.001:
        raise HTTPException(400, f"El monto excede el saldo pendiente ({pend})")
    monto = round(float(data.monto), 2)
    nuevo_saldo = round(pend - monto, 2)
    pago = {"id": uid(), "fecha": data.fecha or iso_now()[:10], "monto": monto,
            "metodo_pago": data.metodo_pago, "cuenta_bancaria_id": data.cuenta_bancaria_id,
            "referencia": data.referencia or "", "notas": data.notas or "",
            "usuario_id": user["id"], "usuario_nombre": user["name"], "created_at": iso_now()}
    pagos = list(compra.get("pagos", []) or []) + [pago]
    abonado = round(float(compra.get("abonado", 0) or 0) + monto, 2)
    estado = "pagada" if nuevo_saldo <= 0 else compra.get("estado", "confirmada")
    await db.compras.update_one({"id": compra_id}, {"$set": {
        "saldo_pendiente": nuevo_saldo, "abonado": abonado, "pagos": pagos, "estado": estado}})
    # Salida de efectivo si se paga en efectivo (afecta el arqueo de caja).
    caja = await caja_abierta_de(user["id"]) if data.metodo_pago == "efectivo" else None
    if caja:
        await db.caja_movimientos.insert_one({
            "id": uid(), "caja_id": caja["id"], "tipo": "gasto",
            "concepto": f"Pago CxP {compra.get('folio')} · {compra.get('proveedor_nombre') or 'S/D'}",
            "monto": monto, "referencia": compra.get("folio"),
            "usuario_id": user["id"], "usuario_nombre": user["name"], "fecha": iso_now()})
    await log_audit(user, "cxp_pagar", "compra", compra_id,
                    f"{compra.get('folio')} monto {monto} saldo {nuevo_saldo}")
    return {"id": compra_id, "folio": compra.get("folio"), "abonado": abonado,
            "saldo_pendiente": nuevo_saldo, "estado": estado, "pago": pago}

# =========================================================================
# REPORTES DE COMPRAS Y GASTOS
# =========================================================================
def _reporte_compras_agregado(docs: list, desde: str = "", hasta: str = "") -> dict:
    compras = [d for d in docs if d.get("tipo") != "gasto"]
    gastos = [d for d in docs if d.get("tipo") == "gasto"]
    total_compras = round(sum(float(d.get("total", 0) or 0) for d in compras), 2)
    total_gastos = round(sum(float(d.get("total", 0) or 0) for d in gastos), 2)

    por_proveedor = {}
    por_producto = {}
    gastos_cat = {}
    gastos_suc = {}
    gastos_cc = {}
    for d in docs:
        prov = d.get("proveedor_nombre") or "S/D"
        por_proveedor[prov] = round(por_proveedor.get(prov, 0) + float(d.get("total", 0) or 0), 2)
        if d.get("tipo") == "gasto":
            cat = d.get("categoria") or "Sin categoría"
            gastos_cat[cat] = round(gastos_cat.get(cat, 0) + float(d.get("total", 0) or 0), 2)
            suc = d.get("sucursal_id") or ""
            gastos_suc[suc] = round(gastos_suc.get(suc, 0) + float(d.get("total", 0) or 0), 2)
            cc = d.get("centro_costo_nombre") or ""
            gastos_cc[cc] = round(gastos_cc.get(cc, 0) + float(d.get("total", 0) or 0), 2)
        for it in d.get("items", []):
            if not it.get("afecta_inventario"):
                continue
            k = it.get("product_id") or it.get("codigo")
            e = por_producto.setdefault(k, {"product_id": it.get("product_id"),
                                            "codigo": it.get("codigo"),
                                            "descripcion": it.get("descripcion"),
                                            "cantidad": 0.0, "costo_total": 0.0})
            e["cantidad"] = round(e["cantidad"] + float(it.get("cantidad", 0) or 0), 3)
            e["costo_total"] = round(e["costo_total"] + float(it.get("importe") or 0) or
                                     float(it.get("cantidad", 0)) * float(it.get("costo", 0)), 2)
    hoy = now_utc().date().isoformat()
    pend = [d for d in docs if float(d.get("saldo_pendiente", 0) or 0) > 0]
    vencidas = [d for d in pend if (d.get("fecha_vencimiento") or "")[:10] and (d.get("fecha_vencimiento") or "")[:10] < hoy]
    pagos = []
    for d in docs:
        for p in d.get("pagos", []):
            pagos.append({"fecha": p.get("fecha"), "monto": p.get("monto"),
                          "metodo_pago": p.get("metodo_pago"), "folio": d.get("folio")})
    pagos.sort(key=lambda x: x.get("fecha") or "", reverse=True)
    return {
        "desde": desde, "hasta": hasta,
        "compras_periodo": total_compras, "gastos_periodo": total_gastos,
        "total_periodo": round(total_compras + total_gastos, 2),
        "facturas": len([d for d in docs if d.get("factura_numero")]),
        "por_proveedor": sorted(por_proveedor.items(), key=lambda x: x[1], reverse=True),
        "por_producto": sorted(por_producto.values(), key=lambda x: x["costo_total"], reverse=True)[:30],
        "gastos_por_categoria": sorted(gastos_cat.items(), key=lambda x: x[1], reverse=True),
        "gastos_por_sucursal": sorted(gastos_suc.items(), key=lambda x: x[1], reverse=True),
        "gastos_por_centro_costo": sorted(gastos_cc.items(), key=lambda x: x[1], reverse=True),
        "cxp_saldo": round(sum(float(d.get("saldo_pendiente", 0) or 0) for d in pend), 2),
        "vencidas": [{"folio": d.get("folio"), "proveedor_nombre": d.get("proveedor_nombre"),
                      "fecha_vencimiento": d.get("fecha_vencimiento"),
                      "saldo": float(d.get("saldo_pendiente", 0) or 0)} for d in vencidas],
        "vencidas_total": round(sum(float(d.get("saldo_pendiente", 0) or 0) for d in vencidas), 2),
        "pagos_realizados": pagos[:100],
        "pagos_total": round(sum(float(p["monto"]) for p in pagos), 2),
    }

@api.get("/compras/reportes")
async def compras_reportes(desde: Optional[str] = None, hasta: Optional[str] = None,
                           tipo: Optional[str] = None, proveedor: Optional[str] = None,
                           producto: Optional[str] = None, categoria: Optional[str] = None,
                           sucursal: Optional[str] = None, centro_costo: Optional[str] = None,
                           usuario: Optional[str] = None, estado: Optional[str] = None,
                           user: dict = Depends(get_current_user)):
    docs = await db.compras.find({"estado": {"$ne": "borrador"}}, {"_id": 0}).to_list(100000)
    if desde:
        docs = [d for d in docs if (d.get("fecha_recepcion") or "")[:10] >= desde]
    if hasta:
        docs = [d for d in docs if (d.get("fecha_recepcion") or "")[:10] <= hasta]
    if tipo and tipo != "todos":
        docs = [d for d in docs if d.get("tipo") == tipo]
    if proveedor:
        docs = [d for d in docs if d.get("proveedor_id") == proveedor or
                proveedor.lower() in (d.get("proveedor_nombre") or "").lower()]
    if producto:
        docs = [d for d in docs if any(producto.lower() in (it.get("descripcion") or "").lower() or
                producto.lower() == str(it.get("codigo") or "").lower() for it in d.get("items", []))]
    if categoria and categoria != "todos":
        docs = [d for d in docs if d.get("categoria") == categoria]
    if sucursal:
        docs = [d for d in docs if d.get("sucursal_id") == sucursal]
    if centro_costo:
        docs = [d for d in docs if d.get("centro_costo_id") == centro_costo or
                centro_costo.lower() in (d.get("centro_costo_nombre") or "").lower()]
    if usuario:
        docs = [d for d in docs if d.get("usuario_id") == usuario]
    if estado and estado != "todos":
        docs = [d for d in docs if d.get("estado") == estado]
    return {"registros": len(docs), **_reporte_compras_agregado(docs, desde, hasta)}

@api.get("/compras/reportes/export")
async def compras_reportes_export(fmt: str = "excel", desde: Optional[str] = None,
                                  hasta: Optional[str] = None, tipo: Optional[str] = None,
                                  proveedor: Optional[str] = None, producto: Optional[str] = None,
                                  categoria: Optional[str] = None, sucursal: Optional[str] = None,
                                  centro_costo: Optional[str] = None,
                                  user: dict = Depends(get_current_user)):
    docs = await db.compras.find({"estado": {"$ne": "borrador"}}, {"_id": 0}).to_list(100000)
    if desde:
        docs = [d for d in docs if (d.get("fecha_recepcion") or "")[:10] >= desde]
    if hasta:
        docs = [d for d in docs if (d.get("fecha_recepcion") or "")[:10] <= hasta]
    if tipo and tipo != "todos":
        docs = [d for d in docs if d.get("tipo") == tipo]
    if proveedor:
        docs = [d for d in docs if d.get("proveedor_id") == proveedor or
                proveedor.lower() in (d.get("proveedor_nombre") or "").lower()]
    if producto:
        docs = [d for d in docs if any(producto.lower() in (it.get("descripcion") or "").lower() or
                producto.lower() == str(it.get("codigo") or "").lower() for it in d.get("items", []))]
    if categoria and categoria != "todos":
        docs = [d for d in docs if d.get("categoria") == categoria]
    if sucursal:
        docs = [d for d in docs if d.get("sucursal_id") == sucursal]
    if centro_costo:
        docs = [d for d in docs if d.get("centro_costo_id") == centro_costo or
                centro_costo.lower() in (d.get("centro_costo_nombre") or "").lower()]
    rep = _reporte_compras_agregado(docs, desde, hasta)
    filtros = f"Periodo: {desde or 'inicio'} a {hasta or 'hoy'} | Registros: {len(docs)}"
    nombre = f"reporte_compras_{desde or 'inicio'}_{hasta or 'hoy'}.{fmt}"
    if fmt == "pdf":
        rows_prov = [[p, f"${m:,.2f}"] for p, m in rep["por_proveedor"][:20]]
        rows_cat = [[c, f"${m:,.2f}"] for c, m in rep["gastos_por_categoria"]]
        rows_prod = [[r["codigo"], r["descripcion"], r["cantidad"], f"${r['costo_total']:,.2f}"]
                     for r in rep["por_producto"]]
        pdf = _tabla_pdf_bytes(
            "Reporte de Compras y Gastos", filtros,
            [("Concepto", "$"), ("Proveedor", "Total")], rows_prov,
            extra_sections=[("Gastos por categoría", ("Categoría", "Total"), rows_cat),
                            ("Productos comprados", ("Código", "Descripción", "Cant.", "Costo total"), rows_prod)])
        media = "application/pdf"
    else:
        pdf = _reporte_compras_excel(rep)
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return StreamingResponse(io.BytesIO(pdf), media_type=media,
                             headers={"Content-Disposition": f"attachment; filename={nombre}"})

@api.get("/compras/{compra_id}")
async def compras_detail(compra_id: str, user: dict = Depends(get_current_user)):
    doc = await db.compras.find_one({"id": compra_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Compra no encontrada")
    return doc


@api.get("/compras/{compra_id}/pdf")
async def compra_pdf(compra_id: str, user: dict = Depends(get_current_user)):
    doc = await db.compras.find_one({"id": compra_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Compra no encontrada")
    proveedor = None
    if doc.get("proveedor_id"):
        proveedor = await db.proveedores.find_one({"id": doc["proveedor_id"]}, {"_id": 0})
    settings = await db.settings.find_one({"_id": "app"}) or {}
    data = storage.build_compra_pdf(doc, settings, proveedor)
    return StreamingResponse(io.BytesIO(data), media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename=compra-{doc.get("folio", compra_id)}.pdf'})

def _reporte_compras_excel(rep: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    hd = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="B3401E")
    def hoja(ws, cols, rows):
        ws.append(cols)
        for c in ws[1]:
            c.font = hd; c.fill = fill; c.alignment = Alignment(horizontal="center")
        for r in rows:
            ws.append(r)
        for col in ws.columns:
            mx = max(len(str(c.value or "")) for c in col) + 2
            ws.column_dimensions[col[0].column_letter].width = min(mx, 45)
    ws = wb.active; ws.title = "Resumen"
    hoja(ws, ["Métrica", "Valor"], [
        ["Compras del periodo", rep["compras_periodo"]], ["Gastos del periodo", rep["gastos_periodo"]],
        ["Total", rep["total_periodo"]], ["Facturas", rep["facturas"]],
        ["CxP saldo pendiente", rep["cxp_saldo"]], ["Vencidas total", rep["vencidas_total"]],
        ["Pagos realizados", rep["pagos_total"]],
    ])
    ws2 = wb.create_sheet("Por proveedor")
    hoja(ws2, ["Proveedor", "Total"], rep["por_proveedor"])
    ws3 = wb.create_sheet("Productos")
    hoja(ws3, ["Código", "Descripción", "Cantidad", "Costo total"],
         [[r["codigo"], r["descripcion"], r["cantidad"], r["costo_total"]] for r in rep["por_producto"]])
    ws4 = wb.create_sheet("Gastos por categoría")
    hoja(ws4, ["Categoría", "Total"], rep["gastos_por_categoria"])
    ws5 = wb.create_sheet("Cuentas por pagar")
    hoja(ws5, ["Folio", "Proveedor", "Vencimiento", "Saldo"],
         [[v["folio"], v["proveedor_nombre"], v["fecha_vencimiento"], v["saldo"]] for v in rep["vencidas"]])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def _tabla_pdf_bytes(titulo: str, meta: str, columns, rows, extra_sections=None) -> bytes:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.enums import TA_LEFT
    buf = io.BytesIO()
    st = getSampleStyleSheet()
    st["Title"].fontName = "Helvetica-Bold"; st["Title"].fontSize = 15
    st["Title"].textColor = colors.HexColor("#8B3A2A"); st["Title"].alignment = TA_LEFT
    subt = ParagraphStyle("subt", parent=st["Normal"], fontSize=8, textColor=colors.HexColor("#666666"), spaceAfter=6)
    doc = SimpleDocTemplate(buf, pagesize=letter)
    flow = []
    logo_path = os.path.join(os.path.dirname(__file__), "brand", "logotipo.png")
    if os.path.isfile(logo_path):
        try:
            flow.append(Image(logo_path, width=100, height=100 * (545 / 1157)))
        except Exception:
            pass
    flow.append(Paragraph(titulo, st["Title"]))
    flow.append(Paragraph(f"{meta} &nbsp;|&nbsp; Generado: {now_utc().strftime('%d/%m/%Y %H:%M')}", subt))
    flow.append(Spacer(1, 6))
    def add_tabla(cols, rows_):
        if not rows_:
            flow.append(Paragraph("Sin registros.", subt))
            return
        data = [list(cols)] + [[str(c or "") for c in r] for r in rows_]
        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#B3401E")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5EFEB")]),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        flow.append(t)
        flow.append(Spacer(1, 8))
    add_tabla(columns, rows)
    for name, cols, rows_ in (extra_sections or []):
        flow.append(Paragraph(f"<b>{name}</b>", subt))
        add_tabla(cols, rows_)
    doc.build(flow)
    buf.seek(0)
    return buf.getvalue()

# =========================================================================
# DASHBOARD
# =========================================================================
@api.get("/dashboard")
async def dashboard(user: dict = Depends(get_current_user)):
    now = now_utc()
    hoy = now.date().isoformat()
    mes = now.strftime("%Y-%m")
    # Modo: admin/propietario/dev ven el negocio completo; el resto solo su operación
    es_global = "*" in effective_permissions(user)
    mq = {"estado": "confirmada"}
    if not es_global:
        mq["usuario_id"] = user["id"]
    sales = await db.sales.find(mq, {"_id": 0}).to_list(5000)
    ventas_hoy = [s for s in sales if s["fecha"][:10] == hoy]
    ventas_mes = [s for s in sales if s["fecha"][:7] == mes]
    total_hoy = round(sum(s["total"] for s in ventas_hoy), 2)
    total_mes = round(sum(s["total"] for s in ventas_mes), 2)
    # Caja: la propia (operador) o el consolidado de todas las cajas abiertas (admin)
    caja = await caja_abierta_de(user["id"])
    total_caja = 0.0
    caja_info = None
    if es_global:
        cajas_abiertas = await db.cajas.find({"estado": "abierta"}, {"_id": 0}).to_list(500)
        for c in cajas_abiertas:
            movs = await db.caja_movimientos.find({"caja_id": c["id"]}, {"_id": 0}).to_list(1000)
            total_caja = round(total_caja + resumen_caja(c, movs)["efectivo_esperado"], 2)
    else:
        if caja:
            movs = await db.caja_movimientos.find({"caja_id": caja["id"]}, {"_id": 0}).to_list(1000)
            res = resumen_caja(caja, movs)
            total_caja = res["efectivo_esperado"]
            caja_info = {"id": caja["id"], "caja_nombre": caja.get("caja_nombre") or "Caja 1",
                         "fondo_inicial": caja.get("fondo_inicial"), "resumen": res}
        else:
            # Sin caja abierta: mostrar el último corte cerrado del usuario
            ultima = await db.cajas.find({"usuario_id": user["id"], "estado": "cerrada"},
                                         {"_id": 0}).sort("fecha_cierre", -1).to_list(1)
            if ultima:
                cierre = ultima[0].get("cierre") or {}
                total_caja = round(float(cierre.get("efectivo_esperado", 0)), 2)
                caja_info = {"id": ultima[0]["id"], "caja_nombre": ultima[0].get("caja_nombre") or "Caja 1",
                             "fondo_inicial": ultima[0].get("fondo_inicial"), "cerrada": True,
                             "resumen": cierre}
    # Productos (datos compartidos del negocio)
    products = await db.products.find({"estado": "activo"}, {"_id": 0}).to_list(5000)
    bajo_stock = [p for p in products if 0 < float(p.get("existencia", 0)) <= float(p.get("stock_minimo", 0))]
    sin_existencia = [p for p in products if float(p.get("existencia", 0)) <= 0]
    clientes = await db.clients.count_documents({})
    # ventas por dia ultimos 7 (alcance del dashboard)
    serie = []
    for i in range(6, -1, -1):
        d = (now - __import__("datetime").timedelta(days=i)).date().isoformat()
        tot = round(sum(s["total"] for s in sales if s["fecha"][:10] == d), 2)
        serie.append({"dia": d[5:], "total": tot})
    recientes = sorted(sales, key=lambda s: s["fecha"], reverse=True)[:8]
    base = {
        "ventas_hoy": total_hoy, "ventas_mes": total_mes,
        "num_ventas_hoy": len(ventas_hoy), "total_caja": total_caja,
        "bajo_stock": len(bajo_stock), "sin_existencia": len(sin_existencia),
        "clientes": clientes, "productos": len(products),
        "serie_ventas": serie, "mode": "global" if es_global else "propio",
        "caja_info": caja_info,
        "ventas_recientes": [{"folio": s["folio"], "cliente": s["cliente_nombre"],
                              "total": s["total"], "fecha": s["fecha"], "estado": s["estado"]} for s in recientes],
        "alertas_stock": [{"codigo": p["codigo"], "descripcion": p["descripcion"],
                           "existencia": p.get("existencia", 0), "stock_minimo": p.get("stock_minimo", 0)}
                          for p in (bajo_stock + sin_existencia)[:10]],
    }
    if es_global:
        base.update(await _dashboard_global_breakdown())
    return base


@api.get("/finanzas")
async def finanzas_resumen(desde: Optional[str] = None, hasta: Optional[str] = None,
                           user: dict = Depends(get_current_user)):
    """Resumen financiero consolidado con filtro por rango de fechas.
    Base para la página Finanzas y para la exportación a PDF/Excel."""
    es_global = "*" in effective_permissions(user)
    hoy = now_utc().date().isoformat()
    d = desde[:10] if desde else (hoy[0:8] + "01")  # default: mes actual
    h = hasta[:10] if hasta else hoy

    sales = await db.sales.find({"estado": "confirmada"}, {"_id": 0}).to_list(300000)
    if not es_global:
        sales = [s for s in sales if s.get("usuario_id") == user["id"]]
    sales = [s for s in sales if d <= (s.get("fecha") or "")[:10] <= h]

    ventas_total = round(sum(float(s.get("total", 0) or 0) for s in sales), 2)
    utilidad = 0.0
    for s in sales:
        for it in (s.get("items") or []):
            cant = float(it.get("cantidad", 0) or 0)
            costo = float(it.get("costo") or 0)
            importe = float(it.get("importe_bruto") or it.get("importe") or
                            (cant * float(it.get("precio") or 0)) or 0)
            utilidad += max(0, importe - cant * costo)
    utilidad = round(utilidad, 2)
    num_ventas = len(sales)
    efectivo = round(sum(float(p.get("monto", 0) or 0)
                         for s in sales for p in (s.get("pagos") or []) if p.get("metodo") == "efectivo"), 2)

    compras = await db.compras.find({"estado": "confirmada"}, {"_id": 0}).to_list(200000)
    compras = [c for c in compras if d <= (c.get("fecha_recepcion") or "")[:10] <= h]
    compras_total = round(sum(float(c.get("total", 0) or 0) for c in compras if c.get("tipo") != "gasto"), 2)
    gastos_total = round(sum(float(c.get("total", 0) or 0) for c in compras if c.get("tipo") == "gasto"), 2)

    cxv = await db.sales.find({"condicion": "credito", "estado": "confirmada", "saldo": {"$gt": 0}},
                              {"_id": 0, "saldo": 1, "fecha": 1, "cliente_id": 1}).to_list(200000)
    cartera = round(sum(float(c.get("saldo", 0) or 0) for c in cxv), 2)
    cli_dias = {}
    for cli in await db.clients.find({}, {"_id": 0, "id": 1, "dias_credito": 1}).to_list(200000):
        cli_dias[cli["id"]] = int(cli.get("dias_credito", 0) or 0)
    vencido = 0.0
    for c in cxv:
        try:
            dv, _ = _dias_vencido(c.get("fecha", ""), cli_dias.get(c.get("cliente_id"), 0), now_utc().date())
        except Exception:
            dv = 0
        if dv > 0:
            vencido += float(c.get("saldo", 0) or 0)
    vencido = round(vencido, 2)

    cxp_docs = await db.compras.find({"estado": {"$in": ["confirmada", "pagada"]}}, {"_id": 0}).to_list(100000)
    cxp_total = round(sum(float(c.get("saldo_pendiente", 0) or 0)
                          for c in cxp_docs if float(c.get("saldo_pendiente", 0) or 0) > 0), 2)

    # Serie mensual de ingresos vs gastos dentro del rango
    serie = {}
    serie_d = {}
    for s in sales:
        m = (s.get("fecha") or "")[:7]
        serie[m] = serie.get(m, 0) + float(s.get("total", 0) or 0)
    for c in compras:
        m = (c.get("fecha_recepcion") or "")[:7]
        serie_d[m] = serie_d.get(m, 0) + float(c.get("total", 0) or 0)
    serie_out = sorted({**serie, **serie_d}.keys())
    serie_rows = [{"mes": m, "ingresos": round(serie.get(m, 0), 2),
                   "egresos": round(serie_d.get(m, 0), 2)} for m in serie_out]

    return {
        "desde": d, "hasta": h,
        "ventas_total": ventas_total, "num_ventas": num_ventas,
        "utilidad_bruta": utilidad, "efectivo": efectivo,
        "compras_total": compras_total, "gastos_total": gastos_total,
        "cartera_cxc": cartera, "vencido_cxc": vencido,
        "cuentas_por_pagar": cxp_total,
        "resultado_neto": round(ventas_total - (compras_total + gastos_total), 2),
        "serie": serie_rows,
    }


@api.get("/finanzas/export")
async def finanzas_export(fmt: str = "excel", desde: Optional[str] = None, hasta: Optional[str] = None,
                          user: dict = Depends(require_permission("exportar"))):
    r = await finanzas_resumen(desde=desde, hasta=hasta, user=user)
    filtros = "Periodo: {} a {}".format(r["desde"], r["hasta"])
    if fmt == "pdf":
        headers = ["Métrica", "Valor"]
        rows = [
            ["Ingresos (ventas)", round(r["ventas_total"], 2)],
            ["N° ventas", r["num_ventas"]],
            ["Efectivo recibido", round(r["efectivo"], 2)],
            ["Utilidad bruta", round(r["utilidad_bruta"], 2)],
            ["Compras de mercancía", round(r["compras_total"], 2)],
            ["Gastos operativos", round(r["gastos_total"], 2)],
            ["Resultado neto", round(r["resultado_neto"], 2)],
            ["Cuentas por cobrar (cartera)", round(r["cartera_cxc"], 2)],
            ["Vencido CxC", round(r["vencido_cxc"], 2)],
            ["Cuentas por pagar", round(r["cuentas_por_pagar"], 2)],
        ]
        data = exports.pdf_bytes("REPORTE FINANCIERO", headers,
                                 [list(x) for x in rows], settings=await db.settings.find_one({"_id": "app"}) or {},
                                 user_name=user.get("name"), filtros=filtros,
                                 col_weights=[3, 2], wrap_cols=[0])
        media = "application/pdf"
        nombre = f"finanzas_{r['desde']}_{r['hasta']}.pdf"
    else:
        headers = ["Métrica", "Valor"]
        rows = [
            {"Métrica": "Ingresos (ventas)", "Valor": round(r["ventas_total"], 2)},
            {"Métrica": "N° ventas", "Valor": r["num_ventas"]},
            {"Métrica": "Efectivo recibido", "Valor": round(r["efectivo"], 2)},
            {"Métrica": "Utilidad bruta", "Valor": round(r["utilidad_bruta"], 2)},
            {"Métrica": "Compras de mercancía", "Valor": round(r["compras_total"], 2)},
            {"Métrica": "Gastos operativos", "Valor": round(r["gastos_total"], 2)},
            {"Métrica": "Resultado neto", "Valor": round(r["resultado_neto"], 2)},
            {"Métrica": "Cuentas por cobrar (cartera)", "Valor": round(r["cartera_cxc"], 2)},
            {"Métrica": "Vencido CxC", "Valor": round(r["vencido_cxc"], 2)},
            {"Métrica": "Cuentas por pagar", "Valor": round(r["cuentas_por_pagar"], 2)},
        ]
        data = exports.excel_bytes(rows, headers, sheet_name="Finanzas", title="REPORTE FINANCIERO - GRUPO RYSA")
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        nombre = f"finanzas_{r['desde']}_{r['hasta']}.xlsx"
    return StreamingResponse(io.BytesIO(data), media_type=media,
                             headers={"Content-Disposition": f"attachment; filename={nombre}"})


async def _dashboard_global_breakdown() -> dict:
    """Desglose global: ventas por caja y por usuario (solo admin/propietario)."""
    sales = await db.sales.find({"estado": "confirmada"}, {"_id": 0}).to_list(100000)
    # Por usuario
    por_usuario = {}
    for s in sales:
        uid_ = s.get("usuario_id") or "?"
        e = por_usuario.setdefault(uid_, {"usuario_id": uid_, "usuario_nombre": s.get("usuario_nombre") or "—",
                                          "num_ventas": 0, "total": 0.0})
        e["num_ventas"] += 1
        e["total"] = round(e["total"] + float(s.get("total", 0)), 2)
    # Por caja
    cajas = await db.cajas.find({}, {"_id": 0}).sort("fecha_apertura", -1).to_list(500)
    ventas_caja = {}
    for s in sales:
        cid = s.get("caja_id")
        if cid:
            e = ventas_caja.setdefault(cid, {"num_ventas": 0, "total": 0.0})
            e["num_ventas"] += 1
            e["total"] = round(e["total"] + float(s.get("total", 0)), 2)
    por_caja = []
    for c in cajas:
        res = {"efectivo_esperado": 0.0}
        v = ventas_caja.get(c["id"], {"num_ventas": 0, "total": 0.0})
        if c["estado"] == "abierta":
            movs = await db.caja_movimientos.find({"caja_id": c["id"]}, {"_id": 0}).to_list(1000)
            res = resumen_caja(c, movs)
        else:
            cierre = c.get("cierre") or {}
            res["efectivo_esperado"] = cierre.get("efectivo_esperado", 0.0)
        por_caja.append({
            "caja_id": c["id"], "caja_nombre": c.get("caja_nombre") or "Caja 1",
            "usuario_nombre": c.get("usuario_nombre") or "—", "estado": c["estado"],
            "fecha_apertura": c.get("fecha_apertura"), "fondo_inicial": c.get("fondo_inicial"),
            "efectivo_esperado": round(float(res.get("efectivo_esperado", 0)), 2),
            "ventas_total": round(float(v["total"]), 2), "num_ventas": v["num_ventas"],
        })
    sin_caja = {"num_ventas": 0, "total": 0.0}
    for s in sales:
        if not s.get("caja_id"):
            sin_caja["num_ventas"] += 1
            sin_caja["total"] = round(sin_caja["total"] + float(s.get("total", 0)), 2)
    if sin_caja["num_ventas"]:
        por_caja.append({"caja_id": None, "caja_nombre": "Sin caja", "usuario_nombre": "—",
                         "estado": "—", "fecha_apertura": "", "fondo_inicial": 0.0,
                         "efectivo_esperado": 0.0, "ventas_total": sin_caja["total"],
                         "num_ventas": sin_caja["num_ventas"]})
    return {
        "por_usuario": sorted(por_usuario.values(), key=lambda r: -r["total"]),
        "por_caja": por_caja,
        "totales_globales": {
            "ventas": round(sum(float(s.get("total", 0)) for s in sales), 2),
            "num_ventas": len(sales),
            "cajas_abiertas": sum(1 for c in cajas if c["estado"] == "abierta"),
        },
    }

# =========================================================================
# EXCEL - IMPORT / EXPORT
# =========================================================================
PROD_COLS = ["codigo", "descripcion", "linea", "clasificacion", "costo", "existencia",
             "unidad_medida", "stock_minimo", "estado"]
CLIENT_COLS = ["codigo", "nombre", "razon_social", "rfc", "telefono", "whatsapp",
               "correo", "direccion", "ciudad", "estado_geo", "cp", "tipo", "estado"]

def df_to_excel_bytes(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Datos")
    buf.seek(0)
    return buf.read()

def _product_export_query(estado=None, q=None, categoria=None, sku=None, linea=None,
                          unidad_medida=None, proveedor=None, min_costo=None,
                          max_costo=None, min_precio=None, max_precio=None):
    """Construye el query combinate de filtros de productos (coincide con la UI)."""
    query = {}
    if estado:
        query["estado"] = estado
    if categoria:
        query["clasificacion"] = categoria
    if linea:
        query["linea"] = linea
    if unidad_medida:
        query["unidad_medida"] = unidad_medida
    if min_costo is not None or max_costo is not None:
        r = {}
        if min_costo is not None:
            r["$gte"] = float(min_costo)
        if max_costo is not None:
            r["$lte"] = float(max_costo)
        query["costo"] = r
    if min_precio is not None or max_precio is not None:
        r = {}
        if min_precio is not None:
            r["$gte"] = float(min_precio)
        if max_precio is not None:
            r["$lte"] = float(max_precio)
        query["precios.0.precio_con_iva"] = r
    if sku:
        query["sku"] = {"$regex": sanitize_search_term(sku), "$options": "i"}
    if proveedor:
        rx = {"$regex": sanitize_search_term(proveedor), "$options": "i"}
        query["$and"] = query.get("$and", []) + [{"$or": [{"proveedor": rx}, {"proveedores": rx}]}]
    if q:
        rx = {"$regex": sanitize_search_term(q), "$options": "i"}
        qor = [{"codigo": rx}, {"descripcion": rx}, {"sku": rx}, {"linea": rx},
               {"clasificacion": rx}, {"sinonimos": rx}, {"codigos_barras": rx}]
        query["$and"] = query.get("$and", []) + [{"$or": qor}]
    return query


async def _all_products(estado=None, q=None, categoria=None, sku=None, linea=None,
                        unidad_medida=None, proveedor=None, min_costo=None, max_costo=None,
                        min_precio=None, max_precio=None, filtro=None):
    """Devuelve TODOS los productos que cumplen los filtros (sin paginación)."""
    query = _product_export_query(estado, q, categoria, sku, linea, unidad_medida,
                                  proveedor, min_costo, max_costo, min_precio, max_precio)
    docs = await db.products.find(query, {"_id": 0}).sort("descripcion", 1).to_list(100000)
    if filtro == "bajo_stock":
        docs = [p for p in docs if 0 < float(p.get("existencia", 0) or 0) <= float(p.get("stock_minimo", 0) or 0)]
    elif filtro == "sin_existencia":
        docs = [p for p in docs if float(p.get("existencia", 0) or 0) <= 0]
    return docs


def _product_export_rows(docs):
    rows = []
    for p in docs or []:
        precio = (p.get("precios") or [{}])
        rows.append({
            "ID": p.get("id"), "Código": p.get("codigo"), "SKU": p.get("sku"),
            "Código de barras": ", ".join(p.get("codigos_barras") or []),
            "Nombre": p.get("descripcion"), "Descripción": p.get("descripcion_larga") or "",
            "Categoría": p.get("clasificacion") or "", "Línea": p.get("linea") or "",
            "Clasificación": p.get("clasificacion") or "", "Costo": p.get("costo"),
            "Precio": (precio[0].get("precio_con_iva") if p.get("precios") else 0),
            "Existencia": p.get("existencia"), "Stock mínimo": p.get("stock_minimo"),
            "Unidad de medida": p.get("unidad_medida") or "PZA",
            "Proveedor": ", ".join(p.get("proveedores") or []) or p.get("proveedor") or "",
            "Sucursal": "", "Estado": p.get("estado") or "activo",
        })
    return rows


PRODUCT_EXPORT_HEADERS = ["ID", "Código", "SKU", "Código de barras", "Nombre", "Descripción",
                          "Categoría", "Línea", "Clasificación", "Costo", "Precio", "Existencia",
                          "Stock mínimo", "Unidad de medida", "Proveedor", "Sucursal", "Estado"]

PRODUCT_FILTER_LABEL = {
    "estado": "Estado", "categoria": "Categoría", "sku": "SKU", "linea": "Línea",
    "unidad_medida": "Unidad de medida", "proveedor": "Proveedor",
    "min_costo": "Costo desde", "max_costo": "Costo hasta",
    "min_precio": "Precio desde", "max_precio": "Precio hasta",
    "filtro": "Stock", "q": "Búsqueda",
}


async def _product_filtros_dict(estado, q, categoria, sku, linea, unidad_medida,
                                proveedor, min_costo, max_costo, min_precio, max_precio, filtro):
    vals = {
        "Estado": estado, "Categoría": categoria, "SKU": sku, "Línea": linea,
        "Unidad de medida": unidad_medida, "Proveedor": proveedor,
        "Búsqueda": q, "Stock": {"bajo_stock": "Bajo stock", "sin_existencia": "Sin existencia"}.get(filtro),
    }
    if min_costo is not None or max_costo is not None:
        vals["Rango de costo"] = f"{min_costo or 0} - {max_costo or '∞'}"
    if min_precio is not None or max_precio is not None:
        vals["Rango de precio"] = f"{min_precio or 0} - {max_precio or '∞'}"
    return {k: v for k, v in vals.items() if v not in (None, "", "all", 0, "0")}


@api.get("/products/export/excel")
async def export_products(estado: Optional[str] = None, q: Optional[str] = None,
                          categoria: Optional[str] = None, sku: Optional[str] = None,
                          linea: Optional[str] = None, unidad_medida: Optional[str] = None,
                          proveedor: Optional[str] = None, min_costo: Optional[float] = None,
                          max_costo: Optional[float] = None, min_precio: Optional[float] = None,
                          max_precio: Optional[float] = None, filtro: Optional[str] = None,
                          user: dict = Depends(require_permission("exportar"))):
    docs = await _all_products(estado, q, categoria, sku, linea, unidad_medida, proveedor,
                               min_costo, max_costo, min_precio, max_precio, filtro)
    data = exports.excel_bytes(_product_export_rows(docs), PRODUCT_EXPORT_HEADERS,
                               sheet_name="Productos", title="CATÁLOGO DE PRODUCTOS - GRUPO RYSA")
    return StreamingResponse(io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=productos.xlsx"})


@api.get("/products/export/pdf")
async def export_products_pdf(estado: Optional[str] = None, q: Optional[str] = None,
                              categoria: Optional[str] = None, sku: Optional[str] = None,
                              linea: Optional[str] = None, unidad_medida: Optional[str] = None,
                              proveedor: Optional[str] = None, min_costo: Optional[float] = None,
                              max_costo: Optional[float] = None, min_precio: Optional[float] = None,
                              max_precio: Optional[float] = None, filtro: Optional[str] = None,
                              user: dict = Depends(require_permission("exportar"))):
    docs = await _all_products(estado, q, categoria, sku, linea, unidad_medida, proveedor,
                               min_costo, max_costo, min_precio, max_precio, filtro)
    rows = _product_export_rows(docs)
    sett = await db.settings.find_one({"_id": "app"}) or {}
    filt = await _product_filtros_dict(estado, q, categoria, sku, linea, unidad_medida,
                                       proveedor, min_costo, max_costo, min_precio, max_precio, filtro)
    data = exports.pdf_bytes("REPORTE DE PRODUCTOS", PRODUCT_EXPORT_HEADERS,
                             [ [r[h] for h in PRODUCT_EXPORT_HEADERS] for r in rows ],
                             settings=sett, user_name=user.get("name"), filtros=filt,
                             col_weights=[1, 1.2, 1.2, 1.4, 3.2, 4.5, 1.4, 1.3, 1.4, 1.3, 1.3, 1.2, 1.2, 1.2, 2.6, 2.4, 1],
                             wrap_cols=[4, 5, 6, 7, 14, 15])
    return StreamingResponse(io.BytesIO(data), media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=productos.pdf"})


@api.get("/products/plantilla/excel")
async def plantilla_products(user: dict = Depends(get_current_user)):
    df = pd.DataFrame(columns=COL_ORDER)
    data = df_to_excel_bytes(df)
    return StreamingResponse(io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_productos_85.xlsx"})

def read_import_table(content: bytes, filename: str):
    """Lee XLSX/XLS/CSV de forma robusta. Devuelve un DataFrame (todo texto)."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        try:
            return pd.read_csv(io.BytesIO(content), dtype=str, keep_default_na=False)
        except Exception:
            return pd.read_csv(io.BytesIO(content), dtype=str, keep_default_na=False, sep=";", encoding="latin-1")
    errores = []
    engines = ["openpyxl", "xlrd"] if name.endswith(".xlsx") else ["xlrd", "openpyxl"]
    for eng in engines:
        try:
            return pd.read_excel(io.BytesIO(content), dtype=str, keep_default_na=False, engine=eng)
        except Exception as e:
            errores.append(f"{eng}: {str(e)[:80]}")
    # Algunos ERP exportan .xls que en realidad es HTML/XML
    try:
        dfs = pd.read_html(io.BytesIO(content))
        if dfs:
            return dfs[0].astype(str).fillna("")
    except Exception as e:
        errores.append(f"html: {str(e)[:80]}")
    try:
        return pd.read_csv(io.BytesIO(content), dtype=str, keep_default_na=False, sep=None, engine="python")
    except Exception as e:
        errores.append(f"csv: {str(e)[:80]}")
    raise HTTPException(400, "No se pudo leer el archivo. Formatos soportados: XLSX, XLS, CSV. " + " | ".join(errores[:2]))

@api.post("/products/import/preview")
async def import_preview(file: UploadFile = File(...), user: dict = Depends(require_permission("importar"))):
    content = await file.read()
    try:
        df = read_import_table(content, file.filename or "")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Error al procesar el archivo: {str(e)[:150]}")
    df.columns = [IMPORT_ALIASES.get(str(c).strip().upper(), str(c).strip().upper()) for c in df.columns]
    rows = df.to_dict("records")
    if not rows:
        raise HTTPException(400, "El archivo no contiene registros o no es un formato válido (XLSX, XLS, CSV).")
    if not any(str(c).strip().upper() == "CODIGO" for c in df.columns):
        raise HTTPException(400, "El archivo no tiene la columna CODIGO. Descarga la plantilla de 85 columnas.")
    all_codes = [str(r.get("CODIGO", "")).strip() for r in rows if str(r.get("CODIGO", "")).strip()]
    existentes_db = set()
    if all_codes:
        async for p in db.products.find({"codigo": {"$in": all_codes}}, {"_id": 0, "codigo": 1}):
            existentes_db.add(p["codigo"])
    preview, codigos = [], set()
    nuevos = existentes = con_errores = 0
    for i, r in enumerate(rows):
        canon = {k: r.get(k, "") for k in COL_ORDER}
        data, errores = parse_row(canon)
        codigo = data.get("codigo", "")
        if codigo and codigo in codigos:
            errores.append({"campo": "CODIGO", "valor": codigo, "motivo": "Código duplicado en el archivo"})
        codigos.add(codigo)
        existe = codigo in existentes_db
        if errores:
            con_errores += 1
        elif existe:
            existentes += 1
        else:
            nuevos += 1
        preview.append({"fila": i + 2, "codigo": codigo, "descripcion": data.get("descrip", ""),
                        "accion": "actualizar" if existe else "crear", "existe": existe,
                        "errores": errores, "data": data})
    return {"total": len(preview), "nuevos": nuevos, "existentes": existentes,
            "con_errores": con_errores, "columnas": COL_ORDER, "preview": preview}

@api.post("/products/import/confirm")
async def import_confirm(payload: dict, user: dict = Depends(require_permission("importar"))):
    rows = payload.get("rows", [])
    mode = payload.get("mode", "ambos")  # nuevos | actualizar | ambos
    actualizar_existencia = bool(payload.get("actualizar_existencia", False))
    creados = actualizados = omitidos = 0
    for row in rows:
        if row.get("errores"):
            omitidos += 1; continue
        d = row.get("data", {})
        codigo = str(d.get("codigo", "")).strip()
        if not codigo:
            omitidos += 1; continue
        existing = await db.products.find_one({"codigo": codigo})
        if existing and mode == "nuevos":
            omitidos += 1; continue
        if not existing and mode == "actualizar":
            omitidos += 1; continue
        doc = build_product_doc(d)
        doc["codigo"] = codigo
        doc = asegurar_codigo_como_barras(doc)
        doc["updated_at"] = iso_now()
        ex_val = doc.pop("existencia", None)
        if existing:
            await db.products.update_one({"codigo": codigo}, {"$set": doc})
            if actualizar_existencia and ex_val is not None:
                prod = await db.products.find_one({"codigo": codigo})
                diff = float(ex_val) - float(prod.get("existencia", 0))
                if abs(diff) > 0.0001:
                    if diff > 0:
                        await registrar_movimiento(prod, "ajuste", diff, 0, user, "Importación", "Ajuste por importación")
                    else:
                        await registrar_movimiento(prod, "ajuste", 0, -diff, user, "Importación", "Ajuste por importación")
            actualizados += 1
        else:
            doc["id"] = uid(); doc["existencia"] = 0; doc["created_at"] = iso_now()
            await db.products.insert_one(doc)
            if ex_val and float(ex_val) > 0:
                prod = await db.products.find_one({"id": doc["id"]})
                await registrar_movimiento(prod, "entrada", float(ex_val), 0, user, "Importación", "Inventario inicial")
            creados += 1
    if creados or actualizados:
        await log_audit(user, "importar", "producto", "", f"{creados} creados, {actualizados} actualizados, {omitidos} omitidos")
    return {"creados": creados, "actualizados": actualizados, "omitidos": omitidos}

def _client_export_query(q=None, estado=None, tipo=None, filtro=None, ciudad=None,
                         vendedor=None, rfc=None, telefono=None, fecha_desde=None, fecha_hasta=None):
    query = {}
    if estado:
        query["estado"] = estado
    if tipo:
        query["tipo"] = tipo
    if filtro:
        if filtro == "con_credito":
            query["credito_autorizado"] = True
        elif filtro == "sin_credito":
            query["credito_autorizado"] = {"$ne": True}
        elif filtro == "con_saldo":
            query["saldo"] = {"$gt": 0}
        elif filtro == "sin_saldo":
            query["$or"] = query.get("$or", []) + [{"saldo": {"$lte": 0}}, {"saldo": {"$exists": False}}]
        elif filtro in ("activo", "suspendido", "inactivo"):
            query["estado"] = filtro
        elif filtro == "con_ofertas":
            query["ofertas"] = True
        elif filtro == "sin_ofertas":
            query["ofertas"] = {"$ne": True}
    if ciudad:
        query["ciudad"] = {"$regex": sanitize_search_term(ciudad), "$options": "i"}
    if rfc:
        query["rfc"] = {"$regex": sanitize_search_term(rfc), "$options": "i"}
    if telefono:
        rx = {"$regex": sanitize_search_term(telefono), "$options": "i"}
        query["$and"] = query.get("$and", []) + [{"$or": [{"telefono": rx}, {"celular": rx},
                                                          {"tel_oficina": rx}, {"whatsapp": rx}]}]
    if vendedor:
        query["vendedor"] = {"$regex": sanitize_search_term(vendedor), "$options": "i"}
    if fecha_desde or fecha_hasta:
        r = {}
        if fecha_desde:
            r["$gte"] = fecha_desde
        if fecha_hasta:
            r["$lte"] = fecha_hasta + "T23:59:59"
        query["$and"] = query.get("$and", []) + [{"$or": [{"created_at": r}, {"fecha_alta": {"$gte": fecha_desde or "", "$lte": fecha_hasta or "9999"}}]}]
    if q:
        rx = {"$regex": sanitize_search_term(q), "$options": "i"}
        qor = [{"codigo": rx}, {"nombre": rx}, {"razon_social": rx}, {"rfc": rx},
               {"representa": rx}, {"telefono": rx}, {"tel_oficina": rx}, {"celular": rx},
               {"correo": rx}, {"correos": rx}, {"ciudad": rx}, {"estado_geo": rx}]
        query["$and"] = query.get("$and", []) + [{"$or": qor}]
    return query


async def _all_clients(q=None, estado=None, tipo=None, filtro=None, ciudad=None,
                       vendedor=None, rfc=None, telefono=None, fecha_desde=None, fecha_hasta=None):
    query = _client_export_query(q=q, estado=estado, tipo=tipo, filtro=filtro, ciudad=ciudad,
                                 vendedor=vendedor, rfc=rfc, telefono=telefono,
                                 fecha_desde=fecha_desde, fecha_hasta=fecha_hasta)
    return await db.clients.find(query, {"_id": 0}).sort("nombre", 1).to_list(100000)


CLIENT_EXPORT_HEADERS = ["ID", "Código", "Nombre", "Razón social", "RFC", "Teléfono", "Email",
                         "Dirección", "Ciudad", "Estado", "Sucursal", "Vendedor", "Tipo de cliente",
                         "Crédito", "Saldo", "Estado", "Fecha de registro"]

TIPO_CLIENTE_LABEL = {
    "publico": "Público General", "menudeo": "Menudeo", "mayoreo": "Mayoreo", "especial": "Especial",
}


def _client_export_rows(clients):
    rows = []
    for c in clients or []:
        direc = ", ".join(filter(None, [c.get("direccion"), c.get("colonia"),
                                        c.get("numero_exterior"), c.get("cp")]))
        rows.append({
            "ID": c.get("id"), "Código": c.get("codigo"), "Nombre": c.get("nombre"),
            "Razón social": c.get("razon_social") or "",
            "RFC": c.get("rfc") or "",
            "Teléfono": c.get("telefono") or c.get("celular") or c.get("whatsapp") or "",
            "Email": c.get("correo") or c.get("correos") or "",
            "Dirección": direc, "Ciudad": c.get("ciudad") or "",
            "Estado": c.get("estado_geo") or "",
            "Sucursal": c.get("almacen") or "",
            "Vendedor": c.get("vendedor") or "",
            "Tipo de cliente": TIPO_CLIENTE_LABEL.get(c.get("tipo"), c.get("tipo")) or "",
            "Crédito": "Sí" if c.get("credito_autorizado") else "No",
            "Saldo": round(float(c.get("saldo", 0) or 0), 2),
            "Estado": c.get("estado") or "activo",
            "Fecha de registro": (c.get("created_at") or c.get("fecha_alta") or "")[:10],
        })
    return rows


async def _client_filtros_dict(q, estado, tipo, filtro, ciudad, vendedor, rfc,
                               telefono, fecha_desde, fecha_hasta):
    FILTRO_LABEL = {
        "con_credito": "Con crédito", "sin_credito": "Sin crédito", "con_saldo": "Con saldo",
        "sin_saldo": "Sin saldo", "activo": "Activos", "suspendido": "Suspendidos",
        "inactivo": "Inactivos", "con_ofertas": "Con ofertas", "sin_ofertas": "Sin ofertas",
    }
    vals = {
        "Búsqueda": q, "Estado": estado,
        "Tipo": TIPO_CLIENTE_LABEL.get(tipo, tipo),
        "Filtro rápido": FILTRO_LABEL.get(filtro), "Ciudad": ciudad, "Vendedor": vendedor,
        "RFC": rfc, "Teléfono": telefono,
        "Registro desde": fecha_desde, "Registro hasta": fecha_hasta,
    }
    return {k: v for k, v in vals.items() if v not in (None, "", "all")}


@api.get("/clients/export/excel")
async def export_clients(q: Optional[str] = None, estado: Optional[str] = None, tipo: Optional[str] = None,
                         filtro: Optional[str] = None, ciudad: Optional[str] = None,
                         vendedor: Optional[str] = None, rfc: Optional[str] = None,
                         telefono: Optional[str] = None, fecha_desde: Optional[str] = None,
                         fecha_hasta: Optional[str] = None,
                         user: dict = Depends(require_permission("exportar"))):
    clients = await _all_clients(q=q, estado=estado, tipo=tipo, filtro=filtro, ciudad=ciudad,
                                 vendedor=vendedor, rfc=rfc, telefono=telefono,
                                 fecha_desde=fecha_desde, fecha_hasta=fecha_hasta)
    data = exports.excel_bytes(_client_export_rows(clients), CLIENT_EXPORT_HEADERS,
                               sheet_name="Clientes", title="CARTERA DE CLIENTES - GRUPO RYSA")
    return StreamingResponse(io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=clientes.xlsx"})


@api.get("/clients/export/pdf")
async def export_clients_pdf(q: Optional[str] = None, estado: Optional[str] = None, tipo: Optional[str] = None,
                             filtro: Optional[str] = None, ciudad: Optional[str] = None,
                             vendedor: Optional[str] = None, rfc: Optional[str] = None,
                             telefono: Optional[str] = None, fecha_desde: Optional[str] = None,
                             fecha_hasta: Optional[str] = None,
                             user: dict = Depends(require_permission("exportar"))):
    clients = await _all_clients(q=q, estado=estado, tipo=tipo, filtro=filtro, ciudad=ciudad,
                                 vendedor=vendedor, rfc=rfc, telefono=telefono,
                                 fecha_desde=fecha_desde, fecha_hasta=fecha_hasta)
    rows = _client_export_rows(clients)
    sett = await db.settings.find_one({"_id": "app"}) or {}
    filt = await _client_filtros_dict(q, estado, tipo, filtro, ciudad, vendedor, rfc,
                                      telefono, fecha_desde, fecha_hasta)
    data = exports.pdf_bytes("REPORTE DE CLIENTES", CLIENT_EXPORT_HEADERS,
                             [ [r[h] for h in CLIENT_EXPORT_HEADERS] for r in rows ],
                             settings=sett, user_name=user.get("name"), filtros=filt,
                             col_weights=[1, 1.2, 3, 3, 1.6, 1.6, 2.4, 4, 1.8, 1.6, 2, 2, 1.8, 1.3, 1.5, 1.2, 1.8],
                             wrap_cols=[2, 3, 6, 7, 8, 10, 11])
    return StreamingResponse(io.BytesIO(data), media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=clientes.pdf"})


@api.get("/clients/plantilla/excel")
async def plantilla_clients(user: dict = Depends(get_current_user)):
    df = pd.DataFrame(columns=CLIENT_LEGACY_ORDER)
    data = df_to_excel_bytes(df)
    return StreamingResponse(io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_clientes.xlsx"})

@api.post("/clients/import/preview")
async def import_clients_preview(file: UploadFile = File(...), user: dict = Depends(require_permission("importar"))):
    content = await file.read()
    df = read_import_table(content, file.filename or "").fillna("")
    df.columns = [str(c).strip().upper() for c in df.columns]
    rows = df.to_dict("records")
    if not rows:
        raise HTTPException(400, "El archivo no contiene registros.")
    if not any(c == "CLAVE" for c in df.columns) and not any(c == "NOMBRE" for c in df.columns):
        raise HTTPException(400, "El archivo no tiene columnas CLAVE/NOMBRE. Descarga la plantilla de clientes.")
    all_codes = [str(r.get("CLAVE", "")).strip() for r in rows if str(r.get("CLAVE", "")).strip()]
    existentes_db = set()
    if all_codes:
        async for c in db.clients.find({"codigo": {"$in": all_codes}}, {"_id": 0, "codigo": 1}):
            existentes_db.add(c["codigo"])
    preview, vistos = [], set()
    nuevos = existentes = con_errores = 0
    for i, r in enumerate(rows):
        data, errores = parse_client_row(r)
        clave = data.get("codigo", "")
        if clave and clave in vistos:
            errores.append({"campo": "CLAVE", "valor": clave, "motivo": "CLAVE duplicada en el archivo"})
        vistos.add(clave)
        existe = clave in existentes_db
        if errores:
            con_errores += 1
        elif existe:
            existentes += 1
        else:
            nuevos += 1
        preview.append({"fila": i + 2, "clave": clave, "nombre": data.get("nombre", ""),
                        "accion": "actualizar" if existe else "crear", "existe": existe,
                        "errores": errores, "data": data})
    return {"total": len(preview), "nuevos": nuevos, "existentes": existentes,
            "con_errores": con_errores, "columnas": CLIENT_LEGACY_ORDER,
            "mapeo_columnas": {
                "archivo": [c for c in df.columns if c],
                "reconocidas": [c for c in df.columns if c in CLIENT_IMPORT_MAP],
                "ignoradas": [c for c in df.columns
                              if c and c not in CLIENT_IMPORT_MAP
                              and c not in ("CLAVE",)],
            },
            "preview": preview}

@api.post("/clients/import/confirm")
async def import_clients(payload: dict, user: dict = Depends(require_permission("importar"))):
    rows = payload.get("rows", [])
    mode = payload.get("mode", "ambos")  # nuevos | actualizar | ambos
    actualizar_saldo = bool(payload.get("actualizar_saldo", False))
    creados = actualizados = omitidos = 0
    for r in rows:
        if r.get("errores"):
            omitidos += 1
            continue
        data = r.get("data", {})
        clave = str(data.get("codigo", "")).strip()
        if not clave:
            omitidos += 1
            continue
        existing = await db.clients.find_one({"codigo": clave})
        if existing:
            if mode == "nuevos":
                omitidos += 1
                continue
            doc = normalize_client_doc(dict(data))
            doc.pop("id", None)
            if actualizar_saldo:
                doc["saldo"] = float(data.get("saldo") or 0.0)  # carga de saldo inicial desde archivo
            else:
                doc.pop("saldo", None)  # por defecto el saldo lo gobiernan las ventas
            doc["codigo"] = clave
            doc["updated_at"] = iso_now()
            await db.clients.update_one({"codigo": clave}, {"$set": doc})
            actualizados += 1
        else:
            if mode == "actualizar":
                omitidos += 1
                continue
            doc = normalize_client_doc(dict(data))
            doc["id"] = uid()
            doc["codigo"] = clave
            doc.setdefault("tipo", "publico")
            doc.setdefault("estado", "activo")
            doc["saldo"] = float(data.get("saldo") or 0.0)
            doc["created_at"] = iso_now()
            await db.clients.insert_one(doc)
            creados += 1
    if creados or actualizados:
        await log_audit(user, "importar", "cliente", "", f"{creados} creados, {actualizados} actualizados, {omitidos} omitidos")
    return {"creados": creados, "actualizados": actualizados, "omitidos": omitidos}

# =========================================================================
# EXPORTACIÓN / IMPORTACIÓN GLOBAL DE DATOS (solo administradores)
# Formato: ZIP con manifiesto de versión + un JSON por entidad.
# =========================================================================
import zipfile

EXPORT_COLLECTIONS = ["sucursales", "users", "products", "clients", "sales",
                      "cajas", "caja_movimientos", "inventory_movements", "abonos",
                      "cfdi_documents", "categories", "suspended_sales", "settings",
                      "mensajes", "plantillas"]
EXPORT_SKIP_SECRET_KEYS = {"password_hash", "api_key", "api_password", "api_user",
                           "token", "secret", "csrf", "refresh_token"}

def _sanitize_export(doc: dict) -> dict:
    out = {}
    for k, v in (doc or {}).items():
        if str(k).lower() in EXPORT_SKIP_SECRET_KEYS:
            continue
        out[k] = v
    return out

@api.get("/datos/export")
async def exportar_datos(user: dict = Depends(require_permission("exportar"))):
    """Exporta la BD a un ZIP (manifiesto + JSON por entidad). Solo admin."""
    if not user_has_permission(user, "config"):
        raise HTTPException(403, "La exportación global requiere rol administrador")
    buf = io.BytesIO()
    payload = {"formato": "rysa-datos", "version": 2, "fecha": iso_now(), "colecciones": {}}
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for col in EXPORT_COLLECTIONS:
            try:
                docs = await getattr(db, col).find({}, {"_id": 0}).to_list(1000000)
                payload["colecciones"][col] = len(docs)
                z.writestr(f"{col}.json", json.dumps([_sanitize_export(d) for d in docs], ensure_ascii=False, default=str))
            except Exception as e:
                payload["colecciones"][col] = f"error: {e}"
        z.writestr("manifest.json", json.dumps(payload, ensure_ascii=False))
    buf.seek(0)
    await log_audit(user, "exportar_datos", "sistema", "", "Exportación global de datos (admin)")
    return StreamingResponse(buf, media_type="application/zip",
                             headers={"Content-Disposition": 'attachment; filename="rysa_export.zip"'})

@api.post("/datos/import")
async def importar_datos(file: UploadFile = File(...), user: dict = Depends(require_permission("config"))):
    """Importa un ZIP de rysa_export: valida versión/integridad y restaura por entidad.
    NOTA: excluye `users` y `settings` (para evitar escalamiento de privilegios desde un archivo)."""
    data = await file.read()
    if len(data) > 100 * 1024 * 1024:
        raise HTTPException(400, "Archivo demasiado grande")
    try:
        z = zipfile.ZipFile(io.BytesIO(data))
    except zipfile.BadZipFile:
        raise HTTPException(400, "Archivo ZIP inválido")
    if "manifest.json" not in z.namelist():
        raise HTTPException(400, "Manifiesto no encontrado: no es un respaldo de RYSA")
    manifest = json.loads(z.read("manifest.json"))
    if manifest.get("formato") != "rysa-datos":
        raise HTTPException(400, "Formato de respaldo no válido")
    if int(manifest.get("version", 1)) < 1:
        raise HTTPException(400, "Versión de respaldo no soportada")
    resultados = {}
    for col in EXPORT_COLLECTIONS:
        if col in ("users", "settings"):
            continue  # no se restauran usuarios (escalamiento) ni settings sensibles
        if f"{col}.json" not in z.namelist():
            continue
        docs = json.loads(z.read(f"{col}.json"))
        n0 = await getattr(db, col).count_documents({})
        creados = 0
        for d in docs:
            if not d.get("id") and not d.get("codigo"):
                continue
            try:
                flt = {"id": d["id"]} if d.get("id") else {"codigo": d["codigo"]}
                if await getattr(db, col).find_one(flt):
                    continue
                await getattr(db, col).insert_one(d)
                creados += 1
            except Exception:
                continue
        resultados[col] = {"previos": n0, "creados": creados}
    await log_audit(user, "importar_datos", "sistema", "", "Importación global de datos (admin)")
    return {"ok": True, "resumen": resultados}

# =========================================================================
# AUDITORÍA
# =========================================================================
@api.get("/audit")
async def audit(user: dict = Depends(require_permission("auditoria.ver"))):
    return await db.audit_logs.find({}, {"_id": 0}).sort("fecha", -1).to_list(300)

# =========================================================================
# HERRAMIENTAS DE DESARROLLADOR / MANTENIMIENTO (solo admin_desarrollador)
# En producción estos endpoints NO existen (404), independientemente del rol.
# =========================================================================
def _dev_only(user: dict = Depends(require_permission("dev.errores"))):
    if _APP_ENV == "production":
        raise HTTPException(status_code=404, detail="No encontrado")
    return user

@api.get("/dev/errores")
async def dev_errores(user: dict = Depends(_dev_only)):
    return {"errores": list(reversed(DEV_ERRORS))}

@api.delete("/dev/errores")
async def dev_limpiar_errores(user: dict = Depends(_dev_only)):
    DEV_ERRORS.clear()
    await log_audit(user, "dev_limpiar", "dev", "errores", "Bitácora de errores limpiada")
    return {"ok": True}

@api.get("/dev/info")
async def dev_info(user: dict = Depends(_dev_only)):
    counts = {}
    for col in ("users", "products", "clients", "sales", "cajas", "caja_movimientos",
                "abonos", "audit_logs", "counters"):
        try:
            counts[col] = await getattr(db, col).count_documents({})
        except Exception:
            counts[col] = 0
    return {
        "entorno": _APP_ENV,
        "python": platform.python_version(),
        "fecha_servidor": iso_now(),
        "roles": {r: sorted(p) for r, p in ROLE_PERMISSIONS.items()},
        "colecciones": counts,
        "errores_en_memoria": len(DEV_ERRORS),
        "admin_system_roles": sorted(ADMIN_SYSTEM_ROLES),
    }


# --- Datos demo de fuerza de ventas (SOLO DEV) --------------------------------
# Crea cuentas de vendedores + clientes con GPS en Palenque, Chiapas, y
# simula el track de ubicaciones del día + visitas. Idempotente: si los
# vendedores demo ya existen solo regenera el track de HOY.
_PALENQUE = (17.5095, -91.9827)  # centro de Palenque, Chiapas

_DEMO_VENDEDORES = [
    {"email": "ramiro.demo@rysa.dev", "name": "Ramiro Gómez"},
    {"email": "lucia.demo@rysa.dev", "name": "Lucía Hernández"},
    {"email": "pedro.demo@rysa.dev", "name": "Pedro Cruz"},
]
_DEMO_PASSWORD = "DemoVendedor2026"

_DEMO_CLIENTES = [
    ("Tienda La Esperanza", 0.004, 0.003), ("Abarrotes El Progreso", -0.005, 0.006),
    ("Miscelánea Doña Mary", 0.007, -0.004), ("Farmacia San Juan", -0.008, -0.003),
    ("Bodega Palenque Norte", 0.011, 0.009), ("Papelería Central", -0.002, 0.012),
    ("Restaurante El Folclor", 0.009, -0.010), ("Distribuidora Maya", -0.012, 0.001),
    ("Tiendita El Ahorro", 0.003, -0.013), ("Comercial Pakal", -0.006, -0.011),
    ("Vinatería Los Comales", 0.013, -0.001), ("Abarrotes Pakal Na", -0.010, -0.008),
]


@api.post("/dev/seed-campo")
async def dev_seed_campo(regenerar_track: bool = True, user: dict = Depends(_dev_only)):
    """SIMULACIÓN (solo desarrollo): vendedores, clientes con GPS en Palenque,
    track de ubicaciones de HOY y visitas. No crea ventas ni afecta finanzas."""
    import random
    random.seed(42)
    lat0, lng0 = _PALENQUE
    hoy = iso_now()[:10]
    resumen = {"vendedores": [], "clientes_creados": 0, "ubicaciones_hoy": 0,
               "visitas_creadas": 0}

    # 1) Vendedores demo (rol vendedor, activos).
    vids = {}
    for dv in _DEMO_VENDEDORES:
        ex = await db.users.find_one({"email": dv["email"]}, {"_id": 0})
        if not ex:
            doc = {"id": uid(), "email": dv["email"], "name": dv["name"],
                   "password_hash": hash_password(_DEMO_PASSWORD), "role": "vendedor",
                   "active": True, "token_version": 0,
                   "modulos": ["clientes"], "created_at": iso_now()}
            await db.users.insert_one(doc)
            ex = doc
            resumen["vendedores"].append(f"{dv['name']} (NUEVO · {_DEMO_PASSWORD})")
        else:
            resumen["vendedores"].append(f"{dv['name']} (ya existía)")
        vids[ex["id"]] = ex["name"]
    vid_list = list(vids.keys())

    # 2) Clientes demo con coordenadas en Palenque, repartidos entre vendedores.
    for i, (nombre, dlat, dlng) in enumerate(_DEMO_CLIENTES):
        codigo = f"DEMO-{i+1:03d}"
        if await db.clients.find_one({"codigo": codigo}):
            continue
        vid = vid_list[i % len(vid_list)]
        await db.clients.insert_one({
            "id": uid(), "codigo": codigo, "nombre": nombre,
            "estado": "activo", "tipo": "menudeo",
            "telefono": f"916{100000+i}", "ciudad": "Palenque",
            "direccion": "Zona demo Palenque, Chiapas",
            "latitud": round(lat0 + dlat, 6), "longitud": round(lng0 + dlng, 6),
            "vendedor_id": vid, "vendedor": vids[vid],
            "dias_credito": random.choice([8, 15, 30]),
            "credito_autorizado": True, "limite_credito": 5000, "saldo": 0,
            "lista_precios": 1, "created_at": iso_now()})
        resumen["clientes_creados"] += 1

    clientes = await db.clients.find({"codigo": {"$regex": "^DEMO-"}}, {"_id": 0}).to_list(100)

    # 3) Track GPS de HOY por vendedor (08:30 → hora actual, entre clientes).
    if regenerar_track:
        now_h = now_utc().hour
        for vid in vid_list:
            # limpiar track previo de hoy para no duplicar
            previos = await db.seller_locations.find(
                {"vendedor_id": vid, "fecha": {"$regex": "^" + hoy}}, {"_id": 0}).to_list(500)
            for p in previos:
                await db.seller_locations.delete_one({"id": p["id"]})
            mis_clientes = [c for c in clientes if c.get("vendedor_id") == vid] or clientes
            puntos = []
            hora_f = 8.5
            paso_min = 35
            while hora_f <= min(now_h + random.random(), 19.0):
                c = random.choice(mis_clientes)
                jitter = lambda v: round(v + random.uniform(-0.0006, 0.0006), 6)  # noqa: E731
                hh = int(hora_f); mm = int((hora_f - hh) * 60)
                puntos.append({
                    "id": uid(), "vendedor_id": vid,
                    "latitud": jitter(c.get("latitud") or lat0),
                    "longitud": jitter(c.get("longitud") or lng0),
                    "precision": round(random.uniform(5, 25), 1),
                    "fuente": "gps",
                    "velocidad_kmh": round(random.uniform(0, 45), 1),
                    "bateria_pct": max(15, 95 - int(hora_f) * 5),
                    "fecha": f"{hoy}T{hh:02d}:{mm:02d}:{random.randint(10,59):02d}",
                })
                hora_f += paso_min / 60.0
                paso_min = random.randint(20, 50)
            for p in reversed(puntos):  # insertar en orden cronológico
                await db.seller_locations.insert_one(p)
            resumen["ubicaciones_hoy"] += len(puntos)

    # 4) Visitas de hoy (algunas realizadas, otras programadas más tarde).
    for vid in vid_list:
        ya = await db.visits.find(
            {"vendedor_id": vid, "fecha": {"$regex": "^" + hoy}}, {"_id": 0}).to_list(200)
        if ya:
            continue
        mis = [c for c in clientes if c.get("vendedor_id") == vid] or clientes
        for k, c in enumerate(random.sample(mis, min(3, len(mis)))):
            realizada = k == 0
            estado = "realizada" if realizada else "programada"
            await db.visits.insert_one({
                "id": uid(), "cliente_id": c["id"], "cliente_nombre": c.get("nombre"),
                "vendedor_id": vid, "vendedor_nombre": vids[vid],
                "tipo": random.choice(["visita", "cobro", "seguimiento"]),
                "estado": estado,
                "fecha": f"{hoy}T{8+k}:00:00",
                "fecha_programada": f"{hoy}T{14+k}:00:00",
                "comentarios": "Visita demo generada automáticamente." if realizada else "",
                "checkin": ({"latitud": c.get("latitud"), "longitud": c.get("longitud"),
                             "hora": f"{hoy}T{8+k}:2{i}:00"} if realizada else None),
                "usuario_id": user["id"], "created_at": iso_now(),
                "updated_at": iso_now()})
            resumen["visitas_creadas"] += 1

    return {"ok": True, **resumen,
            "login_demo": [{"email": e, "password": _DEMO_PASSWORD} for _, e in
                           [(v["name"], v["email"]) for v in _DEMO_VENDEDORES]]}


# --- Diagnóstico del sistema (SOLO DEV) ---------------------------------------
@api.get("/dev/diagnostico")
async def dev_diagnostico(user: dict = Depends(_dev_only)):
    """Salud del sistema: latencia de BD, almacenamiento escribible,
    integridad referencial básica y resumen del entorno."""
    diag = {"generado": iso_now(), "entorno": _APP_ENV,
            "python": platform.python_version(),
            "errores_en_memoria": len(DEV_ERRORS)}

    # 1) Base de datos: latencia de una consulta trivial.
    t0 = time.perf_counter()
    try:
        await db.counters.find_one({})
        diag["bd"] = {"ok": True,
                      "latencia_ms": round((time.perf_counter() - t0) * 1000, 1)}
    except Exception as e:
        diag["bd"] = {"ok": False, "error": str(e)[:300]}

    # 2) Almacenamiento local: escribir / leer / borrar un archivo de prueba.
    try:
        storage.init_storage()
        base = Path(storage.base_upload_dir())
        probe = base / ".__diagnostico_rysa.tmp"
        probe.write_bytes(b"rysa-diag")
        leidos = probe.read_bytes()
        probe.unlink()
        diag["storage"] = {"ok": leidos == b"rysa-diag", "ruta": str(base)}
    except Exception as e:
        diag["storage"] = {"ok": False, "error": str(e)[:300]}

    # 3) Integridad referencial básica (escaneos acotados).
    try:
        client_ids = {c["id"] for c in await db.clients.find({}, {"id": 1}).to_list(20000)}
        user_ids = {u["id"] for u in await db.users.find({}, {"id": 1}).to_list(20000)}
        ventas = await db.sales.find({}, {"cliente_id": 1}).to_list(5000)
        visitas = await db.visits.find({}, {"cliente_id": 1, "vendedor_id": 1}).to_list(5000)
        clientes = await db.clients.find({}, {"vendedor_id": 1}).to_list(20000)
        diag["integridad"] = {
            "ventas_sin_cliente": sum(1 for v in ventas if v.get("cliente_id") and v["cliente_id"] not in client_ids),
            "visitas_sin_cliente": sum(1 for v in visitas if v.get("cliente_id") and v["cliente_id"] not in client_ids),
            "visitas_sin_vendedor": sum(1 for v in visitas if v.get("vendedor_id") and v["vendedor_id"] not in user_ids),
            "clientes_sin_vendedor": sum(1 for c in clientes if c.get("vendedor_id") and c["vendedor_id"] not in user_ids),
            "muestra": {"ventas": len(ventas), "visitas": len(visitas)},
        }
    except Exception as e:
        diag["integridad"] = {"error": str(e)[:300]}
    return diag


# --- Checklist previo a producción (SOLO DEV) ----------------------------------
@api.get("/dev/preproduccion")
async def dev_preproduccion(user: dict = Depends(_dev_only)):
    """Evalúa condiciones recomendadas ANTES de pasar ENVIRONMENT=production.
    El propio módulo desarrollador se desactiva en producción (404)."""
    checks = []

    def add(cid, titulo, ok, detalle, severidad="alta"):
        checks.append({"id": cid, "titulo": titulo, "ok": bool(ok),
                       "detalle": detalle, "severidad": severidad})

    # 1) Sin cuentas demo y sin su contraseña por defecto.
    demos = await db.users.find({"email": {"$regex": "@rysa\\.dev$"}},
                                {"_id": 0}).to_list(100)
    pw_demo = [u.get("email") for u in demos
               if verify_password(_DEMO_PASSWORD, u.get("password_hash") or "")]
    add("usuarios_demo", "Sin cuentas demo (@rysa.dev)", len(demos) == 0,
        f"{len(demos)} cuenta(s) demo" +
        (f" · {len(pw_demo)} con contraseña por defecto" if pw_demo else ""))

    # 2) Sin clientes de prueba.
    n_cli = await db.clients.count_documents({"codigo": {"$regex": "^DEMO-"}})
    add("clientes_demo", "Sin clientes de prueba (DEMO-*)", n_cli == 0,
        f"{n_cli} cliente(s) de prueba")

    # 3) Bitácora de errores vacía.
    add("errores", "Bitácora de errores vacía", len(DEV_ERRORS) == 0,
        f"{len(DEV_ERRORS)} error(es) pendientes", severidad="media")

    # 4) Configuración básica del ERP.
    s = await db.settings.find_one({"_id": "app"}, {"_id": 0}) or {}
    empresa = (s.get("empresa_nombre") or "").strip()
    add("empresa", "Nombre de la empresa configurado", bool(empresa),
        empresa or "settings.empresa_nombre vacío")
    add("logo", "Logo corporativo cargado", bool(s.get("logo_url")),
        "definido" if s.get("logo_url") else "sin logo (recomendado)",
        severidad="baja")

    # 5) Catálogo sano: productos sin código o sin precio de venta.
    prods = await db.products.find({}, {"codigo": 1, "CODIGO": 1,
                                        "PRECIO1": 1, "precio1": 1}).to_list(20000)

    def _cod(p):
        return ((p.get("codigo") or p.get("CODIGO")) or "").strip()

    def _precio(p):
        try:
            return float(p.get("PRECIO1") or p.get("precio1") or 0) > 0
        except (TypeError, ValueError):
            return False

    malos = [p for p in prods if not _cod(p) or not _precio(p)]
    add("productos", f"Catálogo íntegro ({len(prods)} productos)", len(malos) == 0,
        f"{len(malos)} producto(s) sin código o sin precio de venta",
        severidad="media")

    n_sin_cod = sum(1 for c in await db.clients.find({}, {"codigo": 1}).to_list(20000)
                    if not (c.get("codigo") or "").strip())
    add("clientes_codigo", "Clientes con código asignado", n_sin_cod == 0,
        f"{n_sin_cod} cliente(s) sin código")

    # 6) Auditoría operando.
    n_audit = await db.audit_logs.count_documents({})
    add("auditoria", "Auditoría operando", n_audit > 0, f"{n_audit} registro(s)")

    listo = all(c["ok"] for c in checks if c["severidad"] != "baja")
    return {"checks": checks, "listo": listo}


# --- Limpieza de datos de prueba (SOLO DEV) -------------------------------------
@api.delete("/dev/datos-prueba")
async def dev_purgar_datos_prueba(user: dict = Depends(_dev_only)):
    """Elimina TODOS los datos generados para pruebas: cuentas demo (@rysa.dev)
    con sus ubicaciones GPS, clientes DEMO-* y las visitas asociadas."""
    resumen = {"usuarios_eliminados": 0, "ubicaciones_eliminadas": 0,
               "clientes_eliminados": 0, "visitas_eliminadas": 0}

    # 1) Cuentas demo: revocar sesiones y borrar su rastro.
    demos = await db.users.find({"email": {"$regex": "@rysa\\.dev$"}},
                                {"_id": 0}).to_list(100)
    demo_uids = [u["id"] for u in demos]
    for vid in demo_uids:
        try:
            await revoke_user_sessions(vid)
        except Exception:
            pass
        resumen["ubicaciones_eliminadas"] += await db.seller_locations.delete_many(
            {"vendedor_id": vid})
        resumen["visitas_eliminadas"] += await db.visits.delete_many(
            {"vendedor_id": vid})
        resumen["usuarios_eliminados"] += await db.users.delete_one({"id": vid})

    # 2) Clientes de prueba y sus visitas.
    cli_docs = await db.clients.find({"codigo": {"$regex": "^DEMO-"}},
                                     {"_id": 0}).to_list(1000)
    for cid in [c["id"] for c in cli_docs]:
        resumen["visitas_eliminadas"] += await db.visits.delete_many(
            {"cliente_id": cid})
    if cli_docs:
        resumen["clientes_eliminados"] = await db.clients.delete_many(
            {"codigo": {"$regex": "^DEMO-"}})

    await log_audit(user, "dev_purga_pruebas", "sistema", "",
                    f"Purga de datos de prueba: {resumen}")
    return {"ok": True, **resumen}


# =========================================================================
# SUCURSALES
# =========================================================================
@api.get("/sucursales")
async def list_sucursales(user: dict = Depends(get_current_user)):
    sucs = await db.sucursales.find({}, {"_id": 0}).to_list(500)
    sucs.sort(key=lambda s: s.get("codigo", ""))
    return sucs

@api.post("/sucursales")
async def create_sucursal(data: SucursalItem, user: dict = Depends(require_permission("config"))):
    nombre = (data.nombre or "").strip()
    if not nombre:
        raise HTTPException(400, "La sucursal requiere un nombre")
    codigo = (data.codigo or nombre).strip().upper()
    doc = {"id": uid(), "codigo": codigo, "nombre": nombre, "activa": True,
           "direccion": data.direccion, "ciudad": data.ciudad, "estado": data.estado,
           "cp": data.cp, "telefono": data.telefono, "created_at": iso_now()}
    await db.sucursales.insert_one(doc)
    await log_audit(user, "crear", "sucursal", doc["id"], nombre)
    return await db.sucursales.find_one({"id": doc["id"]}, {"_id": 0})

@api.put("/sucursales/{sucursal_id}")
async def update_sucursal(sucursal_id: str, data: SucursalItem,
                          user: dict = Depends(require_permission("config"))):
    doc = data.model_dump()
    await db.sucursales.update_one({"id": sucursal_id}, {"$set": doc})
    await log_audit(user, "editar", "sucursal", sucursal_id, data.nombre)
    return await db.sucursales.find_one({"id": sucursal_id}, {"_id": 0})

@api.delete("/sucursales/{sucursal_id}")
async def delete_sucursal(sucursal_id: str, user: dict = Depends(require_permission("config"))):
    await db.sucursales.delete_one({"id": sucursal_id})
    return {"ok": True}

# =========================================================================
# CONFIGURACIÓN / SETTINGS
# =========================================================================
def aplicar_storage_config(s):
    """Aplica el directorio de almacenamiento local elegido en Configuración."""
    try:
        st = (s or {}).get("storage") or {}
        if st.get("backend") in (None, "", "local"):
            if st.get("upload_dir"):
                storage.set_upload_dir(st["upload_dir"])
            else:
                # Si no hay sobrescritura elegida, vuelve al del entorno
                storage.set_upload_dir(None)
    except Exception:
        pass


def _storage_status():
    storage.init_storage()
    return {"backend": "local", "upload_dir": storage.base_upload_dir()}


@api.get("/settings")
async def get_settings(user: dict = Depends(get_current_user)):
    s = await db.settings.find_one({"_id": "app"}, {"_id": 0})
    if not s:
        s = {}
    st = s.get("storage") or {}
    st = {**st, "backend": st.get("backend") or "local", "upload_dir": st.get("upload_dir") or storage.base_upload_dir()}
    s = {**s, "storage": st}
    if not s.get("unidades_medida"):
        s["unidades_medida"] = list(UNIDADES_DEFAULT)  # UNIDADES_MERGE_DONE
    return s


@api.get("/settings/branding")
async def get_branding():
    """Branding público (logo + nombre) para Login/Landing, con fallback."""
    s = await db.settings.find_one({"_id": "app"}, {"_id": 0}) or {}
    return {
        "empresa_nombre": s.get("empresa_nombre", "Grupo RYSA"),
        "logo_url": s.get("logo_url") or "",
    }


@api.put("/settings")
async def update_settings(data: SettingsInput, user: dict = Depends(require_permission("config"))):
    doc = data.model_dump()
    await db.settings.update_one({"_id": "app"}, {"$set": doc}, upsert=True)
    aplicar_storage_config(doc)
    await log_audit(user, "editar", "configuracion", "app", "Actualización de configuración")
    return doc

# =========================================================================
# ARCHIVOS / OBJECT STORAGE (imágenes de productos/categorías, PDFs de ticket)
# =========================================================================

def _convertir_a_webp(data: bytes) -> tuple:
    """Convierte una imagen a WebP (85% de calidad) usando Pillow.

    Devuelve (bytes_webp, 'image/webp'). Si Pillow no está instalado o la
    conversión falla, devuelve los bytes originales sin cambiar el formato
    (el sistema sigue funcionando, solo sin optimización).
    """
    try:
        from PIL import Image
        from io import BytesIO
        img = Image.open(BytesIO(data))
        img = img.convert("RGB")
        out = BytesIO()
        img.save(out, format="WEBP", quality=85)
        out.seek(0)
        return out.read(), "image/webp"
    except Exception:
        return data, None


@api.post("/uploads/image")
async def upload_image(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """Sube una imagen de producto/categoría convirtiéndola a WebP."""
    data = await file.read()
    if len(data) > 8 * 1024 * 1024:
        raise HTTPException(400, "La imagen no debe superar 8 MB.")
    
    # Validar tipo MIME real por firma de bytes
    real_mime = storage.detect_mime_type(data)
    if real_mime not in ["image/jpeg", "image/png", "image/gif", "image/webp"]:
        raise HTTPException(400, "Formato real no permitido. Usa JPG, PNG, WEBP o GIF.")

    # Optimización: todas las imágenes se guardan en WEBP (formato ligero para
    # catálogos grandes). Si la conversión no aplica, se conserva el original.
    data_guardar = data
    ctype = real_mime
    convertido, new_mime = _convertir_a_webp(data)
    if new_mime:
        data_guardar = convertido
        ctype = new_mime
    
    # Para evitar vulnerabilidades de path traversal y sobrescritura, usamos un UUID aleatorio
    path = f"uploads/{uid()}.webp"
    try:
        result = storage.put_object(path, data_guardar, ctype)
    except Exception as e:
        logger.error("Upload imagen falló: %s", str(e)[:160])
        raise HTTPException(502, "No se pudo subir la imagen al almacenamiento local.")
    stored = result.get("path", path)
    await db.files.insert_one({
        "id": uid(), "storage_path": stored, "original_filename": file.filename,
        "content_type": ctype, "size": result.get("size", len(data_guardar)),
        "original_size": len(data), "original_type": real_mime,
        "is_deleted": False, "created_at": iso_now(),
    })
    return {"path": stored, "url": f"/api/files/{stored}"}

@api.post("/uploads/document")
async def upload_document(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """Sube un documento (factura, evidencia, formato) al almacenamiento local.
    Hasta 15 MB. Se registra en la colección de archivos para servirlo en /api/files."""
    data = await file.read()
    if len(data) > 15 * 1024 * 1024:
        raise HTTPException(400, "El documento no debe superar 15 MB.")
    mime = storage.detect_mime_type(data)
    if mime == "application/octet-stream":
        mime = mimetypes.guess_type(file.filename or "")[0] or "application/octet-stream"
    safe_ext = mimetypes.guess_extension(mime) or ".bin"
    path = f"documents/{uid()}{safe_ext}"
    try:
        result = storage.put_object(path, data, mime)
    except Exception as e:
        logger.error("Upload documento falló: %s", str(e)[:160])
        raise HTTPException(502, "No se pudo subir el documento al almacenamiento local.")
    stored = result.get("path", path)
    await db.files.insert_one({
        "id": uid(), "storage_path": stored, "original_filename": file.filename,
        "content_type": mime, "size": result.get("size", len(data)),
        "is_deleted": False, "created_at": iso_now(),
    })
    return {"path": stored, "url": f"/api/files/{stored}", "filename": file.filename}

@api.get("/files/{path:path}")
async def serve_file(path: str):
    # Defensa en profundidad: rechazo temprano de path traversal antes de tocar
    # el storage (storage.get_safe_local_path ya lo bloquea también).
    if not path or path.startswith("/") or "\\" in path or path.split("/")[0] in ("..", "."):
        raise HTTPException(404, "Archivo no encontrado")
    if ".." in path:
        raise HTTPException(404, "Archivo no encontrado")
    record = await db.files.find_one({"storage_path": path, "is_deleted": False})
    if not record:
        raise HTTPException(404, "Archivo no encontrado")
    try:
        data, ctype = storage.get_object(path)
    except Exception as e:
        logger.error("Error al obtener archivo local %s: %s", path, str(e)[:120])
        raise HTTPException(404, "Archivo no disponible")
    return Response(content=data, media_type=record.get("content_type", ctype))

@api.get("/sales/{sale_id}/qr")
async def sale_qr_png(sale_id: str, size: int = 240, destino: Optional[str] = None):
    """Genera un PNG del QR de verificación del ticket."""
    import qrcode
    from io import BytesIO
    from fastapi.responses import Response as FastResponse
    sale = await db.sales.find_one({"id": sale_id}, {"_id": 0})
    if not sale:
        raise HTTPException(404, "Ticket no encontrado")
    if destino:
        url = destino
    else:
        base_url = os.environ.get("PUBLIC_BASE_URL", "").rstrip("/")
        url = f"{base_url}/verificar/{sale_id}" if base_url else f"/api/sales/{sale_id}/public"
    qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=10, border=2)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    bio = BytesIO()
    img.save(bio, format="PNG")
    return FastResponse(content=bio.getvalue(), media_type="image/png",
                        headers={"Cache-Control": "no-store"})

@api.post("/sales/{sale_id}/ticket-pdf")
async def sale_ticket_pdf(sale_id: str, regenerar: bool = False,
                          user: dict = Depends(get_current_user)):
    """PDF del ticket. Generador ÚNICO (documentos.py): la primera llamada
    crea el archivo y las siguientes devuelven EXACTAMENTE el mismo PDF —
    el que se comparte por WhatsApp/descarga es siempre el original."""
    import documentos as _docs
    try:
        res = await _docs.asegurar_documentos(sale_id, formatos=("ticket",),
                                              regenerar=regenerar)
    except ValueError:
        raise HTTPException(404, "Venta no encontrada")
    except Exception as e:
        logger.error("Ticket PDF falló: %s", str(e)[:160])
        raise HTTPException(502, "No se pudo generar el PDF del ticket.")
    r = res.get("ticket") or {}
    if not r.get("url"):
        raise HTTPException(502, "No se pudo generar el PDF del ticket.")
    return {"path": r["path"], "url": r["url"]}


@api.post("/sales/{sale_id}/letter-pdf")
async def sale_letter_pdf(sale_id: str, regenerar: bool = False,
                          user: dict = Depends(get_current_user)):
    """Comprobante formato carta (Letter 8.5x11) con logotipo oficial.
    Mismo generador central: una sola vez por venta, mismo archivo siempre."""
    import documentos as _docs
    try:
        res = await _docs.asegurar_documentos(sale_id, formatos=("carta",),
                                              regenerar=regenerar)
    except ValueError:
        raise HTTPException(404, "Venta no encontrada")
    except Exception as e:
        logger.error("Formato carta PDF falló: %s", str(e)[:200])
        raise HTTPException(502, "No se pudo generar el comprobante en formato carta.")
    r = res.get("carta") or {}
    if not r.get("url"):
        raise HTTPException(502, "No se pudo generar el comprobante en formato carta.")
    return {"path": r["path"], "url": r["url"]}


@api.post("/sales/{sale_id}/cotizacion-pdf")
async def sale_cotizacion_pdf(sale_id: str, regenerar: bool = False,
                              request: Request = None,
                              user: dict = Depends(get_current_user)):
    """COTIZACIÓN en PDF tamaño carta (2 hojas: cotización + cuentas bancarias).
    Es el documento oficial de este tipo; NO usa ticket térmico. Mismo
    generador central: un solo archivo por cotización."""
    import documentos as _docs
    base_url = str(request.base_url).rstrip("/") if request else ""
    try:
        res = await _docs.asegurar_documentos(sale_id, formatos=("cotizacion",),
                                              regenerar=regenerar, base_url=base_url)
    except ValueError:
        raise HTTPException(404, "Venta no encontrada")
    except Exception as e:
        logger.error("PDF de cotización falló: %s", str(e)[:200])
        raise HTTPException(502, "No se pudo generar el PDF de la cotización.")
    r = res.get("cotizacion") or {}
    if not r.get("url"):
        raise HTTPException(502, "No se pudo generar el PDF de la cotización.")
    return {"path": r["path"], "url": r["url"]}


@api.get("/sales/{sale_id}/ticket-print")
async def sale_ticket_print(sale_id: str, request: Request):
    """Vista HTML de auto-impresión en PAPEL POS80 (@page 80mm) del ticket
    oficial YA generado. El iframe del POS la carga y el navegador imprime
    el documento real a 80mm por la impresora térmica predeterminada."""
    from fastapi.responses import HTMLResponse
    import documentos as _docs
    try:
        res = await _docs.asegurar_documentos(sale_id, formatos=("ticket",))
    except ValueError:
        raise HTTPException(404, "Venta no encontrada")
    except Exception as e:
        logger.error("ticket-print falló: %s", str(e)[:160])
        raise HTTPException(502, "No se pudo preparar la impresión.")
    r = res.get("ticket") or {}
    if not r.get("url"):
        raise HTTPException(502, "No se pudo preparar la impresión.")
    base = str(request.base_url).rstrip("/")
    src = f"{base}{r['url']}"
    folio = ""
    sale_doc = await db.sales.find_one({"id": sale_id}, {"_id": 0, "folio": 1})
    if sale_doc:
        folio = str(sale_doc.get("folio") or "")
    html = f"""<!doctype html><html><head><meta charset="utf-8"><title>Ticket {folio}</title>
<style>
@page {{ size: 80mm auto; margin: 0; }}
html, body {{ margin: 0; padding: 0; background: #fff; }}
embed {{ width: 80mm; display: block; }}
</style></head>
<body><embed id="pdf" src="{src}" type="application/pdf" width="80mm">
<script>
window.addEventListener('load', function () {{
  setTimeout(function () {{ try {{ window.print(); }} catch (e) {{}} }}, 400);
}});
</script>
</body></html>"""
    return HTMLResponse(html)


@api.get("/vendedores")
async def vendedores(user: dict = Depends(get_current_user)):
    return await db.users.find({"active": {"$ne": False}}, {"_id": 0, "id": 1, "name": 1, "role": 1}).to_list(200)

@api.get("/sales/{sale_id}/public")
async def sale_public_verify(sale_id: str):
    """Verificación pública del ticket (para el QR). No requiere autenticación."""
    sale = await db.sales.find_one({"id": sale_id}, {"_id": 0})
    if not sale or sale.get("estado") != "confirmada":
        raise HTTPException(404, "Ticket no encontrado")
    settings = await db.settings.find_one({"_id": "app"}, {"_id": 0}) or {}
    return {
        "ok": True,
        "folio": sale.get("folio"),
        "fecha": sale.get("fecha"),
        "empresa": settings.get("empresa_nombre", "Grupo RYSA"),
        "rfc": settings.get("rfc", ""),
        "cliente": sale.get("cliente_nombre", "Público General"),
        "total": sale.get("total"),
        "moneda": settings.get("moneda", "MXN"),
        "condicion": sale.get("condicion"),
        "items": [{"descripcion": i.get("descripcion"), "cantidad": i.get("cantidad"),
                   "precio": i.get("precio"), "importe": i.get("importe", i.get("cantidad", 0) * i.get("precio", 0))}
                  for i in sale.get("items", [])],
    }

@api.get("/sales-next-folio")
async def sales_next_folio(user: dict = Depends(get_current_user)):
    v = await db.counters.find_one({"_id": "venta"})
    c = await db.counters.find_one({"_id": "cotizacion"})
    vn = (v["seq"] if v else 0) + 1
    cn = (c["seq"] if c else 0) + 1
    return {"venta": f"V{str(vn).zfill(6)}", "cotizacion": f"COT{str(cn).zfill(6)}"}

# =========================================================================
# FACTURACIÓN CFDI 4.0 (PAC: Facty.mx)
# =========================================================================
class PacConfigInput(BaseModel):
    provider: str = "facty"
    environment: str = "sandbox"        # sandbox | produccion
    api_key: Optional[str] = ""          # API Key / Token de Facty.mx (Bearer)
    api_user: Optional[str] = ""         # reservado por compatibilidad (no usado por Facty.mx)
    api_password: Optional[str] = ""     # reservado por compatibilidad (no usado por Facty.mx)
    rfc: Optional[str] = ""
    razon_social: Optional[str] = ""
    regimen_fiscal: Optional[str] = "601"
    serie: Optional[str] = "A"
    folio: Optional[int] = 1
    lugar_expedicion: Optional[str] = ""  # CP
    timbres_alerta: Optional[int] = 20    # avisar cuando queden menos

class ComplementoPagoInput(BaseModel):
    factura_id: str
    monto: float
    metodo: str = "efectivo"      # efectivo | tarjeta | transferencia | spei | deposito | otros
    fecha: Optional[str] = None   # fecha del pago (YYYY-MM-DD)

# -------------------------------------------------------------------------
# Toda la comunicación con el PAC (Facty.mx) vive en pac_provider.py.
# La lógica de negocio del resto de la app (órdenes, POS, cobros, CxC) NO cambia:
# estos endpoints conservan su contrato. Si se cambia de PAC solo se toca pac_provider.py.
# -------------------------------------------------------------------------

async def get_pac_config():
    return await pac_provider.get_config()

def pac_configurado(cfg):
    return bool(cfg and pac_provider.api_key(cfg) and cfg.get("rfc"))

# -------------------------------------------------------------------------
# REFERENCIA (INACTIVA): implementación previa que construía el payload en el
# formato de Facturama (POST /3/cfdis). Se conserva comentada por si se necesita.
# La implementación activa está en pac_provider._payload_factura().
# -------------------------------------------------------------------------
# def facty_request(cfg, method, path, **kwargs):
#     base = FACTY_URLS.get(cfg.get("environment", "sandbox"), FACTY_URLS["sandbox"])
#     headers = {"Authorization": f"Bearer {cfg.get('api_key', '')}", **kwargs.pop("headers", {})}
#     with httpx.Client(base_url=base, headers=headers, timeout=30.0) as c:
#         r = c.request(method, path, **kwargs)
#         if r.status_code >= 400:
#             raise HTTPException(r.status_code, f"PAC (Facty.mx): {r.text[:500]}")
#         return r.json()
#
# def _money(x):
#     return f"{round(float(x) + 1e-9, 2):.2f}"
#
# def sale_to_cfdi_payload(sale, cliente, cfg):
#     forma_map = {"efectivo": "01", "tarjeta": "04", "transferencia": "03",
#                  "spei": "03", "deposito": "03", "otros": "99"}
#     pago = (sale.get("pagos") or [{}])
#     forma = forma_map.get((pago[0].get("metodo") if pago else "efectivo"), "01")
#     rfc = (cliente.get("rfc") if cliente else "") or "XAXX010101000"
#     generico = rfc == "XAXX010101000"
#     receiver = {
#         "Rfc": rfc.upper(),
#         "Name": ((cliente.get("nombre") if cliente else "") or "PUBLICO EN GENERAL").upper(),
#         "CfdiUse": (cliente.get("uso_cfdi") if cliente else "") or ("S01" if generico else "G03"),
#         "FiscalRegime": (cliente.get("reg_fiscal") if cliente else "") or ("616" if generico else "601"),
#         "TaxZipCode": (cliente.get("cp") if cliente else "") or cfg.get("lugar_expedicion") or "00000",
#     }
#     items = []
#     for it in sale.get("items", []):
#         tasa = float(it.get("iva_tasa", 16)) / 100
#         bruto_unit = float(it["precio"])
#         base_unit = round(bruto_unit / (1 + tasa), 2)
#         base = round(base_unit * it["cantidad"], 2)
#         tax = round(base * tasa, 2)
#         items.append({
#             "Quantity": _money(it["cantidad"]), "ProductCode": it.get("clave_sat") or "01010101",
#             "UnitCode": it.get("clave_unidad") or "H87", "Unit": it.get("unidad") or "Pieza",
#             "Description": it["descripcion"], "IdentificationNumber": it.get("codigo", ""),
#             "UnitPrice": _money(base_unit), "Subtotal": _money(base), "TaxObject": "02",
#             "Taxes": [{"Name": "IVA", "Rate": f"{tasa}", "Total": _money(tax), "Base": _money(base),
#                        "IsRetention": False, "IsFederalTax": True}],
#             "Total": _money(base + tax),
#         })
#     return {
#         "CfdiType": "I", "NameId": "1", "ExpeditionPlace": cfg.get("lugar_expedicion") or "00000",
#         "Serie": cfg.get("serie") or "A", "PaymentForm": forma, "PaymentMethod": "PUE",
#         "Exportation": "01", "Currency": "MXN", "Receiver": receiver, "Items": items,
#     }

@api.get("/facturacion/config")
async def get_pac_config_ep(user: dict = Depends(require_permission("config"))):
    cfg = await get_pac_config() or {}
    cfg = dict(cfg)
    cfg["api_key_set"] = bool(cfg.get("api_key"))
    cfg["api_password_set"] = bool(cfg.get("api_password"))
    cfg.pop("api_password", None)
    cfg.pop("api_key", None)
    cfg["configurado"] = pac_configurado(await get_pac_config())
    return cfg

@api.put("/facturacion/config")
async def put_pac_config(data: PacConfigInput, user: dict = Depends(require_permission("config"))):
    doc = data.model_dump()
    existing = await get_pac_config() or {}
    if not doc.get("api_key"):  # conservar API Key si no se reenvía
        doc["api_key"] = existing.get("api_key", "")
    if not doc.get("api_password"):  # conservar contraseña si no se reenvía
        doc["api_password"] = existing.get("api_password", "")
    await db.pac_config.update_one({"_id": "pac"}, {"$set": doc}, upsert=True)
    await log_audit(user, "editar", "facturacion", "config", f"PAC {doc.get('provider')} ({doc.get('environment')})")
    return {"ok": True, "configurado": pac_configurado(doc)}

@api.get("/facturacion/timbres")
async def timbres(user: dict = Depends(get_current_user)):
    cfg = await get_pac_config()
    if not pac_configurado(cfg):
        return {"configurado": False, "disponibles": None, "plan": None}
    try:
        data = await pac_provider.listar_timbres(cfg)
        disp = int(data.get("disponibles") or 0)
        res = {"configurado": True, "disponibles": disp, "plan": data.get("plan"),
               "expira": None, "actualizado": iso_now(),
               "alerta": disp <= int(cfg.get("timbres_alerta", 20))}
        await db.pac_config.update_one({"_id": "pac"}, {"$set": {"timbres_cache": res}})
        return res
    except HTTPException as e:
        cached = (cfg.get("timbres_cache") or {})
        cached["configurado"] = True
        cached["error"] = str(e.detail)[:200]
        return cached


@api.get("/facturacion")
async def list_cfdi(user: dict = Depends(get_current_user)):
    docs = await db.cfdi_documents.find({}, {"_id": 0, "response": 0}).sort("fecha", -1).to_list(2000)
    return docs

@api.get("/facturacion/facturables")
async def ventas_facturables(user: dict = Depends(get_current_user)):
    sales = await db.sales.find({"tipo_venta": {"$ne": "cotizacion"}, "estado": "confirmada",
                                 "facturado": {"$ne": True}}, {"_id": 0}).sort("fecha", -1).to_list(500)
    return sales

class FacturacionInput(BaseModel):
    rfc: Optional[str] = ""
    razon_social: Optional[str] = ""
    cp: Optional[str] = ""
    reg_fiscal: Optional[str] = ""
    uso_cfdi: Optional[str] = ""
    direccion: Optional[str] = ""


@api.post("/facturacion/sale/{sale_id}")
async def emitir_cfdi(sale_id: str, data: Optional[FacturacionInput] = None, user: dict = Depends(require_permission("venta.crear"))):
    cfg = await get_pac_config()
    if not pac_configurado(cfg):
        raise HTTPException(400, "El PAC no está configurado. Ve a Configuración → Facturación y captura tu API Key de Facty.")
    sale = await db.sales.find_one({"id": sale_id}, {"_id": 0})
    if not sale:
        raise HTTPException(404, "Venta no encontrada")
    if sale.get("facturado"):
        raise HTTPException(400, "Esta venta ya fue facturada")
    cliente = await db.clients.find_one({"id": sale.get("cliente_id")}, {"_id": 0}) if sale.get("cliente_id") else None
    receptor = data.model_dump() if data else {}
    receptor = {k: (v or "") for k, v in receptor.items()}
    result = await pac_provider.crear_factura(sale, cliente, cfg, receptor)
    fid = result.get("id")
    uuid_ = result.get("uuid")
    doc = {"id": uid(), "sale_id": sale_id, "folio_venta": sale.get("folio"),
           "pac_id": fid, "uuid": uuid_, "serie": result.get("serie"),
           "folio": result.get("folio"),
           "status": "vigente", "total": result.get("total", sale.get("total")),
           "cliente_nombre": (cliente.get("nombre") if cliente else "PUBLICO EN GENERAL"),
           "rfc": (cliente.get("rfc") if cliente else "") or "XAXX010101000",
           "fecha": iso_now(), "provider": cfg.get("provider") or "facty",
           "response": result.get("raw") or result}
    await db.cfdi_documents.insert_one(doc)
    await db.sales.update_one({"id": sale_id}, {"$set": {"facturado": True, "cfdi_uuid": uuid_, "cfdi_id": fid}})
    await log_audit(user, "facturar", "venta", sale_id, f"CFDI {uuid_}")
    return {"ok": True, "pac_id": fid, "uuid": uuid_, "folio": doc["folio"]}

class MultiFacturaInput(BaseModel):
    sale_ids: List[str] = Field(...)

@api.post("/facturacion/multi")
async def emitir_cfdi_multi(data: MultiFacturaInput, user: dict = Depends(require_permission("venta.facturar"))):
    """Factura varias ventas en UNA sola factura (CFDI).

    Reglas validadas en backend (evitar doble timbrado):
    - cada venta debe existir, estar `confirmada` y NO facturada;
    - todas deben pertenecer al MISMO cliente (RFC único);
    - ninguna puede ser cotización.
    Los conceptos de todas las ventas se agregan en un solo CFDI.
    """
    if not data.sale_ids:
        raise HTTPException(400, "Selecciona al menos una venta")
    cfg = await get_pac_config()
    if not pac_configurado(cfg):
        raise HTTPException(400, "El PAC no está configurado. Ve a Configuración → Facturación y captura tu API Key de Facty.")
    sales = []
    cliente_id = None
    for sid in data.sale_ids:
        s = await db.sales.find_one({"id": sid}, {"_id": 0})
        if not s:
            raise HTTPException(404, f"Venta {sid} no encontrada")
        if s.get("tipo_venta") == "cotizacion":
            raise HTTPException(400, f"La venta {s.get('folio')} es una cotización, no se puede facturar")
        if s.get("estado") != "confirmada":
            raise HTTPException(400, f"La venta {s.get('folio')} no está confirmada")
        if s.get("facturado"):
            raise HTTPException(400, f"La venta {s.get('folio')} ya fue facturada")  # evita doble timbre
        cid = s.get("cliente_id")
        if cid and cliente_id and cid != cliente_id:
            raise HTTPException(400, "Todas las ventas deben pertenecer al mismo cliente para una sola factura")
        if cid:
            cliente_id = cid
        sales.append(s)
    cliente = await db.clients.find_one({"id": cliente_id}, {"_id": 0}) if cliente_id else None

    # Doc agregado: una sola factura con los conceptos de todas las ventas.
    items = []
    for s in sales:
        items.extend(s.get("items", []))
    total = round(sum(float(s.get("total", 0)) for s in sales), 2)
    subtotal = round(sum(float(s.get("subtotal", 0)) for s in sales), 2)
    iva = round(sum(float(s.get("iva_total", 0)) for s in sales), 2)
    ag = {"id": uid(), "folio": "MULTI", "items": items, "total": total,
          "subtotal": subtotal, "iva_total": iva, "descuento_total": round(sum(float(s.get("descuento_total", 0)) for s in sales), 2),
          "condicion": "contado", "cliente_id": cliente_id, "cliente_nombre": cliente.get("nombre") if cliente else "PUBLICO EN GENERAL",
          "facturado": False}
    result = await pac_provider.crear_factura(ag, cliente, cfg)
    fid = result.get("id"); uuid_ = result.get("uuid")
    doc = {"id": uid(), "sale_ids": data.sale_ids, "sale_id": sales[0]["id"],
           "folio_venta": f"{sales[0].get('folio')}+{len(sales)}",
           "pac_id": fid, "uuid": uuid_, "serie": result.get("serie"), "folio": result.get("folio"),
           "status": "vigente", "total": result.get("total", total),
           "cliente_nombre": (cliente.get("nombre") if cliente else "PUBLICO EN GENERAL"),
           "rfc": (cliente.get("rfc") if cliente else "") or "XAXX010101000",
           "fecha": iso_now(), "provider": cfg.get("provider") or "facty",
           "response": result.get("raw") or result}
    await db.cfdi_documents.insert_one(doc)
    for s in sales:
        await db.sales.update_one({"id": s["id"]}, {"$set": {"facturado": True, "cfdi_uuid": uuid_, "cfdi_id": fid}})
    await log_audit(user, "facturar", "venta", ",".join(s["id"] for s in sales), f"CFDI multi {uuid_} ({len(sales)} ventas)")
    return {"ok": True, "pac_id": fid, "uuid": uuid_, "folio": doc["folio"], "ventas": len(sales)}


@api.get("/facturacion/{cfdi_id}/{fmt}")
async def descargar_cfdi(cfdi_id: str, fmt: str, user: dict = Depends(get_current_user)):
    if fmt not in ("xml", "pdf"):
        raise HTTPException(400, "Formato inválido")
    cfg = await get_pac_config()
    doc = await db.cfdi_documents.find_one({"id": cfdi_id}, {"_id": 0})
    if not doc or not pac_configurado(cfg):
        raise HTTPException(404, "CFDI o PAC no disponible")
    if fmt == "xml":
        raw, media = await pac_provider.descargar_xml(doc, cfg)
    else:
        raw, media = await pac_provider.descargar_pdf(doc, cfg)
    return StreamingResponse(io.BytesIO(raw), media_type=media,
        headers={"Content-Disposition": f"attachment; filename={doc.get('folio_venta','cfdi')}.{fmt}"})

@api.post("/facturacion/{cfdi_id}/cancel")
async def cancelar_cfdi(cfdi_id: str, motivo: str = "02", uuid_reemplazo: Optional[str] = None,
                        user: dict = Depends(require_permission("venta.cancelar"))):
    cfg = await get_pac_config()
    doc = await db.cfdi_documents.find_one({"id": cfdi_id})
    if not doc or not pac_configurado(cfg):
        raise HTTPException(404, "CFDI o PAC no disponible")
    if motivo not in ("01", "02", "03", "04"):
        raise HTTPException(422, "Motivo de cancelación inválido")
    if motivo == "01" and not uuid_reemplazo:
        raise HTTPException(422, "El motivo 01 requiere UUID de reemplazo")
    result = await pac_provider.cancelar_factura(doc, motivo, uuid_reemplazo, cfg)
    await db.cfdi_documents.update_one({"id": cfdi_id}, {"$set": {"status": "cancelado", "cancelacion": result}})
    if doc.get("sale_id"):
        await db.sales.update_one({"id": doc["sale_id"]}, {"$set": {"facturado": False}})
    await log_audit(user, "cancelar", "facturacion", cfdi_id, f"motivo {motivo}")
    return {"ok": True, "result": result}


@api.post("/facturacion/complemento-pago")
async def emitir_rep(data: ComplementoPagoInput, user: dict = Depends(require_permission("venta.cancelar"))):
    """Emite el Complemento de Recepción de Pagos (REP) contra una factura PPD.
    Uso manual (preparado, no conectado automáticamente a CxC)."""
    cfg = await get_pac_config()
    if not pac_configurado(cfg):
        raise HTTPException(400, "El PAC no está configurado. Configura tu API Key de Facty.")
    padre = await db.cfdi_documents.find_one({"id": data.factura_id}, {"_id": 0})
    if not padre:
        raise HTTPException(404, "Factura padre no encontrada")
    result = await pac_provider.emitir_complemento_pago(cfg, padre, data.model_dump())
    await log_audit(user, "facturar", "facturacion", padre.get("id"), f"Complemento de pago sobre {padre.get('uuid')}")
    return {"ok": True, "result": result}


# =========================================================================
# REPORTES DE VENTAS Y UTILIDAD
# =========================================================================
def _hi(x):
    return str(x or "").strip().lower()

async def _build_reporte(desde, hasta, group, vendedor_id=None, q=None,
                         categoria=None, tipo=None, user=None, query_extra=None) -> dict:
    """Construye el reporte de ventas/utilidad con filtros opcionales.
    Filtros: vendedor_id, q (producto/código), categoria (clasificación), tipo (contado|credito).
    user: si no tiene reportes.global, se limita a sus propias ventas."""
    from deps import ver_reportes_globales
    now = now_utc()
    d = (desde[:10] if desde else now.strftime("%Y-%m-01"))
    h = (hasta[:10] if hasta else now.date().isoformat())
    query = {"estado": "confirmada"}
    if vendedor_id:
        query["vendedor_id"] = vendedor_id
    if tipo in ("contado", "credito"):
        query["condicion"] = tipo
    if user is not None and not ver_reportes_globales(user):
        # Usuario sin alcance global: solo reporta sus propias ventas.
        query["vendedor_id"] = user["id"]
    if query_extra:
        query.update(query_extra)
    sales = await db.sales.find(query, {"_id": 0}).to_list(50000)
    sales = [s for s in sales if d <= s.get("fecha", "")[:10] <= h]

    # mapa de productos: id -> {costo, clasificacion, linea}
    prods = await db.products.find({}, {"_id": 0, "id": 1, "costo": 1, "clasificacion": 1, "linea": 1}).to_list(50000)
    pmeta = {}
    for p in prods:
        pmeta[p["id"]] = {
            "costo": float(p.get("costo") or 0),
            "clasificacion": _hi(p.get("clasificacion")),
            "linea": _hi(p.get("linea")),
        }

    def linea_cumple(it):
        if not q and not categoria:
            return True
        meta = pmeta.get(it.get("product_id")) or {}
        if q:
            ql = _hi(q)
            hay = (ql in _hi(it.get("codigo")) or ql in _hi(it.get("descripcion")))
            if not hay:
                return False
        if categoria:
            cl = _hi(categoria)
            if cl not in (meta.get("clasificacion") or "") and cl not in (meta.get("linea") or ""):
                return False
        return True

    por_producto = {}
    por_vendedor = {}
    por_categoria = {}
    serie = {}
    tickets_contados = set()
    total_ingreso = total_costo = total_ventas = 0.0
    unidades = 0.0
    for s in sales:
        matching_lines = [it for it in s.get("items", []) if linea_cumple(it)]
        if not matching_lines:
            continue
        sale_total = 0.0
        for it in matching_lines:
            # Reportes históricos: usan el snapshot congelado en la venta
            # (importe_neto/bruto e IVA por línea). Las ventas antiguas (sin
            # snapshot) se reconstruyen tratando `precio` como bruto (legacy).
            tasa = 1 + float(it.get("iva_tasa", 8)) / 100
            if it.get("importe_neto") is not None:
                neto = float(it["importe_neto"])
                if it.get("importe_bruto") is not None:
                    bruto = float(it["importe_bruto"])
                else:
                    bruto = round(neto * tasa, 2)
                iva_line = float(it.get("iva_linea") or 0) if it.get("iva_linea") is not None else round(bruto - neto, 2)
            else:
                bruto = it["cantidad"] * it["precio"] - (it.get("descuento", 0) or 0)
                neto = max(0.0, bruto) / tasa if tasa else max(0.0, bruto)
                iva_line = bruto - neto
            costo_snap = it.get("costo")
            if costo_snap is None:
                costo_snap = pmeta.get(it.get("product_id"), {}).get("costo", 0.0)
            costo = float(costo_snap) * it["cantidad"]
            pid = it.get("product_id") or it.get("codigo")
            p = por_producto.get(pid)
            if not p:
                meta = pmeta.get(it.get("product_id")) or {}
                p = {"codigo": it.get("codigo"), "descripcion": it.get("descripcion"),
                     "cantidad": 0, "ingreso": 0.0, "ingreso_bruto": 0.0, "iva": 0.0,
                     "costo": 0.0,
                     "clasificacion": it.get("clasificacion") or ""}
                por_producto[pid] = p
            p["cantidad"] += it["cantidad"]
            p["ingreso"] += neto
            p["ingreso_bruto"] += bruto
            p["iva"] += iva_line
            p["costo"] += costo
            total_ingreso += neto
            total_costo += costo
            unidades += it["cantidad"]
            sale_total += neto
            # categoria
            meta = pmeta.get(it.get("product_id")) or {}
            cat = meta.get("clasificacion") or meta.get("linea") or "Sin categoría"
            c = por_categoria.get(cat)
            if not c:
                c = {"categoria": cat, "cantidad": 0, "ingreso": 0.0, "costo": 0.0}
                por_categoria[cat] = c
            c["cantidad"] += it["cantidad"]
            c["ingreso"] += neto
            c["costo"] += costo
        # vendedor
        vid = s.get("vendedor_id") or "?"
        v = por_vendedor.get(vid)
        if not v:
            v = {"id": vid, "nombre": s.get("vendedor_nombre") or "Sin vendedor",
                 "tickets": 0, "ventas": 0.0, "utilidad": 0.0}
            por_vendedor[vid] = v
        if s["id"] not in tickets_contados:
            tickets_contados.add(s["id"])
            v["tickets"] += 1
        v["ventas"] += float(s.get("total", 0))
        v["utilidad"] += sale_total
        key = s.get("fecha", "")[:7] if group == "mes" else s.get("fecha", "")[:10]
        serie[key] = serie.get(key, 0) + float(s.get("total", 0))
        total_ventas += float(s.get("total", 0))

    productos = []
    for p in por_producto.values():
        util = p["ingreso"] - p["costo"]
        cant = p["cantidad"] or 0
        productos.append({**p, "ingreso": round(p["ingreso"], 2),
                          "ingreso_bruto": round(p["ingreso_bruto"], 2),
                          "iva": round(p["iva"], 2),
                          "precio_neto": round(p["ingreso"] / cant, 2) if cant else 0,
                          "precio_bruto": round(p["ingreso_bruto"] / cant, 2) if cant else 0,
                          "costo": round(p["costo"], 2),
                          "utilidad": round(util, 2),
                          "margen": round(util / p["ingreso"] * 100, 2) if p["ingreso"] else 0})
    top_vendidos = sorted(productos, key=lambda x: x["cantidad"], reverse=True)[:15]
    top_utilidad = sorted(productos, key=lambda x: x["utilidad"], reverse=True)[:15]
    series = [{"periodo": k, "total": round(v, 2)} for k, v in sorted(serie.items())]
    util_total = round(total_ingreso - total_costo, 2)
    vendedores = []
    for v in por_vendedor.values():
        v["ventas"] = round(v["ventas"], 2)
        v["utilidad"] = round(v["utilidad"], 2)
        v["ticket_promedio"] = round(v["ventas"] / v["tickets"], 2) if v["tickets"] else 0
        vendedores.append(v)
    vendedores.sort(key=lambda x: x["ventas"], reverse=True)
    ql = _hi(q)
    if ql:
        productos = [p for p in productos if ql in _hi(p["codigo"]) or ql in _hi(p["descripcion"])]
    categorias = []
    for c in por_categoria.values():
        util = c["ingreso"] - c["costo"]
        categorias.append({**c, "ingreso": round(c["ingreso"], 2), "costo": round(c["costo"], 2),
                           "utilidad": round(util, 2),
                           "margen": round(util / c["ingreso"] * 100, 2) if c["ingreso"] else 0})
    categorias.sort(key=lambda x: x["utilidad"], reverse=True)
    return {
        "desde": d, "hasta": h, "group": group, "filtros": {"vendedor": vendedor_id, "q": q, "categoria": categoria, "tipo": tipo},
        "totales": {"ventas": round(total_ventas, 2), "ingreso_neto": round(total_ingreso, 2),
                    "costo": round(total_costo, 2), "utilidad": util_total,
                    "margen": round(util_total / total_ingreso * 100, 2) if total_ingreso else 0,
                    "tickets": len(tickets_contados), "unidades": round(unidades, 2)},
        "series": series, "top_vendidos": top_vendidos, "top_utilidad": top_utilidad,
        "productos": sorted(productos, key=lambda x: x["utilidad"], reverse=True),
        "vendedores": vendedores,
        "categorias": categorias,
    }

@api.get("/reports/ventas")
async def reporte_ventas(desde: Optional[str] = None, hasta: Optional[str] = None,
                         group: str = "dia", vendedor_id: Optional[str] = None,
                         q: Optional[str] = None, categoria: Optional[str] = None,
                         tipo: Optional[str] = None, cliente_id: Optional[str] = None,
                         sucursal_id: Optional[str] = None, condicion: Optional[str] = None,
                         user: dict = Depends(require_permission("reportes.ver"))):
    query_extra = {}
    if cliente_id:
        query_extra["cliente_id"] = cliente_id
    if sucursal_id:
        query_extra["sucursal_id"] = sucursal_id
    if condicion in ("contado", "credito"):
        query_extra["condicion"] = condicion
    return await _build_reporte(desde, hasta, group, vendedor_id, q, categoria, tipo, user, query_extra=query_extra)

# --- Exportación de reportes: Excel y PDF ---
def _reporte_excel_bytes(rep: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    hf = Font(bold=True, color="FFFFFF"); fill = PatternFill("solid", fgColor="B95A3A")
    def estilar(ws):
        for c in ws[1]:
            c.font = hf; c.fill = fill
        for col in ws.columns:
            width = max(len(str(c.value)) for c in col if c.value is not None)
            ws.column_dimensions[col[0].column_letter].width = min(width + 2, 30)
    ws = wb.active; ws.title = "Resumen"
    t = rep["totales"]
    ws.append(["Métrica", "Valor"])
    for k in ["ventas", "ingreso_neto", "costo", "utilidad", "margen", "tickets", "unidades"]:
        ws.append([k, t.get(k)])
    estilar(ws)
    ws_c = wb.create_sheet("Categorías")
    ws_c.append(["Categoría", "Cantidad", "Ingreso", "Costo", "Utilidad", "Margen %"])
    for c in rep.get("categorias", []):
        ws_c.append([c["categoria"], c["cantidad"], c["ingreso"], c["costo"], c["utilidad"], c["margen"]])
    estilar(ws_c)
    ws2 = wb.create_sheet("Productos")
    ws2.append(["Código", "Producto", "Cantidad", "Ingreso neto", "IVA", "Ingreso bruto",
                "Precio neto prom.", "Precio bruto prom.", "Costo", "Utilidad", "Margen %"])
    for p in rep["productos"]:
        ws2.append([p["codigo"], p["descripcion"], p["cantidad"], p["ingreso"], p["iva"],
                    p["ingreso_bruto"], p["precio_neto"], p["precio_bruto"], p["costo"], p["utilidad"], p["margen"]])
    estilar(ws2)
    ws3 = wb.create_sheet("Vendedores")
    ws3.append(["Vendedor", "Tickets", "Ventas", "Ticket promedio", "Utilidad"])
    for v in rep.get("vendedores", []):
        ws3.append([v["nombre"], v["tickets"], v["ventas"], v["ticket_promedio"], v["utilidad"]])
    estilar(ws3)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

def _reporte_pdf_bytes(rep: dict) -> bytes:
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, mm
    from reportlab.lib.enums import TA_RIGHT, TA_LEFT, TA_CENTER
    buf = io.BytesIO()
    margins = 18 * mm
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter), leftMargin=margins,
                            rightMargin=margins, topMargin=margins, bottomMargin=margins,
                            title="Reporte de Ventas y Utilidad - Grupo RYSA",
                            author="Grupo RYSA")
    st = getSampleStyleSheet()
    st["Title"].fontName = "Helvetica-Bold"; st["Title"].fontSize = 16; st["Title"].textColor = colors.HexColor("#8B3A2A")
    st["Title"].alignment = TA_LEFT; st["Title"].spaceAfter = 2
    subt = ParagraphStyle("subt", parent=st["Normal"], fontSize=9, textColor=colors.HexColor("#666666"), spaceAfter=3)
    tiny = ParagraphStyle("tiny", parent=st["Normal"], fontSize=8, textColor=colors.HexColor("#888888"))
    h2 = ParagraphStyle("h2b", parent=st["Normal"], fontName="Helvetica-Bold", fontSize=11,
                        textColor=colors.HexColor("#8B3A2A"), spaceBefore=6, spaceAfter=4)

    logo_path = None
    for cand in (os.path.join(os.path.dirname(__file__), "brand", "logotipo.png"),
                 os.getenv("UPLOAD_DIR", "")):
        if cand and os.path.isfile(cand):
            logo_path = cand
            break

    # --- Encabezado con logotipo ---
    enc_data = [[None, None]]
    if logo_path:
        try:
            enc_data = [[Image(logo_path, width=120, height=120 * (545 / 1157)), ""]]
        except Exception:
            enc_data = [[None, None]]
    enc = Table(enc_data, colWidths=[170, None])
    enc.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    elems = [enc]

    # Título y datos
    f = rep["filtros"] or {}
    titulo = Paragraph("Reporte de Ventas y Utilidad", st["Title"])
    meta = Paragraph(
        f"Periodo: {rep['desde']} a {rep['hasta']} &nbsp;&nbsp;|&nbsp;&nbsp; Generado: "
        f"{now_utc().strftime('%d/%m/%Y %H:%M')} &nbsp;&nbsp;|&nbsp;&nbsp; Agrupado por {rep['group']}",
        subt)
    filtros_txt = []
    if f.get("vendedor"):
        filtros_txt.append("vendedor seleccionado")
    if f.get("q"):
        filtros_txt.append(f"producto: {f['q']}")
    if f.get("categoria"):
        filtros_txt.append(f"categoría: {f['categoria']}")
    if f.get("tipo"):
        filtros_txt.append(f"tipo: {f['tipo']}")
    meta2 = Paragraph(("Filtros: " + ", ".join(filtros_txt)) if filtros_txt else "Sin filtros adicionales", tiny)
    head_block = Table([[Paragraph("<b>Grupo RYSA</b>", ParagraphStyle("emp", parent=st["Normal"], fontName="Helvetica-Bold", fontSize=13, textColor=colors.HexColor("#1f2937")))],
                        [titulo], [meta], [meta2]],
                       colWidths=[None])
    head_block.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    header = Table([[enc, head_block]], colWidths=[170, None])
    header.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("LEFTPADDING", (0, 0), (0, 0), 0),
        ("RIGHTPADDING", (1, 0), (1, 0), 4),
    ]))
    elems.append(header)
    # línea divisoria
    rule = Table([[""]], colWidths=[None])
    rule.setStyle(TableStyle([("LINEBELOW", (0, 0), (-1, -1), 1.2, colors.HexColor("#B95A3A")),
                              ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 1)]))
    elems.append(rule)

    # Resumen de totales
    t = rep["totales"]
    tot_style = ParagraphStyle("tot", parent=st["Normal"], fontName="Helvetica-Bold", fontSize=9, alignment=TA_CENTER)
    tot_rows = [["VENTAS", "INGRESO NETO", "COSTO", "UTILIDAD", "MARGEN", "TICKETS"],
                [f"$ {t['ventas']:,.2f}", f"$ {t['ingreso_neto']:,.2f}", f"$ {t['costo']:,.2f}",
                 f"$ {t['utilidad']:,.2f}", f"{t['margen']}%", str(t['tickets'])]]
    tot = Table(tot_rows, colWidths=[None] * 6)
    tot.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2b2b2b")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, 1), 10),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#f4ece6")),
        ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#2b2b2b")),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#c9c9c9")),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d9d9d9")),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elems.append(Spacer(1, 8))
    elems.append(tot)

    def seccion_tabla(title, data, colwidths, numero_cols):
        tab = Table(data, colWidths=colwidths, repeatRows=1)
        tab.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#B95A3A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("ALIGN", (0, 1), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 1), (0, -1), "LEFT"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e0d8d2")),
            ("FONTSIZE", (0, 1), (-1, -1), 7.5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#faf5f1")]),
            ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        return tab

    # Categorías
    elems.append(Paragraph("Reporte por categoría", h2))
    cat_data = [["Categoría", "Cantidad", "Ingreso", "Costo", "Utilidad", "Margen %"]]
    for c in rep["categorias"]:
        cat_data.append([c["categoria"], c["cantidad"], f"$ {c['ingreso']:,.2f}", f"$ {c['costo']:,.2f}",
                         f"$ {c['utilidad']:,.2f}", f"{c['margen']}%"])
    if len(cat_data) == 1:
        cat_data.append(["Sin ventas", "-", "-", "-", "-", "-"])
    elems.append(seccion_tabla("Categorías", cat_data, [None] * 6, 6))
    elems.append(Spacer(1, 6))

    # Productos
    elems.append(Paragraph("Utilidad por producto", h2))
    data = [["Código", "Producto", "Cant.", "Neto", "IVA", "Bruto", "Costo", "Utilidad", "Margen %"]]
    for p in rep["productos"][:120]:
        data.append([p["codigo"], p["descripcion"], p["cantidad"], f"$ {p['ingreso']:,.2f}",
                     f"$ {p['iva']:,.2f}", f"$ {p['ingreso_bruto']:,.2f}",
                     f"$ {p['costo']:,.2f}", f"$ {p['utilidad']:,.2f}", f"{p['margen']}%"])
    if len(data) == 1:
        data.append(["-", "Sin ventas", "-", "-", "-", "-", "-", "-", "-"])
    elems.append(seccion_tabla("Productos", data, [66, None, 50, 72, 66, 72, 66, 72, 55], 9))

    def _pie(canvas, docu):
        canvas.saveState()
        canvas.setFont("Helvetica", 7.5)
        canvas.setFillColor(colors.HexColor("#777777"))
        canvas.drawString(margins, 10 * mm, "Grupo RYSA - Reporte generado automáticamente")
        canvas.drawRightString(docu.pagesize[0] - margins, 10 * mm, f"Página {docu.page}")
        canvas.restoreState()
    doc.build(elems, onFirstPage=_pie, onLaterPages=_pie)
    buf.seek(0)
    return buf.read()

@api.get("/reports/ventas/export")
async def reporte_ventas_export(desde: Optional[str] = None, hasta: Optional[str] = None,
                                group: str = "dia", vendedor_id: Optional[str] = None,
                                q: Optional[str] = None, categoria: Optional[str] = None,
                                tipo: Optional[str] = None, fmt: str = "excel",
                                cliente_id: Optional[str] = None, sucursal_id: Optional[str] = None,
                                condicion: Optional[str] = None,
                                user: dict = Depends(require_permission("exportar"))):
    query_extra = {}
    if cliente_id:
        query_extra["cliente_id"] = cliente_id
    if sucursal_id:
        query_extra["sucursal_id"] = sucursal_id
    if condicion in ("contado", "credito"):
        query_extra["condicion"] = condicion
    rep = await _build_reporte(desde, hasta, group, vendedor_id, q, categoria, tipo, user, query_extra=query_extra)
    suffix = "xlsx" if fmt == "excel" else "pdf"
    media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if fmt == "excel" else "application/pdf"
    if fmt == "excel":
        raw = _reporte_excel_bytes(rep)
    else:
        raw = _reporte_pdf_bytes(rep)
    fname = f"reporte_ventas_{rep['desde']}_{rep['hasta']}.{suffix}"
    return StreamingResponse(io.BytesIO(raw), media_type=media,
        headers={"Content-Disposition": f"attachment; filename={fname}"})

# =========================================================================
# REPORTE DE INVENTARIO VALORIZADO (existencia, costo, margen, rotación)
# =========================================================================
async def _build_inventario(desde: str = "", hasta: str = "", q: str = "",
                            estado: str = "activo", user: dict = None) -> dict:
    """Valoriza el inventario actual: existencia, costo, venta potencial,
    margen y rotación (unidades vendidas en el periodo)."""
    query = {}
    if estado != "todos":
        query["estado"] = estado
    docs = await db.products.find(query, {"_id": 0}).to_list(50000)
    if q:
        ql = _hi(q)
        docs = [p for p in docs if ql in _hi(p.get("codigo")) or ql in _hi(p.get("descripcion"))
                or ql in _hi(p.get("linea")) or ql in _hi(p.get("clasificacion"))]

    # Unidades vendidas por producto en el periodo (para rotación).
    ventas_rango = {}
    d = (desde[:10] or now_utc().strftime("%Y-%m-01"))
    h = (hasta[:10] or now_utc().date().isoformat())
    srows = await db.sales.find({"estado": "confirmada"}, {"_id": 0}).to_list(100000)
    for s in srows:
        if not (d <= s.get("fecha", "")[:10] <= h):
            continue
        for it in s.get("items", []):
            pid = it.get("product_id")
            if not pid:
                continue
            ventas_rango[pid] = ventas_rango.get(pid, 0.0) + float(it.get("cantidad", 0) or 0)

    filas = []
    t_valor = t_potencial = t_costo = 0.0
    t_unidades_vendidas = 0.0
    for p in docs:
        exist = float(p.get("existencia", 0) or 0)
        costo = float(p.get("costo", 0) or 0)
        neto = float(p.get("precio_sin_iva") or 0) or float(p.get("precios", [{}])[0].get("precio_sin_iva", 0) or 0)
        con = float(p.get("precio_con_iva") or 0) or float(p.get("precios", [{}])[0].get("precio_con_iva", 0) or 0)
        valor = round(exist * costo, 2)
        potencial = round(exist * (con or neto), 2)
        util_pot = round(exist * (neto - costo), 2)
        margen = round((neto - costo) / neto * 100, 2) if neto else 0.0
        vendidas = round(ventas_rango.get(p.get("id"), 0.0), 2)
        rotacion = round(vendidas / exist * 100, 2) if exist else 0.0
        t_valor += valor; t_potencial += potencial; t_costo += costo * exist
        t_unidades_vendidas += vendidas
        filas.append({
            "codigo": p.get("codigo"), "descripcion": p.get("descripcion"),
            "linea": p.get("linea") or "", "clasificacion": p.get("clasificacion") or "",
            "unidad_medida": p.get("unidad_medida") or "PZA",
            "existencia": round(exist, 2), "stock_minimo": float(p.get("stock_minimo", 0) or 0),
            "costo": round(costo, 2), "precio_sin_iva": round(neto, 2),
            "precio_con_iva": round(con, 2), "utilidad": round(neto - costo, 2),
            "margen": margen, "valor_inventario": valor, "venta_potencial": potencial,
            "utilidad_potencial": util_pot, "unidades_vendidas": vendidas,
            "rotacion": rotacion, "estado": p.get("estado"),
        })
    filas.sort(key=lambda r: (r["valor_inventario"]), reverse=True)
    return {
        "desde": d, "hasta": h, "filtros": {"estado": estado, "q": q},
        "totales": {
            "productos": len(filas),
            "unidades": round(sum(r["existencia"] for r in filas), 2),
            "valor_inventario": round(t_valor, 2),
            "costo_total": round(t_costo, 2),
            "venta_potencial": round(t_potencial, 2),
            "utilidad_potencial": round(t_potencial - t_costo, 2),
            "margen_promedio": round((t_potencial - t_costo) / t_potencial * 100, 2) if t_potencial else 0,
            "unidades_vendidas": round(t_unidades_vendidas, 2),
        },
        "productos": filas,
    }


@api.get("/reports/inventario")
async def reporte_inventario(desde: Optional[str] = None, hasta: Optional[str] = None,
                             q: Optional[str] = None, estado: str = "activo",
                             user: dict = Depends(get_current_user)):
    return await _build_inventario(desde or "", hasta or "", q or "", estado, user)


@api.get("/reports/inventario/export")
async def reporte_inventario_export(desde: Optional[str] = None, hasta: Optional[str] = None,
                                    q: Optional[str] = None, estado: str = "activo", fmt: str = "excel",
                                    user: dict = Depends(require_permission("exportar"))):
    rep = await _build_inventario(desde or "", hasta or "", q or "", estado, user)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active; ws.title = "Inventario"
    hf = Font(bold=True, color="FFFFFF"); fill = PatternFill("solid", fgColor="B95A3A")
    ws.append(["Código", "Producto", "Línea", "Clasificación", "Existencia", "Costo unit.",
               "Precio neto", "Precio bruto", "Valor inventario", "Venta potencial",
               "Utilidad potencial", "Margen %", "Unid. vendidas", "Rotación %"])
    for col in ws[1]:
        col.font = hf; col.fill = fill
    for r in rep["productos"]:
        ws.append([r["codigo"], r["descripcion"], r["linea"], r["clasificacion"],
                   r["existencia"], r["costo"], r["precio_sin_iva"], r["precio_con_iva"],
                   r["valor_inventario"], r["venta_potencial"], r["utilidad_potencial"],
                   r["margen"], r["unidades_vendidas"], r["rotacion"]])
    for col in ws.columns:
        width = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 30)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"inventario_valorizado_{rep['desde']}_{rep['hasta']}.xlsx"
    return StreamingResponse(io.BytesIO(buf.read()), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


# =========================================================================
# CENTRO DE REPORTES (resumen ejecutivo + día típico)
# =========================================================================
@api.get("/reports/centro")
async def reporte_centro(desde: Optional[str] = None, hasta: Optional[str] = None,
                         user: dict = Depends(require_permission("reportes.ver"))):
    """Resumen ejecutivo del periodo: ticket promedio, hora pico, día de la
    semana con más ventas, métodos de pago y comparativos hoy/ayer/semana."""
    now = now_utc()
    hoy = now.date().isoformat()
    ayer = (now - timedelta(days=1)).date().isoformat()
    d = (desde[:10] if desde else now.strftime("%Y-%m-01"))
    h = (hasta[:10] if hasta else hoy)
    sales = await db.sales.find({"estado": "confirmada"}, {"_id": 0}).to_list(100000)
    sales = [s for s in sales if d <= s.get("fecha", "")[:10] <= h]

    dias = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    por_hora = {}
    por_dia = {dd: {"tickets": 0, "total": 0.0} for dd in dias}
    por_metodo = {}
    ingreso_neto = 0.0
    costo_total = 0.0
    for s in sales:
        hora = str(s.get("hora") or (s.get("fecha", "")[11:16]) or "00:00")[:2]
        por_hora[int(hora)] = por_hora.get(int(hora), 0) + 1
        try:
            idx = date.fromisoformat(s.get("fecha", "")[:10]).weekday()
            por_dia[dias[idx]]["tickets"] += 1
            por_dia[dias[idx]]["total"] += float(s.get("total", 0) or 0)
        except Exception:
            pass
        for p in s.get("pagos", []):
            m = p.get("metodo", "otros")
            if m == "tarjeta" and p.get("card_type") in ("debito", "credito"):
                m = f"tarjeta_{p['card_type']}"
            por_metodo[m] = por_metodo.get(m, 0) + float(p.get("monto", 0) or 0)
        for it in s.get("items", []):
            if it.get("importe_neto") is not None:
                ingreso_neto += float(it["importe_neto"])
            else:
                t = 1 + float(it.get("iva_tasa", 8)) / 100
                ingreso_neto += max(0.0, float(it["cantidad"]) * float(it["precio"]) - float(it.get("descuento", 0) or 0)) / t if t else 0
            costo_total += float(it.get("costo") or 0) * float(it.get("cantidad", 0) or 0)

    total_ventas = round(sum(float(s.get("total", 0) or 0) for s in sales), 2)
    num_ventas = len(sales)
    # Comparativos
    def rango_total(ff):
        return round(sum(float(s["total"]) for s in sales if s.get("fecha", "")[:10] == ff), 2)
    hoy_t = rango_total(hoy); ayer_t = rango_total(ayer)
    return {
        "desde": d, "hasta": h,
        "resumen": {
            "ventas": total_ventas, "tickets": num_ventas,
            "ticket_promedio": round(total_ventas / num_ventas, 2) if num_ventas else 0,
            "ingreso_neto": round(ingreso_neto, 2),
            "costo": round(costo_total, 2),
            "utilidad": round(ingreso_neto - costo_total, 2),
            "hoy": hoy_t, "ayer": ayer_t,
            "delta_hoy_ayer": round(hoy_t - ayer_t, 2),
        },
        "por_hora": [{"hora": f"{hh:02d}:00", "tickets": por_hora.get(hh, 0)} for hh in range(24)],
        "por_dia_semana": [{"dia": dd, "tickets": por_dia[dd]["tickets"], "total": round(por_dia[dd]["total"], 2)} for dd in dias],
        "por_metodo": [{"metodo": k, "monto": round(v, 2)} for k, v in sorted(por_metodo.items(), key=lambda x: -x[1])],
        "hora_pico": max(por_hora.items(), key=lambda x: x[1])[0] if por_hora else None,
        "dia_pico": max(dias, key=lambda dd: por_dia[dd]["tickets"]) if num_ventas else None,
    }

# =========================================================================
# STARTUP
# =========================================================================
@app.on_event("startup")
async def startup():
    try:
        storage.init_storage()
        logger.info("Almacenamiento local inicializado en: %s", storage.UPLOAD_DIR)
    except Exception as e:
        logger.error("Storage init falló: %s", str(e)[:160])
    try:
        st_cfg = await db.settings.find_one({"_id": "app"}, {"storage": 1})
        if st_cfg:
            aplicar_storage_config(st_cfg)
            logger.info("Storage configurado desde Configuración: %s", storage.base_upload_dir())
    except Exception as e:
        logger.warning("No se pudo aplicar storage desde settings: %s", str(e)[:120])

    # Backfill: asegura que todo producto tenga su código registrado también como
    # código de barras (requisito: el código se usa para llenar el código de barras).
    try:
        filled = 0
        async for p in db.products.find({"$or": [{"codigos_barras": {"$in": ["", None]}}, {"codigos_barras": {"$exists": False}}]}):
            doc = asegurar_codigo_como_barras({**p, "codigo": p.get("codigo", "")})
            if doc.get("codigos_barras") != p.get("codigos_barras"):
                await db.products.update_one({"id": p["id"]}, {"$set": {"codigos_barras": doc["codigos_barras"]}})
                filled += 1
        if filled:
            logger.info("Backfill de códigos de barras: %d productos actualizados", filled)
    except Exception as e:
        logger.warning("Backfill de códigos de barras falló: %s", str(e)[:120])
    await db.users.create_index("email", unique=True)
    await db.users.create_index("id", unique=True)
    await db.products.create_index("codigo")
    try:
        idx = await db.clients.index_information()
        if "codigo_1" in idx and not idx["codigo_1"].get("unique"):
            await db.clients.drop_index("codigo_1")
        await db.clients.create_index("codigo", unique=True)
    except Exception as e:
        logger.warning("Índice único clients.codigo: %s", str(e)[:120])

    # Sincronizar categorías desde clasificaciones de productos
    try:
        creadas = 0
        async for r in db.products.aggregate([
            {"$match": {"clasificacion": {"$nin": ["", None]}}},
            {"$group": {"_id": "$clasificacion"}},
        ]):
            await db.categories.update_one({"nombre": r["_id"]}, {"$set": {"nombre": r["_id"]}}, upsert=True)
            creadas += 1
        logger.info("Categorías sincronizadas desde clasificaciones: %d", creadas)
    except Exception as e:
        logger.warning("Sync de categorías falló: %s", str(e)[:120])

    # Seed admin SOLO en desarrollo y SOLO con credenciales de variables de
    # entorno (jamás hardcodeadas). Sin ADMIN_EMAIL/ADMIN_PASSWORD o con una
    # contraseña < 12 caracteres simplemente se omite el seed.
    env = os.environ.get("ENVIRONMENT", "development").lower()
    admin_email = os.environ.get("ADMIN_EMAIL", "").strip().lower()
    admin_pw = os.environ.get("ADMIN_PASSWORD", "")

    if env == "production":
        # En producción el seed del administrador NO se ejecuta automáticamente:
        # el primer usuario admin debe crearse manualmente.
        if not await db.users.find_one({"role": "admin"}):
            logger.warning(
                "Seed de admin desactivado en producción. Crea el usuario administrador "
                "manualmente antes de abrir el sistema."
            )
    else:
        if not admin_email or not admin_pw:
            logger.warning(
                "Seed de admin omitido: define ADMIN_EMAIL y ADMIN_PASSWORD "
                "(≥ 12 caracteres, nunca hardcodeados) para desarrollo."
            )
        elif not _password_ok(admin_pw):
            logger.warning("Seed de admin omitido: ADMIN_PASSWORD debe tener al menos 12 caracteres.")
        else:
            existing = await db.users.find_one({"email": admin_email})
            if not existing:
                await db.users.insert_one({
                    "id": uid(), "email": admin_email, "name": os.environ.get("ADMIN_NAME", "Admin"),
                    "role": "admin", "password_hash": hash_password(admin_pw),
                    "active": True, "token_version": 0, "created_at": iso_now()})
                logger.info("Admin seed creado: %s", admin_email)
            elif not verify_password(admin_pw, existing["password_hash"]):
                await db.users.update_one({"email": admin_email},
                                          {"$set": {"password_hash": hash_password(admin_pw)}})
                logger.info("Admin password actualizado (solo desarrollo).")
        
    # Cliente Público General por defecto
    if not await db.clients.find_one({"codigo": "PUBLICO"}):
        await db.clients.insert_one({
            "id": uid(), "codigo": "PUBLICO", "nombre": "Público General",
            "razon_social": "", "rfc": "XAXX010101000", "telefono": "", "whatsapp": "",
            "correo": "", "direccion": "", "ciudad": "", "estado_geo": "", "cp": "",
            "tipo": "publico", "lista_precios": 1, "condicion_pago": "contado",
            "limite_credito": 0, "saldo": 0, "estado": "activo", "created_at": iso_now()})
            
    # Configuración por defecto
    if not await db.settings.find_one({"_id": "app"}):
        await db.settings.insert_one({
            "_id": "app", "empresa_nombre": "Grupo RYSA", "rfc": "", "telefono": "",
            "correo": "contacto@gruporysa.com", "direccion": "", "ciudad": "", "estado": "",
            "cp": "", "iva_tasa": 8.0, "moneda": "MXN",
            "listas_precios_nombres": ["Precio 1", "Precio 2", "Precio 3", "Precio 4", "Precio 5"],
            "sucursales": [{"nombre": "Matriz", "direccion": "", "ciudad": "", "estado": "",
                            "cp": "", "telefono": "", "activa": True}]})

@app.on_event("shutdown")
async def shutdown():
    client.close()

# Middleware para inyección de Headers de Seguridad Básicos
@app.middleware("http")
async def add_security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    return response

app.include_router(api)

# Módulo de operación en campo: vendedores, visitas, ubicaciones y supervisión.
import field_ops  # noqa: E402
app.include_router(field_ops.router)

# Módulo DESARROLLADOR: diagnóstico, depuración y limpieza transaccional.
# Las rutas destructivas solo se registran si entorno != production Y
# DEVELOPER_MODE=true (ver developer.py).
app.include_router(_devmod.router)

# Configuración dinámica de CORS
env = os.environ.get("ENVIRONMENT", "development").lower()
cors_origins_env = os.environ.get("CORS_ORIGINS", "")
if cors_origins_env:
    origins = [o.strip() for o in cors_origins_env.split(",") if o.strip()]
else:
    if env == "production":
        origins = ["https://gruporysa.com"]
    else:
        # En desarrollo permitimos localhost comunes
        origins = ["http://localhost:3000", "http://localhost:5173", "http://127.0.0.1:3000", "http://127.0.0.1:5173"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["X-Total-Count"],
)
