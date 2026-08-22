"""RYSA - Servicio reutilizable de exportación (Excel + PDF).

Arquitectura compartida para TODOS los módulos (clientes, productos, ventas,
compras, gastos, proveedores, cxc, cxp, inventario, reportes, vendedores).
Cada vista de listado construye una tabla (headers + rows) y la entrega aquí;
la generación de Excel y PDF queda centralizada y consistente.
"""
import io
import os

import pandas as pd  # noqa: F401  (API pública para reuso)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

TERRA = "C1401E"
INK = "1F1F1F"


def excel_bytes(rows, headers, sheet_name="Datos", title=None):
    """Genera un XLSX profesional: encabezados con fondo, fila de título
    opcional, bordes, congelado de panel y anchos de columna ajustados.

    rows: lista de dict; se toma el valor de cada clave de `headers`.
    headers: lista de str con el orden y nombre de las columnas.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Datos")[:31]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=TERRA)
    thin = Side(style="thin", color="C9C9C9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row_start = 1
    if title:
        ws.cell(row=1, column=1, value=title)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
        c = ws.cell(row=1, column=1)
        c.font = Font(bold=True, size=13, color=INK)
        c.alignment = Alignment(horizontal="left", vertical="center")
        row_start = 2

    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=row_start, column=ci, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.freeze_panes = ws.cell(row=row_start + 1, column=1).coordinate

    for r in rows:
        ws.append([r.get(h) for h in headers])
    for row in ws.iter_rows(min_row=row_start + 1, max_col=len(headers)):
        for c in row:
            c.border = border

    for ci in range(1, len(headers) + 1):
        cell = ws.cell(row=row_start, column=ci)
        width = 10
        for row_idx in range(row_start + 1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=ci).value
            if v is not None:
                width = max(width, min(len(str(v)) + 2, 40))
        ws.column_dimensions[cell.column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def pdf_bytes(title, headers, rows, settings=None, user_name="", filtros=None,
              col_weights=None, wrap_cols=None):
    """Genera un PDF apaisado con logo RYSA, título del módulo, fecha y hora
    de generación, usuario, filtros aplicados, total de registros y tabla.

    title:   texto principal (ej. "REPORTE DE CLIENTES").
    headers: lista de encabezados de la tabla.
    rows:    lista de listas (una por registro, columnas = `headers`).
    settings: dict con empresa_nombre, logo_url, etc. (opcional).
    user_name: nombre del usuario que generó el reporte.
    filtros:  dict/lista de (etiqueta, valor) aplicados (opcional).
    col_weights: lista opcional con el peso relativo de cada columna (índice).
                 Las columnas con texto largo (descripción, dirección) pueden
                 recibir más ancho que las cortas (código, cantidad, moneda).
    wrap_cols:  set/lista opcional de índices de columna cuyas celdas se
                envuelven (word-wrap) con Paragraph en lugar de texto plano,
                evitando que las descripciones se amontonen/corten.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, Image, HRFlowable)

    TERRA_C = colors.HexColor("#C1401E")
    INK_C = colors.HexColor("#1F1F1F")
    GRIS = colors.HexColor("#6B7280")
    LINEA = colors.HexColor("#E5E0DA")

    settings = settings or {}
    margins = 15 * mm
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(letter),
        leftMargin=margins, rightMargin=margins, topMargin=margins, bottomMargin=margins,
        title=f"{title} - Grupo RYSA", author="Grupo RYSA")

    st = getSampleStyleSheet()
    tstyle = ParagraphStyle("titulo", parent=st["Normal"], fontName="Helvetica-Bold",
                            fontSize=15, leading=18, textColor=TERRA_C, alignment=TA_LEFT,
                            spaceAfter=2)

    def es(font="Helvetica", size=9, leading=12, color=INK_C, bold=False, align=TA_LEFT):
        return ParagraphStyle("s", fontName=("Helvetica-Bold" if bold else font),
                              fontSize=size, leading=leading, textColor=color, alignment=align)

    sEmpresa = es("Helvetica", 13, 16, color=INK_C, bold=True)
    sRazon = es("Helvetica", 8.5, 11, color=GRIS)
    sMeta = es("Helvetica", 9, 13, color=INK_C)
    sTiny = es("Helvetica", 8, 11, color=GRIS)
    sHead = es("Helvetica", 8, 10, color=colors.white, bold=True)
    sCel = es("Helvetica", 7, 9, color=INK_C)

    story = []

    # --- Encabezado: logo (si aplica) + datos de la empresa ---
    logo_img = _resolve_logo(settings)
    logos_left = [[Image(logo_img, width=34 * mm, height=30 * mm)]] if logo_img else [[Paragraph("", sRazon)]]
    emp_lines = [Paragraph("<b>%s</b>" % (settings.get("empresa_nombre") or "Grupo RYSA"), sEmpresa)]
    for lin in _empresa_lines(settings):
        if lin.strip():
            emp_lines.append(Paragraph(lin, sRazon))
    head = Table([[logos_left, emp_lines]], colWidths=[48 * mm, None])
    head.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    story.append(head)
    story.append(HRFlowable(width="100%", thickness=2, color=TERRA_C, spaceBefore=6, spaceAfter=8))

    # --- Título + fecha/hora + usuario ---
    story.append(Paragraph(title, tstyle))
    from datetime import datetime
    now = datetime.now()
    met = "Generado: {}".format(now.strftime("%d/%m/%Y %H:%M"))
    if user_name:
        met += "  ·  Usuario: {}".format(user_name)
    story.append(Paragraph(met, sMeta))

    # --- Filtros aplicados ---
    flist = []
    if isinstance(filtros, dict):
        flist = ["{}: {}".format(k, v) for k, v in filtros.items() if v not in (None, "", "all")]
    elif isinstance(filtros, list):
        flist = [str(v) for v in filtros if v not in (None, "", "all")]
    filt_txt = "Filtros: " + ("; ".join(flist) if flist else "ninguno (todos los registros)")
    story.append(Paragraph(filt_txt, sTiny))
    story.append(HRFlowable(width="100%", thickness=0.6, color=LINEA, spaceBefore=3, spaceAfter=6))

    # --- Total de registros ---
    story.append(Paragraph("<b>Total:</b> {} registros".format(len(rows)), sMeta))
    story.append(Spacer(1, 5))

    # --- Tabla de resultados ---
    # Las celdas se envuelven con Paragraph solo en las columnas indicadas
    # (wrap_cols) para que las descripciones/direcciones largas hagan word-wrap;
    # el resto usa strings planos para mantener el rendimiento con miles de filas.
    ncol = len(headers) or 1
    wrap_set = set(wrap_cols or [])

    data = [ [Paragraph("<b>%s</b>" % _txt(h), sHead) for h in headers] ]
    for r in rows:
        row = []
        for i, v in enumerate(r):
            if i in wrap_set:
                row.append(Paragraph(_txt(v), sCel))
            else:
                row.append(_plain(v))
        data.append(row)

    usable = _page_width() - 2 * margins
    if col_weights:
        weights = [float(col_weights[i]) if i < len(col_weights) else 1.0 for i in range(ncol)]
        tot_w = sum(weights) or 1.0
        col_w = [(usable - 4 * ncol) * (w / tot_w) for w in weights]
    else:
        col_w = [(usable - 4 * ncol) / ncol for _ in range(ncol)]
    tbl = Table(data, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), TERRA_C),
        ("GRID", (0, 0), (-1, -1), 0.4, LINEA),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FBF7F4")]),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("TEXTCOLOR", (0, 1), (-1, -1), INK_C),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(tbl)

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ---------- helpers internos ----------

def _page_width():
    from reportlab.lib.pagesizes import letter, landscape
    return landscape(letter)[0]


def _txt(v):
    """Convierte un valor a texto seguro para Paragraph (escapando XML)."""
    if v is None:
        return "—"
    s = str(v).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    return s


def _plain(v):
    """Convierte un valor a string plano para celda de tabla (sin layout)."""
    if v is None:
        return ""
    return "" if v == "" else str(v)


def _empresa_lines(settings):
    lines = []
    for f in ["razon_social", "direccion", "ciudad"]:
        v = settings.get(f)
        if v:
            lines.append(str(v))
    extras = []
    if settings.get("telefono"):
        extras.append("Tel: " + str(settings["telefono"]))
    if settings.get("correo"):
        extras.append(str(settings["correo"]))
    if settings.get("rfc"):
        extras.append("RFC: " + str(settings["rfc"]))
    lines.extend(extras)
    return lines


def _resolve_logo(settings):
    """Devuelve la ruta local del logo RYSA (settings.logo_url o brand), o None.

    Se cachea en memoria para no pagar el costo de disco por cada llamada
    (una exportación grande llama esto una vez, pero el proceso vive en prod).
    """
    global _LOGO_CACHE
    lu = (settings or {}).get("logo_url") or ""
    key = lu
    if key in _LOGO_CACHE:
        return _LOGO_CACHE[key]
    found = _resolve_logo_uncached(lu)
    _LOGO_CACHE[key] = found
    if len(_LOGO_CACHE) > 16:
        _LOGO_CACHE.clear()
    return found


_LOGO_CACHE = {}


def _resolve_logo_uncached(lu):
    candidates = []
    rel = ""
    if lu and "/api/files/" in lu:
        rel = lu.split("/api/files/", 1)[1]
        candidates.append(rel)
    candidates.append(os.path.join(os.path.dirname(__file__), "brand", "logotipo.png"))
    env_upload = os.environ.get("UPLOAD_DIR")
    if env_upload and rel:
        candidates.insert(1, os.path.join(env_upload, rel))
    for cand in candidates:
        if cand and os.path.isfile(cand):
            return cand
    try:
        import storage as _storage
        if rel:
            p = _storage.get_safe_local_path(rel)
            if p and os.path.isfile(p):
                return p
        base = _storage.base_upload_dir()
        for f in ("logo.png", "logotipo.png", "isotipo.png", "logo.jpg"):
            if os.path.isfile(os.path.join(base, f)):
                return os.path.join(base, f)
    except Exception:
        pass
    return None